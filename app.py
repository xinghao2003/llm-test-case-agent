# app.py
import os
import io
import json
import time
import shutil
import zipfile
import tempfile
import subprocess
import logging
import platform
import mimetypes
import traceback
import hashlib
from datetime import datetime, timezone
from pathlib import Path
from typing import List, Dict, Any, Tuple, Optional, Set, Union

from dotenv import load_dotenv
import gradio as gr
from gradio import ChatMessage
import yaml
import pathspec
from pathspec import PathSpec

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ---------- Configuration ----------

MAX_TOTAL_TEXT = 250_000
ALLOWED_EXTENSIONS = {".zip", ".txt", ".md", ".pdf", ".jpg", ".jpeg", ".png"}

INITIAL_ASSISTANT_MESSAGE = (
    "Hi! I'm your test generator assistant. Attach your codebase (ZIP), documentation (PDF), or images, "
    "then describe what tests you need. I'll generate test specifications and code for you."
)

SYSTEM_PROMPT = """\
You are a senior QA engineer and software architect generating and improving automated test plans for a Python project.

## Goals
1) Read and understand all provided context (PDFs, images, flattened repo text, and conversation history).
2) Read and understand the user's current request/requirements.
3) Analyze the codebase and create comprehensive test materials:
   - Python pytest test code files under appropriate `tests/` paths.
   - Helper modules under `tests/utils/` if beneficial.
   - Comprehensive Markdown test specifications (e.g., `TEST_SPEC.md` or under `docs/`).
   - Configuration files or auxiliary files that support testing (e.g., `conftest.py`, `pytest.ini`).
   - When necessary for test integration, you may create or edit files in the original codebase (e.g., adding imports, fixtures, or pytest hooks).
4) **PRIMARY FOCUS**: Prioritize creating new test files over editing existing source files. Only edit existing files when absolutely necessary for test framework integration.
5) **IMPORTANT**: Do NOT fix bugs, syntax errors, or any issues in the original codebase that are unrelated to test execution. Focus exclusively on test generation. If you encounter bugs or errors in the codebase, document them in test specifications but do not attempt to fix them.
6) Tests must be runnable with `pytest` out of the box (assume `pytest` only; no plugins unless absolutely needed).
7) Prefer pure-std-lib where possible. If external deps are strictly necessary, note them in the spec file.
8) Ensure meaningful file paths (POSIX-style), consistent naming, and idempotent generation.
9) Include minimal fixtures, parametrization, and positive/negative cases where it helps coverage.

## Function Calling
Use the `create_or_edit_file` function to create new files ONLY. You **cannot** edit existing files from the original codebase.

When you need to:
- **Create a new file**: Call `create_or_edit_file` with the full file path and complete content. 
  - Allowed paths: `tests/**`, `docs/**`, `test_config/**`, `test_utils/**`, or similar new testing directories.
  - **Do NOT** create or edit files in the main application source directories.
- **Edit for testing support**: Only create new configuration or helper files; never modify application source code.
- **Create multiple files**: Call `create_or_edit_file` multiple times in a single response for parallel execution.

## Content Requirements
- For `markdown` files, produce full content (headings, scenarios, cases, traceability).
- For Python test files:
  - Use pytest (`test_*.py`) naming.
  - Use clear test functions, possibly fixtures.
  - Add docstrings explaining intent.
  - Where app APIs are unclear, stub interfaces using reasonable assumptions, with TODOs clearly marked.
  - Avoid environment-dependent operations unless explicitly described in context.
- For new files:
  - Provide the COMPLETE file content.
  - Maintain consistency with the codebase style.
  - Do not reference or depend on undocumented internal app APIs.

### TEST_SPEC.md Template Guidance
When generating or editing `TEST_SPEC.md` (or similar Markdown specs), use the structure below. If specific details (for example, execution evidence or assigned tester) are unknown, populate the field with `TBD` rather than fabricating data.

# **Test Case Specification Template**

**Module Name:**
**Test Case Spec ID:**

---

### **Test Case Description**

> *Describe the purpose of this test case.*

---

### **Prerequisites**

1.
2.

---

### **Environmental Information**

| Item                   | Details          |
| ---------------------- | ---------------- |
| **Operating System**   | TBD              |
| **System Type**        | Laptop / Desktop |
| **Browser / Platform** | TBD              |

---

### **Test Scenario**

> *Describe what is being tested and why.*

---

## **Test Case Details**

| **Test Case ID** | **Test Steps**   | **Test Input** | **Expected Results** | **Verification Notes** | **Comments** |
| ---------------- | ---------------- | -------------- | -------------------- | ---------------------- | ------------ |
| 1                | 1. <br>2. <br>3. | TBD            | TBD                  | Not executed yet       | TBD          |

---

### **Notes**

* Mark unknown or future-state information as `TBD`.
* Attach screenshots, logs, or evidence externally when available.
* Keep *Expected Results* clear and measurable.

---

## **Example**

**Module Name:** Login
**Test Case ID:** gfg_01

---

### **Test Case Description**

To verify login functionality of the GeeksforGeeks website.

---

### **Prerequisites**

1. Stable internet connection
2. Browser (Chrome, Firefox, Internet Explorer, etc.)

---

### **Environmental Information**

| Item                   | Details                              |
| ---------------------- | ------------------------------------ |
| **Operating System**   | Windows / Linux / Mac                |
| **System Type**        | Laptop / Desktop                     |
| **Browser / Platform** | Chrome / Firefox / Internet Explorer |

---

### **Test Scenario**

Checking that after entering the correct username and password, the user can successfully log in.

---

## **Test Case Details**

| **Test Case ID** | **Test Steps**                                              | **Test Input**                                   | **Expected Results**                        | **Verification Notes** | **Comments**    |
| ---------------- | ----------------------------------------------------------- | ------------------------------------------------ | ------------------------------------------- | ---------------------- | --------------- |
| 1                | 1. Enter Username<br>2. Enter Password<br>3. Click on Login | Username: `geeksforgeeks`<br>Password: `geek123` | "Welcome to GeeksforGeeks!" message appears | Execution pending      | No issues found |

## Summary Output Format
After generating test files, provide a structured summary that includes:

1. **What Was Done**
   - List modules/features tested.
   - Name the files created (e.g., `tests/test_auth.py`, `TEST_SPEC.md`).
   - Describe the type of tests added (unit, integration, e2e, etc.).

2. **Why It Was Done**
   - Explain the business/technical rationale for each test suite.
   - Link tests to specific requirements or risks identified in the codebase.
   - Highlight critical paths or high-priority features covered.

3. **Test Coverage**
   - Summarize which functions, classes, or workflows are tested.
   - List the number of test cases per module.
   - Identify positive (happy path) vs. negative (error) test scenarios included.

4. **Gaps & Limitations**
   - Document what is NOT tested and why (e.g., external API mocks, environment-specific behaviors).
   - Note any assumptions made or areas requiring manual testing.
   - Mention known limitations of the pytest suite (e.g., no performance testing, no UI testing).

5. **Usage Instructions**
   - How to run the full test suite: `pytest tests/`
   - How to run specific test files or modules: `pytest tests/test_auth.py -v`
   - Any required setup (environment variables, test data, fixtures).
   - How to generate a coverage report: `pytest --cov=<module> tests/`

## Process
1) Review conversation history to understand what has been generated previously.
2) Summarize the repo: key modules, public APIs (if detectable), and how they relate to the current request.
3) For initial requests: Generate comprehensive test scenarios and cases in NEW files only.
4) For improvement requests: Only create additional test files; never modify original application code.
5) Use function calls to create new files.
6) Provide a structured summary following the **Summary Output Format** above instead of embedding code.
7) Do NOT include code snippets or detailed source code in your summary response.

## Constraints
- Call the `create_or_edit_file` function to persist **new test and documentation files only**.
- Always provide the complete file content to the function (not diffs or patches).
- Use parallel function calls when creating multiple independent files.
- Reference previous iterations when making improvements.
- **NEVER edit application source code files. Only create new test files.**
- **In your final summary message, focus on explaining WHAT, WHY, COVERAGE, GAPS, and USAGE—not the code itself.**
"""



def normalize_uploaded_entry(entry: Any) -> Tuple[Optional[Path], Optional[str]]:
    """Extract a filesystem path and display name from a Gradio upload payload."""
    candidate_path: Optional[Path] = None
    display_name: Optional[str] = None

    if isinstance(entry, dict):
        raw_path = entry.get("path") or entry.get("name") or entry.get("file")
        if raw_path:
            candidate_path = Path(raw_path)
        display_name = entry.get("orig_name") or entry.get("name")
    elif isinstance(entry, Path):
        candidate_path = entry
    elif isinstance(entry, str):
        candidate_path = Path(entry)
    elif hasattr(entry, "name"):
        candidate_path = Path(getattr(entry, "name"))
    elif hasattr(entry, "path"):
        candidate_path = Path(getattr(entry, "path"))

    if not display_name and candidate_path is not None:
        display_name = candidate_path.name

    return candidate_path, display_name

# ---------- Logger Setup ----------

def setup_logger() -> logging.Logger:
    log_dir = Path("logs")
    log_dir.mkdir(exist_ok=True)
    logger = logging.getLogger("test_generator")
    logger.setLevel(logging.DEBUG)

    log_file = log_dir / f"app_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    formatter = logging.Formatter(
        '%(asctime)s - %(levelname)s - [%(funcName)s] - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    fh = logging.FileHandler(log_file, encoding='utf-8')
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(formatter)

    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)
    ch.setFormatter(formatter)

    if not logger.handlers:
        logger.addHandler(fh)
        logger.addHandler(ch)
    return logger

logger = setup_logger()

DEFAULT_TEMPERATURE = 0.4
DEFAULT_MODEL_NAME = "gemini-flash-latest"


def get_model_temperature() -> float:
    raw_value = os.getenv("GEMINI_TEMPERATURE")
    if raw_value is None:
        return DEFAULT_TEMPERATURE
    try:
        return float(raw_value)
    except ValueError:
        logger.warning(
            "Invalid GEMINI_TEMPERATURE=%s; falling back to %.2f",
            raw_value,
            DEFAULT_TEMPERATURE,
        )
        return DEFAULT_TEMPERATURE


def get_model_name() -> str:
    return os.getenv("GEMINI_MODEL", DEFAULT_MODEL_NAME)

# ---------- Gemini Client ----------

try:
    from google import genai
    from google.genai import types
except ImportError:
    raise SystemExit(
        "google-genai not installed. Run: pip install google-genai\n"
        "Docs: https://pypi.org/project/google-genai/"
    )

def create_gemini_client() -> genai.Client:
    load_dotenv()
    api_key = os.getenv("GOOGLE_API_KEY")
    if not api_key:
        raise RuntimeError("GOOGLE_API_KEY environment variable is not set.")
    return genai.Client(api_key=api_key)

# ---------- Tool: create_or_edit_file ----------

def get_file_creation_tool_definition() -> Dict[str, Any]:
    return {
        "name": "create_or_edit_file",
        "description": "Create a new file or edit an existing file by replacing its entire content.",
        "parameters": {
            "type": "object",
            "properties": {
                "file_path": {"type": "string", "description": "POSIX path to create/edit (under temp work dir)"},
                "content": {"type": "string", "description": "Complete file content"},
                "description": {"type": "string", "description": "Short purpose/notes"}
            },
            "required": ["file_path", "content", "description"]
        }
    }

def execute_file_creation(file_path: str, content: str, description: str, work_dir: Path, original_codebase_files: Optional[Set[str]] = None) -> Dict[str, Any]:
    try:
        normalized_path = file_path.replace("\\", "/").lstrip("/")
        target_path = work_dir / normalized_path
        
        # Log if editing a file from original codebase (advisory only, not blocked)
        if original_codebase_files:
            # Normalize the target path for comparison
            path_parts = normalized_path.lower()
            
            # Check if file exists in original codebase
            for orig_file in original_codebase_files:
                if orig_file.lower() == path_parts or path_parts.endswith('/' + orig_file.lower()):
                    logger.warning(
                        f"Modifying original codebase file: {file_path}\n"
                        f"Description: {description}"
                    )
                    break
        
        target_path.parent.mkdir(parents=True, exist_ok=True)
        target_path.write_text(content, encoding='utf-8')
        logger.info(f"File created/edited: {target_path}")
        return {
            "status": "success",
            "file_path": str(target_path),
            "message": f"File '{file_path}' created/edited ({len(content)} bytes)",
            "description": description
        }
    except Exception as e:
        logger.error(f"Error creating file {file_path}: {e}", exc_info=True)
        return {"status": "error", "file_path": file_path, "message": str(e)}

# ---------- Repo-to-Text (single, de-duplicated implementation) ----------

def check_tree_command() -> bool:
    return shutil.which('tree') is not None

def is_ignored_path(file_path: str) -> bool:
    return ('.git' in file_path) or os.path.basename(file_path).startswith('repo-to-text_')

def run_tree_command(path: str) -> str:
    if platform.system() == "Windows":
        cmd = ["cmd", "/c", "tree", "/a", "/f", path]
    else:
        cmd = ["tree", "-a", "-f", "--noreport", path]
    result = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, encoding='utf-8', check=True)
    return result.stdout

def extract_full_path(line: str, path: str) -> Optional[str]:
    idx = line.find('./')
    if idx == -1:
        idx = line.find(path)
    return line[idx:].strip() if idx != -1 else None

def _mark_non_empty_dirs(relative_path: str, non_empty_dirs: Set[str]) -> None:
    dir_path = os.path.dirname(relative_path)
    while dir_path:
        non_empty_dirs.add(dir_path)
        dir_path = os.path.dirname(dir_path)

def _should_ignore_file(
    file_path: str,
    relative_path: str,
    gitignore_spec: Optional[PathSpec],
    content_ignore_spec: Optional[PathSpec],
    tree_and_content_ignore_spec: Optional[PathSpec]
) -> bool:
    relative_path = relative_path.replace(os.sep, '/')
    if relative_path.startswith('./'):
        relative_path = relative_path[2:]
    if os.path.isdir(file_path):
        relative_path += '/'

    return (
        is_ignored_path(file_path) or
        bool(gitignore_spec and gitignore_spec.match_file(relative_path)) or
        bool(content_ignore_spec and content_ignore_spec.match_file(relative_path)) or
        bool(tree_and_content_ignore_spec and tree_and_content_ignore_spec.match_file(relative_path))
    )

def _process_line(
    line: str,
    path: str,
    gitignore_spec: Optional[PathSpec],
    tree_and_content_ignore_spec: Optional[PathSpec],
    non_empty_dirs: Set[str]
) -> Optional[str]:
    full_path = extract_full_path(line, path)
    if not full_path or full_path == '.':
        return None

    try:
        relative_path = os.path.relpath(full_path, path).replace(os.sep, '/')
    except (ValueError, OSError):
        relative_path = os.path.basename(full_path)

    if _should_ignore_file(full_path, relative_path, gitignore_spec, None, tree_and_content_ignore_spec):
        return None

    if not os.path.isdir(full_path):
        _mark_non_empty_dirs(relative_path, non_empty_dirs)

    if not os.path.isdir(full_path) or os.path.dirname(relative_path) in non_empty_dirs:
        return line.replace('./', '', 1)
    return None

def _filter_tree_output(
    tree_output: str,
    path: str,
    gitignore_spec: Optional[PathSpec],
    tree_and_content_ignore_spec: Optional[PathSpec]
) -> str:
    lines = tree_output.splitlines()
    non_empty_dirs: Set[str] = set()
    filtered = [_process_line(line, path, gitignore_spec, tree_and_content_ignore_spec, non_empty_dirs) for line in lines]
    return '\n'.join(filter(None, filtered))

def get_tree_structure(
    path: str = '.',
    gitignore_spec: Optional[PathSpec] = None,
    tree_and_content_ignore_spec: Optional[PathSpec] = None
) -> str:
    if not check_tree_command():
        return ""
    try:
        tree_output = run_tree_command(path)
    except Exception:
        return ""
    if not gitignore_spec and not tree_and_content_ignore_spec:
        return tree_output
    return _filter_tree_output(tree_output, path, gitignore_spec, tree_and_content_ignore_spec)

def load_ignore_specs(path: str = '.', cli_ignore_patterns: Optional[List[str]] = None) -> Tuple[Optional[PathSpec], Optional[PathSpec], PathSpec]:
    gitignore_spec = None
    content_ignore_spec = None
    tree_and_content_ignore_list: List[str] = []
    use_gitignore = True

    repo_settings_path = os.path.join(path, '.repo-to-text-settings.yaml')
    if os.path.exists(repo_settings_path):
        try:
            with open(repo_settings_path, 'r', encoding='utf-8') as f:
                settings: Dict[str, Any] = yaml.safe_load(f) or {}
                use_gitignore = settings.get('gitignore-import-and-ignore', True)
                if 'ignore-content' in settings:
                    content_ignore_spec = pathspec.PathSpec.from_lines('gitwildmatch', settings['ignore-content'])
                if 'ignore-tree-and-content' in settings:
                    tree_and_content_ignore_list.extend(settings.get('ignore-tree-and-content', []))
        except Exception:
            pass

    if cli_ignore_patterns:
        tree_and_content_ignore_list.extend(cli_ignore_patterns)

    if use_gitignore:
        gitignore_path = os.path.join(path, '.gitignore')
        if os.path.exists(gitignore_path):
            try:
                with open(gitignore_path, 'r', encoding='utf-8') as f:
                    gitignore_spec = pathspec.PathSpec.from_lines('gitwildmatch', f)
            except Exception:
                pass

    tree_and_content_ignore_spec = pathspec.PathSpec.from_lines('gitwildmatch', tree_and_content_ignore_list)
    return gitignore_spec, content_ignore_spec, tree_and_content_ignore_spec

def _read_file_content(file_path: str) -> str:
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return f.read()
    except UnicodeDecodeError:
        with open(file_path, 'rb') as f_bin:
            return f_bin.read().decode('latin1')
    except FileNotFoundError:
        if os.path.islink(file_path) and not os.path.exists(file_path):
            try:
                target = os.readlink(file_path)
            except OSError:
                target = ''
            return f"[symlink] -> {target}"
        raise

def _load_additional_specs(path: str = '.') -> Dict[str, Any]:
    kv: Dict[str, Any] = {'maximum_word_count_per_file': None}
    repo_settings_path = os.path.join(path, '.repo-to-text-settings.yaml')
    if os.path.exists(repo_settings_path):
        try:
            with open(repo_settings_path, 'r', encoding='utf-8') as f:
                settings = yaml.safe_load(f) or {}
                max_words = settings.get('maximum_word_count_per_file')
                if isinstance(max_words, int) and max_words > 0:
                    kv['maximum_word_count_per_file'] = max_words
        except Exception:
            pass
    return kv

def _generate_output_content(
    path: str,
    tree_structure: str,
    gitignore_spec: Optional[PathSpec],
    content_ignore_spec: Optional[PathSpec],
    tree_and_content_ignore_spec: Optional[PathSpec],
    maximum_word_count_per_file: Optional[int] = None
) -> List[str]:
    output_segments: List[str] = []
    current: List[str] = []
    current_wc = 0
    project_name = os.path.basename(os.path.abspath(path))

    def count_words(text: str) -> int:
        return len(text.split())

    def finalize():
        nonlocal current_wc
        if current:
            output_segments.append("".join(current))
            current.clear()
            current_wc = 0

    def add(chunk: str):
        nonlocal current_wc
        wc = count_words(chunk)
        if maximum_word_count_per_file and current and current_wc + wc > maximum_word_count_per_file:
            finalize()
        current.append(chunk)
        current_wc += wc

    add('<repo-to-text>\n')
    add(f'Directory: {project_name}\n\n')
    add('Directory Structure:\n')
    add('<directory_structure>\n.\n')
    if os.path.exists(os.path.join(path, '.gitignore')):
        add('├── .gitignore\n')
    add(tree_structure + '\n' + '</directory_structure>\n')

    for root, _, files in os.walk(path):
        for filename in files:
            file_path = os.path.join(root, filename)
            relative_path = os.path.relpath(file_path, path)
            if _should_ignore_file(file_path, relative_path, gitignore_spec, content_ignore_spec, tree_and_content_ignore_spec):
                continue
            cleaned_relative_path = relative_path.replace('./', '', 1)
            add(f'\n<content full_path="{cleaned_relative_path}">\n')
            add(_read_file_content(file_path))
            add('\n</content>\n')

    add('\n</repo-to-text>\n')
    finalize()
    return output_segments or ["<repo-to-text>\n</repo-to-text>\n"]

def save_repo_to_text(
        path: str = '.',
        output_dir: Optional[str] = None,
        to_stdout: bool = False,
        cli_ignore_patterns: Optional[List[str]] = None
) -> str:
    gitignore_spec, content_ignore_spec, tree_and_content_ignore_spec = load_ignore_specs(path, cli_ignore_patterns)
    additional_specs = _load_additional_specs(path)
    tree_structure = get_tree_structure(path, gitignore_spec, tree_and_content_ignore_spec)
    segments = _generate_output_content(
        path,
        tree_structure,
        gitignore_spec,
        content_ignore_spec,
        tree_and_content_ignore_spec,
        additional_specs.get('maximum_word_count_per_file')
    )

    if to_stdout:
        for s in segments:
            print(s, end='')
        return "".join(segments)

    timestamp = datetime.now(timezone.utc).strftime('%Y-%m-%d-%H-%M-%S-UTC')
    base_stem = f'repo-to-text_{timestamp}'
    output_paths: List[str] = []

    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)

    if len(segments) == 1:
        filename = f"{base_stem}.txt"
        full = os.path.join(output_dir, filename) if output_dir else filename
        Path(full).write_text(segments[0], encoding='utf-8')
        print(f'[SUCCESS] Repository structure and contents saved to "{os.path.basename(full)}"')
        output_paths.append(full)
    else:
        for i, segment in enumerate(segments, 1):
            filename = f"{base_stem}_part_{i}.txt"
            full = os.path.join(output_dir, filename) if output_dir else filename
            Path(full).write_text(segment, encoding='utf-8')
            output_paths.append(full)
        print(f"[SUCCESS] Saved to {len(output_paths)} files:")
        for fp in output_paths:
            print(f'  - "{os.path.basename(fp)}"')

    return output_paths[0] if output_paths else ""

# ---------- Helpers: file ingest & Google Files ----------

def ensure_dir(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path

def copy_file_to_work_dir(src: Path, work_dir: Path) -> Path:
    dst = work_dir / (src.stem + ".txt" if src.suffix.lower() == ".md" else src.name)
    shutil.copy2(src, dst)
    return dst

def categorize_upload(file_path: Path) -> Optional[str]:
    ext = file_path.suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        logger.warning(f"Unsupported file extension: {file_path.name} ({ext})")
        return None
    return "zip" if ext == ".zip" else "context"

def get_mime_type(file_path: Path) -> str:
    mime_type, _ = mimetypes.guess_type(str(file_path))
    if mime_type:
        return mime_type
    return {
        '.md': 'text/markdown',
        '.txt': 'text/plain',
        '.pdf': 'application/pdf',
        '.jpg': 'image/jpeg',
        '.jpeg': 'image/jpeg',
        '.png': 'image/png',
        '.gif': 'image/gif',
    }.get(file_path.suffix.lower(), 'application/octet-stream')

def upload_file_to_gemini(client: genai.Client, file_path: Path) -> Dict[str, str]:
    return upload_path_to_gemini(client, file_path, None)

def upload_path_to_gemini(client: genai.Client, path: Path, mime_hint: Optional[str] = None) -> Optional[Dict[str, str]]:
    mime_type = mime_hint or get_mime_type(path)
    try:
        config = types.UploadFileConfig(mime_type=mime_type)
    except (ImportError, AttributeError):
        config = None

    uploaded_file = None
    last_error: Optional[Exception] = None

    attempt_id = 0

    def _log_failure(exc: Exception, label: str) -> None:
        logger.warning(f"Upload attempt {label} failed for {path}: {exc}")

    attempts: List[Tuple[str, Any]] = []

    if config:
        attempts.append(("file+config", lambda: client.files.upload(file=str(path), config=config)))

    attempts.append(("file+mime", lambda: client.files.upload(file=str(path), mime_type=mime_type)))

    def attempt_file_handle():
        with open(path, 'rb') as fh:
            return client.files.upload(file=fh, mime_type=mime_type)

    attempts.append(("fh+mime", attempt_file_handle))

    if hasattr(client.files, "upload_from_path"):
        attempts.append(("upload_from_path", lambda: client.files.upload_from_path(path=str(path), mime_type=mime_type)))

    for label, func in attempts:
        attempt_id += 1
        try:
            uploaded_file = func()
            if uploaded_file:
                break
        except KeyError as exc:
            last_error = exc
            _log_failure(exc, label)
        except Exception as exc:  # pylint: disable=broad-except
            last_error = exc
            _log_failure(exc, label)

    if not uploaded_file:
        logger.error(f"Giving up uploading {path}: {last_error}")
        return None

    logger.info(f"Uploaded to Google Files: {path.name}")
    file_uri = getattr(uploaded_file, "uri", None)
    if not file_uri:
        logger.warning(f"Uploaded file object missing URI for {path}; response: {uploaded_file}")
        return None
    return {"file_uri": file_uri, "mime_type": mime_type, "file_path": str(path)}


def prepare_file_for_upload(source: Path) -> Tuple[Path, Optional[Path]]:
    """Return path to upload and optional temp artifact to clean up."""
    try:
        if source.suffix.lower() == ".md":
            with tempfile.NamedTemporaryFile(
                mode='w',
                encoding='utf-8',
                suffix='.txt',
                prefix=source.stem + '_',
                dir=str(source.parent),
                delete=False
            ) as tmp_fh:
                tmp_fh.write(source.read_text(encoding='utf-8'))
                temp_path = Path(tmp_fh.name)
            logger.info(f"Converted markdown to text for upload: {source.name} -> {temp_path.name}")
            return temp_path, temp_path
    except Exception as exc:  # pylint: disable=broad-except
        logger.warning(f"Failed converting markdown {source} to text: {exc}; uploading original file.")
    return source, None

def extract_and_flatten_repo(zip_path: Path) -> Tuple[str, str]:
    temp_dir = Path(tempfile.mkdtemp(prefix="codebase_"))
    with zipfile.ZipFile(zip_path, 'r') as zf:
        zf.extractall(temp_dir)
    context_dir = ensure_dir(Path("context"))
    output_file_path = save_repo_to_text(path=str(temp_dir), output_dir=str(context_dir), to_stdout=False)
    output_file = Path(output_file_path)
    content = output_file.read_text(encoding="utf-8", errors="replace")
    if len(content) > MAX_TOTAL_TEXT:
        content = content[:MAX_TOTAL_TEXT]
        logger.warning(f"Truncated {zip_path.stem} from {len(content)} to {MAX_TOTAL_TEXT} chars")
    return output_file.as_posix(), content

def save_json_file(data: Dict[str, Any], directory: str, prefix: str) -> Path:
    dir_path = ensure_dir(Path(directory))
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    file_path = dir_path / f"{prefix}_{timestamp}.json"
    file_path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding='utf-8')
    logger.info(f"Saved {prefix} to: {file_path.name} ({file_path.stat().st_size} bytes)")
    return file_path

def save_llm_context(user_story: str, repo_files: List[Tuple[str, str]], uploaded_files: List[Path]) -> Path:
    context_data = {
        "timestamp": datetime.now().isoformat(),
        "user_story": user_story,
        "code_files": [{"path": path, "size_chars": len(content)} for path, content in repo_files],
        "uploaded_files": [{"name": f.name, "size_bytes": f.stat().st_size} for f in uploaded_files],
        "file_content_sample": {path: content[:500] for path, content in repo_files[:5]}
    }
    return save_json_file(context_data, "context_logs", "llm_context")

def convert_markdown_to_pdf(md_file_path: Path, margin_top: str = "1in", margin_bottom: str = "1in", 
                            margin_left: str = "1in", margin_right: str = "1in") -> Optional[Path]:
    """
    Convert a Markdown file to PDF using markdown-pdf library.
    
    Args:
        md_file_path: Path to the markdown file
        margin_top: Top margin (e.g., "1in", "2cm")
        margin_bottom: Bottom margin (e.g., "1in", "2cm")
        margin_left: Left margin (e.g., "1in", "2cm")
        margin_right: Right margin (e.g., "1in", "2cm")
    
    Returns:
        The PDF file path if successful, None otherwise (markdown will be used as fallback).
    """
    pdf_file_path = md_file_path.with_suffix('.pdf')
    
    try:
        from markdown_pdf import MarkdownPdf, Section
        md_content = md_file_path.read_text(encoding='utf-8')
        pdf = MarkdownPdf(toc_level=0, optimize=True)
        pdf.add_section(Section(md_content))
        pdf.meta["title"] = md_file_path.stem
        pdf.save(str(pdf_file_path))
        logger.info(f"Converted {md_file_path} to {pdf_file_path} using markdown_pdf library")
        return pdf_file_path
    except ImportError:
        logger.debug("markdown_pdf library not installed, skipping PDF conversion")
    except Exception as e:
        logger.warning(f"Error converting {md_file_path} to PDF with markdown_pdf: {e}")
    
    # Fall back to markdown only
    logger.info(f"Could not convert {md_file_path} to PDF. Keeping markdown file only.")
    return None


def _write_debug_log(debug_dir: Path, md_filename: str, stage: str, data: Dict[str, Any]) -> None:
    """
    Write debug information to md2excel folder for troubleshooting.
    """
    try:
        debug_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')
        log_name = md_filename.replace('.md', '')
        log_file = debug_dir / f"{log_name}_{stage}_{timestamp}.json"
        
        with open(log_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, default=str)
    except Exception as e:
        logger.error(f"Failed to write debug log: {e}")

def process_markdown_files_to_excel(
    client: genai.Client,
    work_dir: Path,
    status_updates: Optional[List[Dict[str, Any]]] = None
) -> Dict[str, Dict[str, Any]]:
    """
    Scan work_dir for markdown files and attempt to convert them to Excel.
    Returns a mapping of original md path -> conversion result.
    This runs an independent LLM analyzer for each markdown file.
    Files are converted in-place (same directory as original markdown).
    """
    results = {}
    md_files = list(work_dir.glob("**/*.md"))
    
    if not md_files:
        return results
    
    logger.info(f"Found {len(md_files)} markdown files to analyze for Excel conversion")
    logger.info(f"Work dir for markdown processing: {work_dir}")
    
    debug_dir = ensure_dir(work_dir / "md2excel")
    
    for md_file in md_files:
        try:
            logger.info(f"Processing markdown: {md_file.name}")
            event_id: Optional[str] = None
            if status_updates is not None:
                event_id = f"md2excel_{hashlib.sha1(str(md_file).encode('utf-8')).hexdigest()[:8]}"
                status_updates.append({
                    "event": "start",
                    "event_id": event_id,
                    "file": str(md_file),
                    "display_name": md_file.name,
                    "message": f"Creating Excel file from {md_file.name}"
                })
            _write_debug_log(debug_dir, md_file.name, "input", {
                "file": md_file.name,
                "file_size_bytes": md_file.stat().st_size,
                "file_path": str(md_file),
                "timestamp": datetime.now(timezone.utc).isoformat()
            })
            
            # Output directory is same as markdown directory (in-place conversion)
            output_dir = md_file.parent
            
            conversion_result = call_md_to_excel_analyzer(client, md_file, output_dir, debug_dir)
            
            if conversion_result["status"] == "success":
                results[str(md_file)] = conversion_result
                logger.info(f"Successfully converted {md_file.name} to {conversion_result['format']}")
                _write_debug_log(debug_dir, md_file.name, "success", conversion_result)
                if status_updates is not None:
                    fmt = (conversion_result.get("format") or "").lower()
                    file_path_value = conversion_result.get("file_path")
                    if fmt == "excel" and file_path_value:
                        message = f"Created Excel file {Path(file_path_value).name}"
                    elif fmt == "pdf" and file_path_value:
                        message = f"Created PDF {Path(file_path_value).name}"
                    elif fmt == "markdown" and file_path_value:
                        message = f"Kept markdown file {Path(file_path_value).name}"
                    else:
                        message = "Completed conversion"
                    status_updates.append({
                        "event": "success",
                        "event_id": event_id,
                        "file": str(md_file),
                        "display_name": md_file.name,
                        "message": message,
                        "details": conversion_result
                    })
            else:
                logger.warning(f"Failed to convert {md_file.name}: {conversion_result.get('message', 'Unknown error')}")
                results[str(md_file)] = conversion_result
                _write_debug_log(debug_dir, md_file.name, "failure", conversion_result)
                if status_updates is not None:
                    status_updates.append({
                        "event": "error",
                        "event_id": event_id,
                        "file": str(md_file),
                        "display_name": md_file.name,
                        "message": conversion_result.get('message', 'Unknown error'),
                        "details": conversion_result
                    })
                
        except Exception as e:
            logger.error(f"Error processing {md_file.name}: {e}")
            error_info = {
                "status": "error",
                "message": str(e),
                "exception_type": type(e).__name__,
                "traceback": traceback.format_exc()
            }
            results[str(md_file)] = error_info
            _write_debug_log(debug_dir, md_file.name, "exception", error_info)
            if status_updates is not None:
                status_updates.append({
                    "event": "error",
                    "event_id": event_id,
                    "file": str(md_file),
                    "display_name": md_file.name,
                    "message": str(e),
                    "details": error_info
                })
    
    return results


def write_outputs_to_zip_from_workdir(
    result: Dict[str, Any],
    work_dir: Path,
    original_codebase_dir: Optional[Path] = None,
    client: Optional[genai.Client] = None,
    status_updates: Optional[List[Dict[str, Any]]] = None
) -> Path:
    output_dir = ensure_dir(Path("results"))
    timestamp = datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')
    zip_filename = output_dir / f"generated_files_{timestamp}.zip"

    original_dir = Path(original_codebase_dir) if original_codebase_dir and Path(original_codebase_dir).exists() else None
    original_root_dir = None
    root_folder_name = "generated_output"

    if original_dir:
        try:
            candidates = [p for p in original_dir.iterdir() if not p.name.startswith('__MACOSX')]
        except FileNotFoundError:
            candidates = []

        if len(candidates) == 1 and candidates[0].is_dir():
            original_root_dir = candidates[0]
            root_folder_name = candidates[0].name
        else:
            original_root_dir = original_dir
            root_folder_name = original_dir.name

    created_files: Dict[Path, Path] = {}
    for entry in result.get("created_files", []):
        file_path = Path(entry.get("file_path", ""))
        if not file_path.exists():
            continue
        try:
            rel_path = file_path.relative_to(work_dir)
        except ValueError:
            logger.debug(f"Skipping generated file outside work dir: {file_path}")
            continue
        created_files[rel_path] = file_path

    # Process markdown files: attempt Excel conversion, fallback to PDF or keep as-is
    # Keep original markdown AND add converted files alongside
    converted_files_to_add: List[Tuple[Path, Path]] = []  # (rel_path, abs_path) pairs to add to ZIP
    
    if client:
        try:
            logger.info("Starting markdown-to-Excel analysis...")
            conversion_results = process_markdown_files_to_excel(client, work_dir, status_updates=status_updates)
            
            for md_path_str, conv_result in conversion_results.items():
                original_md = Path(md_path_str)
                
                # Add converted file (Excel or PDF) to the ZIP
                output_file = conv_result.get("file_path")
                if output_file and Path(output_file).exists():
                    try:
                        converted_path = Path(output_file)
                        rel_converted = converted_path.relative_to(work_dir)
                        converted_files_to_add.append((rel_converted, converted_path))
                        
                        fmt = conv_result.get('format', 'unknown')
                        logger.info(f"Will add {converted_path.name} ({fmt}) alongside {original_md.name}")
                    except ValueError:
                        logger.debug(f"Converted file outside work dir: {output_file}")
                else:
                    logger.debug(f"No converted file for {original_md.name}")
                    
        except Exception as e:
            logger.error(f"Markdown-to-Excel processing failed: {e}")
    
    # Add all converted files to created_files (alongside originals, not replacing)
    for rel_path, abs_path in converted_files_to_add:
        created_files[rel_path] = abs_path
        logger.info(f"Added converted file to ZIP: {rel_path}")


    try:
        with zipfile.ZipFile(zip_filename, 'w', zipfile.ZIP_DEFLATED) as zf:
            if original_root_dir:
                for path in original_root_dir.rglob('*'):
                    if not path.is_file():
                        continue
                    rel = path.relative_to(original_root_dir)
                    # Skip original file if a generated file overrides the same relative path
                    if rel in created_files:
                        continue
                    arcname = Path(root_folder_name) / rel
                    zf.write(path, arcname)

            for rel_path, abs_path in created_files.items():
                arcname = Path(root_folder_name) / rel_path
                zf.write(abs_path, arcname)

        logger.info(f"ZIP file created: {zip_filename}")
        return zip_filename
    except Exception as e:
        logger.error(f"Failed to create ZIP: {e}")
        raise

def collect_and_upload_created_files(client: genai.Client, work_dir: Path, since_ts: float) -> List[Dict[str, str]]:
    """Upload files created/edited since `since_ts` under work_dir to Google Files."""
    uploaded: List[Dict[str, str]] = []
    for root, _, files in os.walk(work_dir):
        for fn in files:
            fp = Path(root) / fn
            try:
                if fp.stat().st_mtime >= since_ts:
                    upload_target, cleanup_path = prepare_file_for_upload(fp)
                    upload_info = upload_path_to_gemini(client, upload_target)
                    if upload_info:
                        uploaded.append(upload_info)
                    if cleanup_path and cleanup_path != fp:
                        try:
                            cleanup_path.unlink(missing_ok=True)
                        except Exception as exc:  # pylint: disable=broad-except
                            logger.warning(f"Failed to clean temporary upload artifact {cleanup_path}: {exc}")
            except FileNotFoundError:
                continue
            except Exception as exc:  # pylint: disable=broad-except
                logger.error(f"Failed to upload generated file {fp}: {exc}")
    return uploaded

# ---------- Excel Test Case Helpers (Internal, not exposed to LLM yet) ----------

# Core styles (reusable)
_EXCEL_GREEN = PatternFill("solid", fgColor="FF00B050") if OPENPYXL_AVAILABLE else None
_EXCEL_LIGHT_GRAY = PatternFill("solid", fgColor="FFD3D3D3") if OPENPYXL_AVAILABLE else None
_EXCEL_WHITE_BOLD = Font(color="FFFFFFFF", bold=True) if OPENPYXL_AVAILABLE else None
_EXCEL_WRAP = Alignment(wrap_text=True, vertical="center") if OPENPYXL_AVAILABLE else None
_EXCEL_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True) if OPENPYXL_AVAILABLE else None
_EXCEL_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True) if OPENPYXL_AVAILABLE else None
_EXCEL_THIN = Side(style="thin", color="FF000000") if OPENPYXL_AVAILABLE else None
_EXCEL_BORDER_ALL = Border(left=_EXCEL_THIN, right=_EXCEL_THIN, top=_EXCEL_THIN, bottom=_EXCEL_THIN) if OPENPYXL_AVAILABLE else None
_EXCEL_HEADERS = ["Test Case ID", "Test Steps", "Test Input", "Expected Results", "Actual Results", "Status", "Comments"]

def _apply_table_header(ws, start_row: int) -> None:
    """Apply header styling to a row."""
    if not OPENPYXL_AVAILABLE:
        return
    for i, h in enumerate(_EXCEL_HEADERS, start=1):
        cell = ws.cell(row=start_row, column=i, value=h)
        cell.fill = _EXCEL_GREEN
        cell.font = _EXCEL_WHITE_BOLD
        cell.alignment = _EXCEL_CENTER

def _apply_grid(ws, max_row: int, max_col: int = 9) -> None:
    """Apply borders and alignment to grid cells."""
    if not OPENPYXL_AVAILABLE:
        return
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = _EXCEL_BORDER_ALL
            cell.alignment = _EXCEL_WRAP if c != 1 else _EXCEL_LEFT

def _status_dropdown(ws, start_row: int, end_row: int = 500) -> None:
    """Add data validation dropdown for Status column."""
    if not OPENPYXL_AVAILABLE:
        return
    dv = DataValidation(type="list", formula1='"Pass,Fail,Blocked,N/A"', allow_blank=True)
    dv.add(f"F{start_row}:F{end_row}")
    ws.add_data_validation(dv)

def init_testcase_excel(
    path: Union[str, Path],
    module_name: str,
    spec_id: str,
    description: str,
    prerequisites: str,
    env_info: str,
    scenario: str,
) -> Optional[Path]:
    """
    Create a test case workbook with metadata and the steps table.
    Returns the saved file path, or None if openpyxl is not available.
    Session-independent: file created in results/ directory with timestamp.
    """
    if not OPENPYXL_AVAILABLE:
        logger.warning("openpyxl not installed, cannot create Excel test case file")
        return None
    
    path = Path(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Case Template"

    # Column widths
    for col, w in {"A": 18, "B": 26, "C": 30, "D": 30, "E": 30, "F": 12, "G": 18, "H": 10, "I": 10}.items():
        ws.column_dimensions[col].width = w

    def label(row, text, height=18):
        ws.row_dimensions[row].height = height
        cell = ws.cell(row=row, column=1, value=text)
        cell.fill = _EXCEL_GREEN
        cell.font = _EXCEL_WHITE_BOLD
        cell.alignment = _EXCEL_LEFT
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)

    # Top metadata
    label(1, "Module Name:-")
    ws["B1"].value = module_name
    label(2, "Test Case Spec ID")
    ws["B2"].value = spec_id

    ws.merge_cells(start_row=3, start_column=1, end_row=5, end_column=1)
    c = ws.cell(row=3, column=1, value="Test Case Description")
    c.fill = _EXCEL_GREEN
    c.font = _EXCEL_WHITE_BOLD
    c.alignment = _EXCEL_CENTER
    ws.merge_cells(start_row=3, start_column=2, end_row=5, end_column=9)
    ws["B3"].value = description
    ws["B3"].alignment = _EXCEL_LEFT

    ws.merge_cells(start_row=6, start_column=1, end_row=8, end_column=1)
    c = ws.cell(row=6, column=1, value="Prerequisites:")
    c.fill = _EXCEL_GREEN
    c.font = _EXCEL_WHITE_BOLD
    c.alignment = _EXCEL_CENTER
    ws.merge_cells(start_row=6, start_column=2, end_row=8, end_column=9)
    ws["B6"].value = prerequisites
    ws["B6"].alignment = _EXCEL_LEFT

    ws.merge_cells(start_row=9, start_column=1, end_row=11, end_column=1)
    c = ws.cell(row=9, column=1, value="Environmental Information:-")
    c.fill = _EXCEL_GREEN
    c.font = _EXCEL_WHITE_BOLD
    c.alignment = _EXCEL_CENTER
    ws.merge_cells(start_row=9, start_column=2, end_row=11, end_column=9)
    ws["B9"].value = env_info
    ws["B9"].alignment = _EXCEL_LEFT

    ws.merge_cells(start_row=12, start_column=1, end_row=12, end_column=9)
    c = ws.cell(row=12, column=1, value="Test Scenario")
    c.fill = _EXCEL_GREEN
    c.font = _EXCEL_WHITE_BOLD
    c.alignment = _EXCEL_LEFT
    ws.merge_cells(start_row=13, start_column=1, end_row=13, end_column=9)
    ws["A13"].value = scenario
    ws["A13"].alignment = _EXCEL_LEFT

    # Steps table
    table_header_row = 15
    _apply_table_header(ws, table_header_row)
    _status_dropdown(ws, table_header_row + 1, table_header_row + 500)
    ws.freeze_panes = ws[f"A{table_header_row + 1}"]

    # Borders/grid (leave extra space for future rows)
    _apply_grid(ws, table_header_row + 500)

    # Save
    wb.save(path)
    logger.info(f"Created Excel test case file: {path}")
    return path

def add_testcase_to_excel(
    path: Union[str, Path],
    case_id: str,
    steps: List[str],
    test_input: Union[str, List[str]],
    expected_results: str,
) -> Optional[int]:
    """
    Appends a multi-row test case record to an existing Excel file.
    - steps: list[str] of actions (each becomes a new row)
    - test_input: str or list[str] (broadcast if str)
    - expected_results: placed on the FIRST row of the block
    Returns the first row index written, or None if openpyxl is not available.
    """
    if not OPENPYXL_AVAILABLE:
        logger.warning("openpyxl not installed, cannot add test case to Excel file")
        return None
    
    path = Path(path)
    if not path.exists():
        logger.error(f"Excel file not found: {path}")
        return None
    
    wb = load_workbook(path)
    ws = wb.active

    # Find first empty row under table header
    header_row = 15
    row = header_row + 1
    while ws.cell(row=row, column=1).value or ws.cell(row=row, column=2).value:
        row += 1

    # Normalize test_input to list
    if isinstance(test_input, list):
        inputs = test_input + [""] * max(0, len(steps) - len(test_input))
    else:
        inputs = [test_input] + [""] * (len(steps) - 1)

    first_row = row
    for i, step in enumerate(steps):
        ws.cell(row=row, column=1, value=case_id if i == 0 else "")
        ws.cell(row=row, column=2, value=step)
        ws.cell(row=row, column=3, value=inputs[i] if i < len(inputs) else "")
        ws.cell(row=row, column=4, value=expected_results if i == 0 else "")
        # Cols E,F,G left blank by default (Actual, Status, Comments)

        # Apply light gray fill to first row of test case
        if i == 0:
            for c in range(1, 8):  # Columns A through G
                ws.cell(row=row, column=c).fill = _EXCEL_LIGHT_GRAY

        row += 1

    wb.save(path)
    logger.info(f"Added test case {case_id} to Excel file: {path}")
    return first_row

# ---------- Markdown-to-Excel Analyzer (Internal LLM Session) ----------

MD_TO_EXCEL_ANALYZER_PROMPT = """\
You are a deterministic converter that MUST respond by calling exactly ONE tool.

### Workflow
1. Read the markdown content.
2. Determine whether the required fields are present (Module Name, Test Case Spec ID, Description, Prerequisites, Environmental Information, Test Scenario, and at least one test case with ID, steps, expected results).
3. Always invoke one function:
    - If the required fields exist (even if some values are TBD), call `process_md_to_excel`.
    - Otherwise call `markdown_to_pdf_fallback`.

### Non-negotiable rules
- You MUST issue exactly one function call. Plain-text answers are forbidden.
- Never refuse, apologize, or ask questions. Default to `markdown_to_pdf_fallback` when unsure.
- When using `process_md_to_excel`, pass every test case row you can extract. Convert `<br>` separators into individual step strings.
- Preserve wording from the markdown. Trim whitespace but do not invent details.

### Example payloads (for guidance only)
```
process_md_to_excel({
    "module_name": "Example Module",
    "spec_id": "EX-001",
    "description": "Example description",
    "prerequisites": "1. Item one\n2. Item two",
    "env_info": "| Item | Details |...",
    "scenario": "Scenario text",
    "test_cases": [
        {
            "case_id": "CASE-001",
            "steps": ["Step 1", "Step 2"],
            "test_input": "Input details",
            "expected_results": "Expected outcome"
        }
    ]
})

markdown_to_pdf_fallback({
    "reason": "Missing test case table"
})
```

Return only the tool call and the brief final acknowledgement mandated by the system.
"""

def create_md_to_excel_analyzer_tool_defs() -> List[Dict[str, Any]]:
    """Return tool definitions for the MD-to-Excel analyzer LLM session."""
    return [
        {
            "name": "process_md_to_excel",
            "description": "Process markdown test specification and create Excel workbook",
            "parameters": {
                "type": "object",
                "properties": {
                    "module_name": {"type": "string", "description": "Module name being tested"},
                    "spec_id": {"type": "string", "description": "Test specification ID"},
                    "description": {"type": "string", "description": "Test case description"},
                    "prerequisites": {"type": "string", "description": "Prerequisites (multiline)"},
                    "env_info": {"type": "string", "description": "Environmental information"},
                    "scenario": {"type": "string", "description": "Test scenario"},
                    "test_cases": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "case_id": {"type": "string", "description": "Test case ID"},
                                "steps": {"type": "array", "items": {"type": "string"}, "description": "Ordered test steps"},
                                "test_input": {"type": "string", "description": "Input data (string or JSON array string)"},
                                "expected_results": {"type": "string", "description": "Expected results"}
                            },
                            "required": ["case_id", "steps", "expected_results"]
                        },
                        "description": "List of test cases to add"
                    }
                },
                "required": ["module_name", "spec_id", "description", "prerequisites", "env_info", "scenario", "test_cases"]
            }
        },
        {
            "name": "markdown_to_pdf_fallback",
            "description": "Convert markdown file to PDF as fallback when not suitable for Excel",
            "parameters": {
                "type": "object",
                "properties": {
                    "reason": {"type": "string", "description": "Why conversion to Excel is not possible"}
                },
                "required": ["reason"]
            }
        }
    ]

def execute_process_md_to_excel(
    module_name: str,
    spec_id: str,
    description: str,
    prerequisites: str,
    env_info: str,
    scenario: str,
    test_cases: Union[str, List[Dict[str, Any]]],
    output_file: Optional[str] = None
) -> Dict[str, Any]:
    """
    Execute markdown-to-Excel conversion: initialize template and add test cases.
    Runs in parallel: init + all adds happen concurrently.
    """
    try:
        if not output_file:
            raise ValueError("output_file must be provided for Excel generation")
        output_path = Path(output_file)
        
        # Parse test_cases if it's a JSON string (from LLM)
        if isinstance(test_cases, str):
            try:
                import json
                test_cases = json.loads(test_cases)
            except (json.JSONDecodeError, ValueError):
                logger.warning(f"Could not parse test_cases as JSON: {test_cases[:100]}")
                test_cases = []
        
        # Ensure test_cases is a list
        if not isinstance(test_cases, list):
            logger.warning(f"test_cases is not a list: {type(test_cases)}")
            test_cases = []
        
        # Initialize Excel template
        excel_path = init_testcase_excel(
            output_path,
            module_name=module_name,
            spec_id=spec_id,
            description=description,
            prerequisites=prerequisites,
            env_info=env_info,
            scenario=scenario
        )
        
        if not excel_path:
            return {
                "status": "error",
                "message": "Failed to initialize Excel file (openpyxl not available)"
            }
        
        # Add all test cases (sequential but fast)
        added_count = 0
        errors = []
        
        for test_case in test_cases:
            try:
                case_id = test_case.get("case_id", "")
                steps = test_case.get("steps", [])
                test_input = test_case.get("test_input", "")
                expected = test_case.get("expected_results", "")
                
                if not all([case_id, steps, expected]):
                    errors.append(f"Skipped {case_id}: missing required fields")
                    continue
                
                # If test_input is a JSON string array, parse it
                if isinstance(test_input, str) and test_input.startswith("["):
                    try:
                        import json
                        test_input = json.loads(test_input)
                    except (json.JSONDecodeError, ValueError):
                        # Keep as string if JSON parsing fails
                        pass
                
                row = add_testcase_to_excel(
                    excel_path,
                    case_id=case_id,
                    steps=steps,
                    test_input=test_input,
                    expected_results=expected
                )
                
                if row is not None:
                    added_count += 1
                else:
                    errors.append(f"Failed to add {case_id}")
                    
            except Exception as e:
                logger.error(f"Error adding test case: {e}")
                errors.append(f"Exception for {test_case.get('case_id', 'unknown')}: {str(e)}")
        
        return {
            "status": "success",
            "file_path": str(excel_path),
            "test_cases_added": added_count,
            "errors": errors if errors else None
        }
        
    except Exception as e:
        logger.error(f"MD-to-Excel conversion failed: {e}")
        return {
            "status": "error",
            "message": str(e)
        }

def execute_markdown_to_pdf_fallback(
    reason: str,
    markdown_file: Optional[str] = None,
    output_file: Optional[str] = None
) -> Dict[str, Any]:
    """
    Fallback conversion: markdown to PDF. If PDF fails, keeps markdown.
    """
    try:
        if not markdown_file:
            raise ValueError("markdown_file must be provided for fallback conversion")
        md_path = Path(markdown_file)
        if not md_path.exists():
            return {"status": "error", "message": f"Markdown file not found: {markdown_file}"}
        
        # Try PDF conversion
        pdf_path = convert_markdown_to_pdf(md_path)
        
        if pdf_path:
            return {
                "status": "success",
                "format": "pdf",
                "file_path": str(pdf_path),
                "message": "Converted to PDF",
                "reason": reason
            }
        else:
            # Fallback: keep markdown as-is and reference the original path
            import shutil
            if output_file:
                output_path = Path(output_file)
            else:
                output_path = md_path.with_suffix('.md')
            if output_path.suffix.lower() != ".md":
                output_path = output_path.with_suffix(".md")
            if output_path.resolve() != md_path.resolve():
                shutil.copy2(md_path, output_path)
                fallback_path = output_path
            else:
                fallback_path = md_path
            return {
                "status": "success",
                "format": "markdown",
                "file_path": str(fallback_path),
                "message": "PDF conversion unavailable; kept markdown file",
                "reason": reason
            }
            
    except Exception as e:
        logger.error(f"Markdown fallback conversion failed: {e}")
        return {
            "status": "error",
            "message": str(e)
        }

def call_md_to_excel_analyzer(client: genai.Client, md_file_path: Path, output_dir: Path, debug_dir: Optional[Path] = None) -> Dict[str, Any]:
    """
    Spawn an independent LLM session to analyze markdown and convert to Excel.
    Returns status of conversion and file paths.
    """
    md_content = md_file_path.read_text(encoding='utf-8', errors='replace')
    
    # Log the markdown content being analyzed
    if debug_dir:
        _write_debug_log(debug_dir, md_file_path.name, "markdown_content", {
            "content_length": len(md_content),
            "first_300_chars": md_content[:300],
            "has_test_spec_markers": "##" in md_content or "Test Case" in md_content
        })
    
    # Build a more directive user message that explicitly tells the LLM to call a tool
    user_message = f"""\
You MUST analyze this markdown and call exactly ONE tool.

MARKDOWN FILE: {md_file_path.name}

---
{md_content}
---

INSTRUCTIONS:
1. If this markdown contains Module Name, Spec ID, Description, Prerequisites, Environment, Scenario, and well-formed Test Cases:
    Call: process_md_to_excel with the extracted fields.

2. If any required section is missing or malformed:
    Call: markdown_to_pdf_fallback with a short reason (e.g., "missing prerequisites").

YOU MUST CALL ONE OF THESE TOOLS. Do not respond without calling a tool.
"""

    if debug_dir:
        _write_debug_log(debug_dir, md_file_path.name, "user_message", {
            "message_length": len(user_message),
            "first_400_chars": user_message[:400]
        })

    analyzer_tools = create_md_to_excel_analyzer_tool_defs()
    
    config = types.GenerateContentConfig(
        temperature=0.1,  # Minimal temperature for deterministic tool calling
        tools=[types.Tool(function_declarations=analyzer_tools)],
        system_instruction=MD_TO_EXCEL_ANALYZER_PROMPT
    )
    
    if debug_dir:
        _write_debug_log(debug_dir, md_file_path.name, "config", {
            "model": get_model_name(),
            "temperature": 0.1,
            "num_tools": len(analyzer_tools),
            "tool_names": [t.get("name") for t in analyzer_tools]
        })
    
    # Use non-streaming approach first for reliability
    try:
        response = client.models.generate_content(
            model=get_model_name(),
            contents=[types.Content(role='user', parts=[types.Part.from_text(text=user_message)])],
            config=config
        )
    except Exception as e:
        error_detail = {
            "error_type": type(e).__name__,
            "error_message": str(e),
            "traceback": traceback.format_exc()
        }
        if debug_dir:
            _write_debug_log(debug_dir, md_file_path.name, "stream_error", error_detail)
        raise
    
    result = {
        "file_path": None,
        "format": None,
        "status": "pending",
        "message": "",
        "tool_calls": []  # Track all tool calls
    }
    
    response_text = ""
    
    print(response)
    print(response.candidates)
    
    # Process the response (non-streaming)
    if response.candidates:
        candidate = response.candidates[0]
        
        # Process function calls AND text responses
        if candidate.content and candidate.content.parts:
            for part in candidate.content.parts:
                # Handle function calls
                if part.function_call:
                    func_name = part.function_call.name
                    args = {k: v for k, v in part.function_call.args.items()}
                    
                    # Log the tool call
                    tool_call_info = {
                        "function": func_name,
                        "args_keys": list(args.keys()),
                        "args_summary": {k: (v[:100] if isinstance(v, str) else v) for k, v in args.items()}
                    }
                    result["tool_calls"].append(tool_call_info)
                    
                    if debug_dir:
                        _write_debug_log(debug_dir, md_file_path.name, f"tool_call_{func_name}", {
                            "function": func_name,
                            "full_args": args
                        })
                    
                    try:
                        if func_name == "process_md_to_excel":
                            args.setdefault("output_file", str(output_dir / f"{md_file_path.stem}.xlsx"))
                            exec_result = execute_process_md_to_excel(**args)
                            result.update({
                                "format": "excel",
                                "status": exec_result.get("status"),
                                "file_path": exec_result.get("file_path"),
                                "message": f"Created Excel with {exec_result.get('test_cases_added', 0)} test cases"
                            })
                            logger.info(f"MD-to-Excel: {result['message']}")
                            
                            if debug_dir:
                                _write_debug_log(debug_dir, md_file_path.name, "exec_result_excel", exec_result)
                            
                        elif func_name == "markdown_to_pdf_fallback":
                            args.setdefault("markdown_file", str(md_file_path))
                            args.setdefault("output_file", str(output_dir / f"{md_file_path.stem}.pdf"))
                            exec_result = execute_markdown_to_pdf_fallback(**args)
                            result.update({
                                "format": exec_result.get("format", "unknown"),
                                "status": exec_result.get("status"),
                                "file_path": exec_result.get("file_path"),
                                "message": exec_result.get("message", "Fallback conversion completed")
                            })
                            logger.info(f"MD-to-PDF/Fallback: {result['message']}")
                            
                            if debug_dir:
                                _write_debug_log(debug_dir, md_file_path.name, "exec_result_pdf", exec_result)
                    
                    except Exception as exec_error:
                        error_detail = {
                            "tool_function": func_name,
                            "error_type": type(exec_error).__name__,
                            "error_message": str(exec_error),
                            "traceback": traceback.format_exc(),
                            "args_passed": args
                        }
                        if debug_dir:
                            _write_debug_log(debug_dir, md_file_path.name, f"exec_error_{func_name}", error_detail)
                        logger.error(f"Error executing {func_name}: {exec_error}")
                        raise
                
                # Handle text responses (for debugging)
                if part.text:
                    response_text += part.text
    
    # If no tool calls were made, report a clear error
    if not result["tool_calls"]:
        if debug_dir:
            _write_debug_log(debug_dir, md_file_path.name, "llm_response_no_tools", {
                "response_text": response_text,
                "note": "LLM did not call any tools"
            })
        logger.warning("LLM did not call any tools for %s", md_file_path.name)
        result["message"] = (
            "Markdown-to-Excel analyzer did not invoke required tools; conversion skipped."
            if not response_text else
            f"LLM response without tool call: {response_text[:200]}"
        )
        result["status"] = "error"

    
    # Final summary log
    if debug_dir:
        _write_debug_log(debug_dir, md_file_path.name, "final_result", result)
    
    return result

def _python_type_to_schema(type_obj: Dict[str, Any]) -> str:
    """Helper to convert python type dict to schema type string."""
    if isinstance(type_obj, dict):
        if type_obj.get("type") == "array":
            return "ARRAY"
        if type_obj.get("type") == "object":
            return "OBJECT"
        if type_obj.get("type") == "string":
            return "STRING"
        if "oneOf" in type_obj:
            return "STRING"  # Default for oneOf
    return "STRING"

# ---------- Pipeline ----------

def process_uploads(uploads, work_dir: Path) -> Tuple[List[Path], List[Path]]:
    context_files, zip_files = [], []
    if not uploads:
        return context_files, zip_files
    for upload in uploads:
        candidate, _ = normalize_uploaded_entry(upload)
        if candidate is None:
            logger.warning("Skipping upload with no resolvable path: %r", upload)
            continue
        if not candidate.exists():
            logger.warning("Skipping missing uploaded file: %s", candidate)
            continue
        src_path = candidate
        category = categorize_upload(src_path)
        if not category:
            continue
        dst_path = copy_file_to_work_dir(src_path, work_dir)
        (zip_files if category == "zip" else context_files).append(dst_path)
    return context_files, zip_files

def process_repositories(zip_files: List[Path]) -> List[Tuple[str, str]]:
    repo_contents = []
    for zip_file in zip_files:
        try:
            logger.info(f"Processing ZIP: {zip_file.name}")
            file_path, content = extract_and_flatten_repo(zip_file)
            repo_contents.append((file_path, content))
        except Exception as e:
            logger.error(f"Failed to process {zip_file.name}: {e}")
    return repo_contents

def build_llm_request(
    user_story: str,
    repo_contents: List[Tuple[str, str]],
    uploaded_refs: List[Any],
    conversation_history: List[Dict[str, str]] = None
) -> List[Any]:
    
    user_parts: List[types.Part] = []

    if conversation_history:
        buf = io.StringIO()
        buf.write("### CONVERSATION HISTORY\n")
        for i, turn in enumerate(conversation_history, 1):
            buf.write(f"\n--- Turn {i} ---\nUser: {turn['user']}\n")
            if 'assistant_summary' in turn:
                buf.write(f"Assistant: {turn['assistant_summary']}\n")
        user_parts.append(types.Part.from_text(text=buf.getvalue()))

    user_parts.append(
        types.Part.from_text(text="### CURRENT USER REQUEST\n" + user_story.strip())
    )

    for ref in uploaded_refs or []:
        if not isinstance(ref, dict):
            continue
        file_uri = ref.get('file_uri')
        if not file_uri:
            continue
        mime_type = ref.get('mime_type') or 'application/octet-stream'
        user_parts.append(types.Part.from_uri(file_uri=file_uri, mime_type=mime_type))

    for file_path, content in repo_contents:
        chunk = content[:MAX_TOTAL_TEXT]
        user_parts.append(
            types.Part.from_text(
                text=f"### FLATTENED CODEBASE: {file_path}\n{chunk}"
            )
        )

    return [types.Content(role='user', parts=user_parts)]

def call_gemini_model(client: genai.Client, contents: List[Any], work_dir: Path, original_codebase_files: Optional[Set[str]] = None):
    """
    Stream Gemini model responses with thoughts and tool calls.
    Yields dictionaries with:
    - type: 'thought' | 'tool_call' | 'tool_response' | 'text' | 'final'
    - content: the actual content
    - metadata: additional info for display
    """
    file_tool = get_file_creation_tool_definition()

    def _extract_response_parts(resp):
        if not resp or not getattr(resp, "candidates", None):
            return None, None, []
        candidate = resp.candidates[0]
        content = getattr(candidate, "content", None)
        parts = getattr(content, "parts", None) or []
        return candidate, content, parts

    config = types.GenerateContentConfig(
        temperature=get_model_temperature(),
        thinking_config=types.ThinkingConfig(
            thinking_budget=-1,
        ),
        tools=[types.Tool(function_declarations=[file_tool])],
        system_instruction=SYSTEM_PROMPT
    )

    created_files: List[Dict[str, Any]] = []
    thinking_buffer = ""
    thinking_message_id = "thinking"

    # Stream the initial response
    response_stream = client.models.generate_content_stream(
        model=get_model_name(),
        contents=contents,
        config=config
    )

    last_candidate = None
    last_content = None
    accumulated_text = ""

    for chunk in response_stream:
        candidate, assistant_content, assistant_parts = _extract_response_parts(chunk)
        last_candidate = candidate
        last_content = assistant_content

        if candidate is None:
            continue

        # Check for thinking content (if model has extended thinking)
        for part in assistant_parts:
            # Stream thinking/reasoning
            if hasattr(part, 'thought') and part.thought:
                thinking_buffer += part.thought
                yield {
                    "type": "thought",
                    "content": thinking_buffer,
                    "metadata": {
                        "title": "🤔 Analyzing requirements...",
                        "id": thinking_message_id,
                        "status": "pending"
                    }
                }

            # Stream regular text
            if getattr(part, "text", None):
                accumulated_text += part.text
                yield {
                    "type": "text",
                    "content": accumulated_text,
                    "metadata": {}
                }

            # Handle function calls
            if getattr(part, "function_call", None):
                fc = part.function_call
                logger.info(f"Function call: {fc.name}")

                if fc.name == "create_or_edit_file":
                    file_path = fc.args.get("file_path", "")
                    description = fc.args.get("description", "")

                    # Show tool call
                    yield {
                        "type": "tool_call",
                        "content": f"**File:** `{file_path}`\n**Purpose:** {description}",
                        "metadata": {
                            "title": f"🛠️ Creating: {file_path}",
                            "id": f"tool_call_{len(created_files)}",
                            "status": "pending"
                        }
                    }

                    # Execute the tool with codebase file restrictions
                    result = execute_file_creation(
                        file_path=file_path,
                        content=fc.args.get("content", ""),
                        description=description,
                        work_dir=work_dir,
                        original_codebase_files=original_codebase_files
                    )
                    created_files.append(result)

                    # Show tool result
                    if result.get("status") == "success":
                        yield {
                            "type": "tool_response",
                            "content": f"✅ {result.get('message', 'File created successfully')}",
                            "metadata": {
                                "title": f"✅ Completed: {file_path}",
                                "id": f"tool_result_{len(created_files)-1}",
                                "status": "done"
                            }
                        }
                    else:
                        yield {
                            "type": "tool_response",
                            "content": f"❌ Error: {result.get('message', 'Unknown error')}",
                            "metadata": {
                                "title": f"❌ Failed: {file_path}",
                                "id": f"tool_result_{len(created_files)-1}",
                                "status": "done"
                            }
                        }

    # Mark thinking as done if it was shown
    if thinking_buffer:
        yield {
            "type": "thought",
            "content": thinking_buffer,
            "metadata": {
                "title": "🤔 Analysis complete",
                "id": thinking_message_id,
                "status": "done"
            }
        }

    # If there were function calls, continue the conversation
    if created_files:
        # Build tool response content
        tool_parts = [
            types.Part.from_function_response(
                name="create_or_edit_file",
                response=result
            )
            for result in created_files
        ]

        # Add assistant's last content with explicit role
        if last_content is not None:
            # Ensure the content has role="model" (Gemini uses "model" not "assistant")
            assistant_content = types.Content(role="model", parts=last_content.parts if hasattr(last_content, 'parts') else [])
            contents.append(assistant_content)
        else:
            # Add empty model response if none exists
            contents.append(types.Content(role="model", parts=[]))
        
        # Add tool responses from user
        contents.append(types.Content(role="user", parts=tool_parts))

        # Get final summary from model
        final_stream = client.models.generate_content_stream(
            model=get_model_name(),
            contents=contents,
            config=config
        )

        base_text = accumulated_text
        final_text = ""
        for chunk in final_stream:
            _, _, parts = _extract_response_parts(chunk)
            for part in parts:
                if getattr(part, "text", None):
                    final_text += part.text
                    yield {
                        "type": "text",
                        "content": base_text + final_text,
                        "metadata": {}
                    }

    # Yield final result
    combined_text = accumulated_text if not created_files else (base_text + final_text)
    if not combined_text and last_candidate is not None:
        finish_info = getattr(last_candidate, "finish_reason", "")
        if finish_info:
            combined_text = f"[finish_reason: {finish_info}]"

    yield {
        "type": "final",
        "status": "success",
        "created_files": created_files,
        "summary": combined_text,
        "total_files_created": sum(1 for f in created_files if f.get("status") == "success")
    }

def infer(user_story: str, attached_files: List[Path], create_zip: bool, conversation_history: List[Dict[str, str]], session_state: Dict[str, Any]):
    session_state = dict(session_state or {})
    session_id = session_state.get('session_id') or datetime.now().strftime('%Y%m%d_%H%M%S')
    logger.info(f"=== Processing request in session: {session_id} ===")

    progress_lines: List[str] = []

    def log_progress(line: str):
        """Log progress to backend only, not shown in chat"""
        progress_lines.append(line)
        logger.info(line)

    sanitized_story = (user_story or "").strip()
    attachment_paths = [Path(p) for p in (attached_files or [])]

    if not sanitized_story and not attachment_paths:
        yield {
            "type": "final",
            "message": "⚠️ Please provide a request or attach files.",
            "zip_path": None,
            "conversation_history": conversation_history,
            "session_state": session_state
        }
        return

    if not sanitized_story and attachment_paths:
        sanitized_story = "User uploaded new context files. Continue processing with these artifacts."

    try:
        log_progress("🔌 Initializing Gemini client…")
        client = create_gemini_client()
        log_progress("🔌 Gemini client ready ✅")

        repo_contents = session_state.get('repo_contents', [])
        uploaded_refs = session_state.get('uploaded_refs', [])
        work_dir_path = session_state.get('work_dir')
        work_dir = Path(work_dir_path) if work_dir_path else None
        original_codebase_dir = session_state.get('original_codebase_dir')
        original_codebase_path = Path(original_codebase_dir) if original_codebase_dir else None
        original_codebase_files = session_state.get('original_codebase_files', set())
        context_files: List[Path] = []
        zip_files: List[Path] = []

        if not session_state or work_dir is None:
            work_dir = Path(tempfile.mkdtemp(prefix="poc_ctx_"))
            log_progress("📥 Ingesting uploads…")
            context_files, zip_files = process_uploads(attachment_paths, work_dir)
            log_progress("📥 Ingesting uploads ✅")

            log_progress("🧩 Flattening codebase (zip → text)…")
            repo_contents = process_repositories(zip_files)
            log_progress("🧩 Flattening codebase (zip → text) ✅")

            original_codebase_path = None
            original_codebase_files = set()
            if zip_files:
                original_codebase_path = Path(tempfile.mkdtemp(prefix="codebase_"))
                with zipfile.ZipFile(zip_files[0], 'r') as zf:
                    zf.extractall(original_codebase_path)
                    # Track original file names for restriction checks
                    original_codebase_files = set(zf.namelist())

            log_progress("☁️ Uploading context to Google Files…")
            uploaded_refs = [upload_file_to_gemini(client, f) for f in context_files] if context_files else []
            for file_path, _ in repo_contents:
                flattened_ref = upload_path_to_gemini(client, Path(file_path), "text/plain")
                uploaded_refs.append(flattened_ref)
            log_progress("☁️ Uploading context to Google Files ✅")

            session_state = {
                'session_id': session_id,
                'work_dir': str(work_dir),
                'repo_contents': repo_contents,
                'uploaded_refs': uploaded_refs,
                'context_files': [str(f) for f in context_files],
                'original_codebase_dir': str(original_codebase_path) if original_codebase_path else None,
                'original_codebase_files': list(original_codebase_files)
            }
            save_llm_context(sanitized_story, repo_contents, context_files)
        else:
            work_dir = Path(work_dir)
            original_codebase_path = Path(original_codebase_path) if original_codebase_path else None
            original_codebase_files = set(session_state.get('original_codebase_files', []))

        log_progress("🤖 Calling LLM (with create_file tool)…")
        turn_start = time.time()

        contents = build_llm_request(sanitized_story, repo_contents, uploaded_refs, conversation_history)

        # Stream the model responses - pass original codebase files for validation
        result = None
        for stream_update in call_gemini_model(client, contents, work_dir, original_codebase_files):
            if stream_update.get("type") == "final":
                result = stream_update
            else:
                update_type = stream_update.get("type", "")
                if update_type in ["thought", "tool_call", "tool_response", "text"]:
                    yield stream_update

        log_progress("🤖 LLM response received ✅")

        if result is None:
            result = {
                "status": "error",
                "created_files": [],
                "summary": "No response received from the model",
                "total_files_created": 0
            }

        log_progress("☁️ Uploading newly created files to Google Files…")
        new_refs = collect_and_upload_created_files(client, work_dir, since_ts=turn_start)
        if new_refs:
            session_state.setdefault("uploaded_refs", []).extend(new_refs)
        log_progress("☁️ Uploading newly created files to Google Files ✅")

        save_json_file(result, "results", "function_call_result")

        zip_path: Optional[Path] = None
        zip_path: Optional[Path] = None
        conversion_updates: List[Dict[str, Any]] = []
        if create_zip:
            log_progress("📦 Bundling ZIP (original + generated files)…")
            original_dir = session_state.get('original_codebase_dir')
            original_dir_path = Path(original_dir) if original_dir else None
            zip_path = write_outputs_to_zip_from_workdir(
                result=result,
                work_dir=work_dir,
                original_codebase_dir=original_dir_path,
                client=client,
                status_updates=conversion_updates
            )
            log_progress("📦 ZIP bundle ready ✅")

        for event in conversion_updates:
            event_id = event.get("event_id") or f"md2excel_{hashlib.sha1(event.get('file', '').encode('utf-8')).hexdigest()[:8]}"
            display_name = event.get("display_name") or Path(event.get("file", "")).name
            if event.get("event") == "start":
                yield {
                    "type": "tool_call",
                    "content": event.get("message", ""),
                    "metadata": {
                        "title": f"🛠️ {display_name}",
                        "id": event_id,
                        "status": "pending"
                    }
                }
            elif event.get("event") == "success":
                yield {
                    "type": "tool_response",
                    "content": event.get("message", ""),
                    "metadata": {
                        "title": f"✅ {display_name}",
                        "id": event_id,
                        "status": "done"
                    }
                }
            elif event.get("event") == "error":
                yield {
                    "type": "tool_response",
                    "content": f"❌ {event.get('message', '')}",
                    "metadata": {
                        "title": f"❌ {display_name}",
                        "id": event_id,
                        "status": "done"
                    }
                }

        summary = result.get('summary', '')
        success_count = result.get('total_files_created', 0)
        created_files = result.get('created_files', [])

        updated_history = conversation_history + [{
            'user': sanitized_story,
            'assistant_summary': (summary[:500] if summary else f"Created {success_count} files")
        }]

        # Yield final result with summary, file list, and zip
        yield {
            "type": "final",
            "message": summary,
            "created_files": created_files,
            "total_files_created": success_count,
            "zip_path": str(zip_path) if zip_path else None,
            "conversation_history": updated_history,
            "session_state": session_state
        }

    except Exception as e:
        logger.error(f"Inference failed: {e}", exc_info=True)
        progress_lines.append(f"❌ Error: {str(e)}")
        yield {
            "type": "final",
            "message": "\n".join(progress_lines) if progress_lines else f"❌ Error: {str(e)}",
            "zip_path": None,
            "conversation_history": conversation_history,
            "session_state": session_state
        }

# ---------- Chat Interface ----------


def chat_handler(
    message: Any,
    history: List[Dict[str, Any]],
    conversation_state_value: List[Dict[str, Any]],
    session_state_value: Dict[str, Any]
):
    conv_state = list(conversation_state_value or [])
    sess_state = dict(session_state_value or {})
    create_zip_opt = True  # Always create ZIP

    if isinstance(message, dict):
        message_dict = message
    else:
        message_dict = {"text": str(message or ""), "files": []}

    message_text = (message_dict.get("text") or "").strip()
    file_infos: List[Tuple[Optional[Path], str]] = []
    for entry in message_dict.get("files") or []:
        path, display = normalize_uploaded_entry(entry)
        if not path and not display:
            continue
        label = display or (path.name if path else "(file)")
        file_infos.append((path, label))

    if not message_text and not file_infos:
        yield "⚠️ Please enter a message or attach files.", conv_state, sess_state
        return

    missing_files = [label for path, label in file_infos if path is None or not path.exists()]
    if missing_files:
        yield f"⚠️ Skipping missing file(s): {', '.join(missing_files)}", conv_state, sess_state

    attachments_to_process = [path for path, _ in file_infos if path and path.exists()]
    if file_infos and sess_state.get("work_dir"):
        warning = (
            "⚠️ New file uploads are ignored after the first message. Clear the chat to start a fresh session if you "
            "need to provide different files."
        )
        yield warning, conv_state, sess_state
        attachments_to_process = []

    sanitized_story = message_text
    if not sanitized_story and attachments_to_process:
        sanitized_story = "User uploaded new context files. Continue processing with these artifacts."

    if not sanitized_story and not attachments_to_process:
        yield "⚠️ Nothing to process yet. Provide instructions or new files.", conv_state, sess_state
        return

    try:
        inference = infer(sanitized_story, attachments_to_process, bool(create_zip_opt), conv_state, sess_state)
        current_conv = conv_state
        current_session = sess_state

        # Accumulate ALL messages in a list
        all_messages = []
        streaming_summary_index = None  # Track which message is the streaming summary

        for update in inference:
            update_type = update.get("type")
            current_session = update.get("session_state", current_session)
            current_conv = update.get("conversation_history", current_conv)

            if update_type == "thought":
                # Add model's thinking as collapsible
                content = update.get("content", "")
                metadata = update.get("metadata", {})
                thought_msg = ChatMessage(
                    role="assistant",
                    content=content,
                    metadata=metadata
                )
                all_messages.append(thought_msg)
                # Yield the accumulated list so far
                yield all_messages, current_conv, current_session

            elif update_type == "tool_call":
                # Add tool call as collapsible
                content = update.get("content", "")
                metadata = update.get("metadata", {})
                tool_msg = ChatMessage(
                    role="assistant",
                    content=content,
                    metadata=metadata
                )
                all_messages.append(tool_msg)
                # Yield the accumulated list so far
                yield all_messages, current_conv, current_session

            elif update_type == "tool_response":
                # Add tool result
                content = update.get("content", "")
                metadata = update.get("metadata", {})
                result_msg = ChatMessage(
                    role="assistant",
                    content=content,
                    metadata=metadata
                )
                all_messages.append(result_msg)
                # Yield the accumulated list so far
                yield all_messages, current_conv, current_session

            elif update_type == "text":
                # Stream the summary text
                content = update.get("content", "")

                if streaming_summary_index is None:
                    # First chunk - create new message
                    summary_msg = ChatMessage(
                        role="assistant",
                        content=content
                    )
                    all_messages.append(summary_msg)
                    streaming_summary_index = len(all_messages) - 1
                else:
                    # Update existing message with new content
                    all_messages[streaming_summary_index] = ChatMessage(
                        role="assistant",
                        content=content
                    )

                # Yield the accumulated list with streaming summary
                yield all_messages, current_conv, current_session

            elif update_type == "final":
                # Add summary only if it wasn't already streamed
                summary = update.get("message", "")
                if summary and streaming_summary_index is None:
                    # Summary wasn't streamed, add it now
                    all_messages.append(ChatMessage(
                        role="assistant",
                        content=summary
                    ))
                elif summary and streaming_summary_index is not None:
                    # Summary was streamed, just ensure final content is correct
                    all_messages[streaming_summary_index] = ChatMessage(
                        role="assistant",
                        content=summary
                    )

                # Add created files list as separate message
                created_files = update.get("created_files", [])
                total_files = update.get("total_files_created", 0)
                if created_files:
                    files_list = []
                    files_list.append(f"### 📁 Generated Files ({total_files})")
                    files_list.append("")
                    for file_info in created_files:
                        if file_info.get("status") == "success":
                            path = file_info.get("file_path", "")
                            desc = file_info.get("description", "")
                            files_list.append(f"✅ **{path}**")
                            if desc:
                                files_list.append(f"   *{desc}*")
                            files_list.append("")

                    all_messages.append(ChatMessage(
                        role="assistant",
                        content="\n".join(files_list)
                    ))

                # Add ZIP file attachment
                bundle_path = update.get("zip_path")
                if bundle_path:
                    bundle = Path(bundle_path)
                    if bundle.exists():
                        all_messages.append(ChatMessage(
                            role="assistant",
                            content=gr.File(
                                value=str(bundle),
                                label="📦 Download Generated Bundle"
                            )
                        ))

                # Yield final accumulated list
                yield all_messages, current_conv, current_session
                return

            sess_state = current_session
            conv_state = current_conv

        # If loop completes without final message
        all_messages.append(ChatMessage(
            role="assistant",
            content="❌ Inference interrupted unexpectedly."
        ))
        yield all_messages, current_conv, current_session

    except Exception as exc:
        logger.error("chat_handler failed: %s", exc, exc_info=True)
        all_messages.append(ChatMessage(
            role="assistant",
            content=f"❌ Error: {exc}"
        ))
        yield all_messages, conv_state, sess_state


with gr.Blocks(css="footer {visibility: hidden}") as demo:
    conversation_state = gr.State([])
    session_state = gr.State({})

    gr.ChatInterface(
        fn=chat_handler,
        type="messages",
        multimodal=True,
        save_history=True,
        chatbot=gr.Chatbot(
            label="Test Generator",
            height=600,
            type="messages",
            placeholder=INITIAL_ASSISTANT_MESSAGE  # Show as placeholder instead of initial message
        ),
        textbox=gr.MultimodalTextbox(
            placeholder="Describe your test requirements or attach files (ZIP, PDF, images)…",
            file_count="multiple",
            sources=["upload"],
            file_types=["file"]
        ),
        additional_inputs=[conversation_state, session_state],
        additional_outputs=[conversation_state, session_state],
        cache_examples=False
    )

if __name__ == "__main__":
    demo.launch()
