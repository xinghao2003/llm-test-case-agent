# app.py
import os
import io
import json
import time
import shutil
import zipfile
import tempfile
import subprocess
import logging
import platform
import mimetypes
import traceback
import hashlib
import textwrap
from datetime import datetime, timezone
from pathlib import Path
from typing import List, Dict, Any, Tuple, Optional, Set, Union, Callable

from dotenv import load_dotenv
import streamlit as st
import yaml
import pathspec
from pathspec import PathSpec

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from llm_sandbox import SandboxSession
    LLM_SANDBOX_AVAILABLE = True
except ImportError:
    SandboxSession = None
    LLM_SANDBOX_AVAILABLE = False

# ---------- Configuration ----------

MAX_TOTAL_TEXT = 250_000
ALLOWED_EXTENSIONS = {".zip", ".txt", ".md", ".pdf", ".jpg", ".jpeg", ".png"}

INITIAL_ASSISTANT_MESSAGE = (
    "Hi! I'm your test generator assistant. Attach your codebase (ZIP), documentation (PDF), or images, "
    "then describe what tests you need. I'll generate test specifications and code for you."
)

SESSION_SELECTOR_DEFAULT = "Select a previous session..."

PEPPERMINT_PRIMARY = "#3EB489"
PEPPERMINT_DARK = "#0F3F31"
PEPPERMINT_LIGHT = "#ECFFF7"
PEPPERMINT_ACCENT = "#A2F5D5"

SYSTEM_PROMPT = """\
You are a senior QA engineer and software architect generating and improving automated test plans for a Python project.

## Goals
1) Read and understand all provided context (PDFs, images, flattened repo text, and conversation history).
2) Read and understand the user's current request/requirements.
3) Analyze the codebase and create comprehensive test materials:
   - Python pytest test code files under appropriate `tests/` paths.
   - Helper modules under `tests/utils/` if beneficial.
   - Comprehensive Markdown test specifications (e.g., `TEST_SPEC.md` or under `docs/`).
   - Configuration files or auxiliary files that support testing (e.g., `conftest.py`, `pytest.ini`).
   - When necessary for test integration, you may create or edit files in the original codebase (e.g., adding imports, fixtures, or pytest hooks).
4) **PRIMARY FOCUS**: Prioritize creating new test files over editing existing source files. Only edit existing files when absolutely necessary for test framework integration.
5) **IMPORTANT**: Do NOT fix bugs, syntax errors, or any issues in the original codebase that are unrelated to test execution. Focus exclusively on test generation. If you encounter bugs or errors in the codebase, document them in test specifications but do not attempt to fix them.
6) Tests must be runnable with `pytest` out of the box (assume `pytest` only; no plugins unless absolutely needed).
7) Prefer pure-std-lib where possible. If external deps are strictly necessary, note them in the spec file.
8) Ensure meaningful file paths (POSIX-style), consistent naming, and idempotent generation.
9) Include minimal fixtures, parametrization, and positive/negative cases where it helps coverage.

## Function Calling
Use the `create_or_edit_file` function to create new files ONLY. You **cannot** edit existing files from the original codebase.

When you need to:
- **Create a new file**: Call `create_or_edit_file` with the full file path and complete content. 
  - Allowed paths: `tests/**`, `docs/**`, `test_config/**`, `test_utils/**`, or similar new testing directories.
  - **Do NOT** create or edit files in the main application source directories.
- **Edit for testing support**: Only create new configuration or helper files; never modify application source code.
- **Create multiple files**: Call `create_or_edit_file` multiple times in a single response for parallel execution.

## Content Requirements
- For `markdown` files, produce full content (headings, scenarios, cases, traceability).
- For Python test files:
  - Use pytest (`test_*.py`) naming.
  - Use clear test functions, possibly fixtures.
  - Add docstrings explaining intent.
  - Where app APIs are unclear, stub interfaces using reasonable assumptions, with TODOs clearly marked.
  - Avoid environment-dependent operations unless explicitly described in context.
- For new files:
  - Provide the COMPLETE file content.
  - Maintain consistency with the codebase style.
  - Do not reference or depend on undocumented internal app APIs.

### TEST_SPEC.md Template Guidance
When generating or editing `TEST_SPEC.md` (or similar Markdown specs), use the structure below. If specific details (for example, execution evidence or assigned tester) are unknown, populate the field with `TBD` rather than fabricating data.

# **Test Case Specification Template**

**Module Name:**
**Test Case Spec ID:**

---

### **Test Case Description**

> *Describe the purpose of this test case.*

---

### **Prerequisites**

1.
2.

---

### **Environmental Information**

| Item                   | Details          |
| ---------------------- | ---------------- |
| **Operating System**   | TBD              |
| **System Type**        | Laptop / Desktop |
| **Browser / Platform** | TBD              |

---

### **Test Scenario**

> *Describe what is being tested and why.*

---

## **Test Case Details**

| **Test Case ID** | **Test Steps**   | **Test Input** | **Expected Results** | **Verification Notes** | **Comments** |
| ---------------- | ---------------- | -------------- | -------------------- | ---------------------- | ------------ |
| 1                | 1. <br>2. <br>3. | TBD            | TBD                  | Not executed yet       | TBD          |

---

### **Notes**

* Mark unknown or future-state information as `TBD`.
* Attach screenshots, logs, or evidence externally when available.
* Keep *Expected Results* clear and measurable.

---

## **Example**

**Module Name:** Login
**Test Case ID:** gfg_01

---

### **Test Case Description**

To verify login functionality of the GeeksforGeeks website.

---

### **Prerequisites**

1. Stable internet connection
2. Browser (Chrome, Firefox, Internet Explorer, etc.)

---

### **Environmental Information**

| Item                   | Details                              |
| ---------------------- | ------------------------------------ |
| **Operating System**   | Windows / Linux / Mac                |
| **System Type**        | Laptop / Desktop                     |
| **Browser / Platform** | Chrome / Firefox / Internet Explorer |

---

### **Test Scenario**

Checking that after entering the correct username and password, the user can successfully log in.

---

## **Test Case Details**

| **Test Case ID** | **Test Steps**                                              | **Test Input**                                   | **Expected Results**                        | **Verification Notes** | **Comments**    |
| ---------------- | ----------------------------------------------------------- | ------------------------------------------------ | ------------------------------------------- | ---------------------- | --------------- |
| 1                | 1. Enter Username<br>2. Enter Password<br>3. Click on Login | Username: `geeksforgeeks`<br>Password: `geek123` | "Welcome to GeeksforGeeks!" message appears | Execution pending      | No issues found |

## Summary Output Format
After generating test files, provide a structured summary that includes:

1. **What Was Done**
   - List modules/features tested.
   - Name the files created (e.g., `tests/test_auth.py`, `TEST_SPEC.md`).
   - Describe the type of tests added (unit, integration, e2e, etc.).

2. **Why It Was Done**
   - Explain the business/technical rationale for each test suite.
   - Link tests to specific requirements or risks identified in the codebase.
   - Highlight critical paths or high-priority features covered.

3. **Test Coverage**
   - Summarize which functions, classes, or workflows are tested.
   - List the number of test cases per module.
   - Identify positive (happy path) vs. negative (error) test scenarios included.

4. **Gaps & Limitations**
   - Document what is NOT tested and why (e.g., external API mocks, environment-specific behaviors).
   - Note any assumptions made or areas requiring manual testing.
   - Mention known limitations of the pytest suite (e.g., no performance testing, no UI testing).

5. **Usage Instructions**
   - How to run the full test suite: `pytest tests/`
   - How to run specific test files or modules: `pytest tests/test_auth.py -v`
   - Any required setup (environment variables, test data, fixtures).
   - How to generate a coverage report: `pytest --cov=<module> tests/`

## Process
1) Review conversation history to understand what has been generated previously.
2) Summarize the repo: key modules, public APIs (if detectable), and how they relate to the current request.
3) For initial requests: Generate comprehensive test scenarios and cases in NEW files only.
4) For improvement requests: Only create additional test files; never modify original application code.
5) Use function calls to create new files.
6) Provide a structured summary following the **Summary Output Format** above instead of embedding code.
7) Do NOT include code snippets or detailed source code in your summary response.

## Constraints
- Call the `create_or_edit_file` function to persist **new test and documentation files only**.
- Always provide the complete file content to the function (not diffs or patches).
- Use parallel function calls when creating multiple independent files.
- Reference previous iterations when making improvements.
- **NEVER edit application source code files. Only create new test files.**
- **In your final summary message, focus on explaining WHAT, WHY, COVERAGE, GAPS, and USAGE—not the code itself.**
"""



def normalize_uploaded_entry(entry: Any) -> Tuple[Optional[Path], Optional[str]]:
    """Extract a filesystem path and display name from an upload payload."""
    candidate_path: Optional[Path] = None
    display_name: Optional[str] = None

    if isinstance(entry, dict):
        raw_path = entry.get("path") or entry.get("name") or entry.get("file")
        if raw_path:
            candidate_path = Path(raw_path)
        display_name = entry.get("orig_name") or entry.get("name")
    elif isinstance(entry, Path):
        candidate_path = entry
    elif isinstance(entry, str):
        candidate_path = Path(entry)
    elif hasattr(entry, "name"):
        candidate_path = Path(getattr(entry, "name"))
    elif hasattr(entry, "path"):
        candidate_path = Path(getattr(entry, "path"))

    if not display_name and candidate_path is not None:
        display_name = candidate_path.name

    return candidate_path, display_name

# ---------- Logger Setup ----------

def setup_logger() -> logging.Logger:
    log_dir = Path("logs")
    log_dir.mkdir(exist_ok=True)
    logger = logging.getLogger("test_generator")
    logger.setLevel(logging.DEBUG)

    log_file = log_dir / f"app_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    formatter = logging.Formatter(
        '%(asctime)s - %(levelname)s - [%(funcName)s] - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    fh = logging.FileHandler(log_file, encoding='utf-8')
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(formatter)

    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)
    ch.setFormatter(formatter)

    if not logger.handlers:
        logger.addHandler(fh)
        logger.addHandler(ch)
    return logger

logger = setup_logger()

DEFAULT_TEMPERATURE = 0.4
DEFAULT_MODEL_NAME = "gemini-flash-latest"


def get_model_temperature() -> float:
    raw_value = os.getenv("GEMINI_TEMPERATURE")
    if raw_value is None:
        return DEFAULT_TEMPERATURE
    try:
        return float(raw_value)
    except ValueError:
        logger.warning(
            "Invalid GEMINI_TEMPERATURE=%s; falling back to %.2f",
            raw_value,
            DEFAULT_TEMPERATURE,
        )
        return DEFAULT_TEMPERATURE


def get_model_name() -> str:
    return os.getenv("GEMINI_MODEL", DEFAULT_MODEL_NAME)


def apply_peppermint_theme() -> None:
    """Inject a peppermint-inspired visual theme into the Streamlit app."""
    peppermint_css = f"""
    <style>
    :root {{
        --peppermint-primary: {PEPPERMINT_PRIMARY};
        --peppermint-dark: {PEPPERMINT_DARK};
        --peppermint-light: {PEPPERMINT_LIGHT};
        --peppermint-accent: {PEPPERMINT_ACCENT};
        --peppermint-ink: #0A1F1A;
    }}

    .stApp {{
        background: radial-gradient(circle at 20% 20%, rgba(62, 180, 137, 0.12), transparent 60%),
                    radial-gradient(circle at 80% 0%, rgba(162, 245, 213, 0.35), transparent 55%),
                    var(--peppermint-light);
        color: var(--peppermint-ink);
    }}

    .stApp h1, .stApp h2, .stApp h3, .stApp h4, .stApp h5, .stApp h6 {{
        color: var(--peppermint-dark);
    }}

    section[data-testid="stSidebar"] {{
        background: linear-gradient(180deg, rgba(62, 180, 137, 0.15), rgba(236, 255, 247, 0.85));
        border-right: 1px solid rgba(62, 180, 137, 0.25);
    }}

    .stButton>button,
    div[data-testid="stDownloadButton"]>button {{
        background-color: var(--peppermint-primary);
        color: white;
        border-radius: 20px;
        border: 1px solid rgba(10, 31, 26, 0.1);
        transition: transform 0.1s ease, box-shadow 0.2s ease;
    }}

    .stButton>button:hover,
    div[data-testid="stDownloadButton"]>button:hover {{
        transform: translateY(-1px);
        box-shadow: 0 6px 18px rgba(62, 180, 137, 0.35);
    }}

    .stTabs [data-baseweb="tab"] {{
        background-color: rgba(62, 180, 137, 0.08);
        color: var(--peppermint-dark);
    }}

    .stTabs [data-baseweb="tab"]:hover {{
        background-color: rgba(62, 180, 137, 0.15);
    }}

    .stTabs [data-baseweb="tab"][aria-selected="true"] {{
        background-color: white;
        border-bottom: 3px solid var(--peppermint-primary);
    }}

    .stSelectbox>div>div>div,
    .stTextInput>div>div>input,
    .stTextArea>div>div>textarea,
    .stFileUploader>div {{
        border-radius: 14px;
        border: 1px solid rgba(62, 180, 137, 0.35);
        background-color: rgba(255, 255, 255, 0.9);
    }}

    .stSelectbox>div>div>div:focus-within,
    .stTextInput>div>div>input:focus,
    .stTextArea>div>div>textarea:focus {{
        border: 1px solid var(--peppermint-primary);
        box-shadow: 0 0 0 2px rgba(62, 180, 137, 0.25);
    }}

    [data-testid="stProgress"] div[role="progressbar"] {{
        background-color: var(--peppermint-primary);
    }}

    .stAlert {{
        border-left: 6px solid var(--peppermint-primary);
        background-color: rgba(162, 245, 213, 0.25);
    }}
    </style>
    """
    st.markdown(peppermint_css, unsafe_allow_html=True)

# ---------- Gemini Client ----------

try:
    from google import genai
    from google.genai import types
except ImportError:
    raise SystemExit(
        "google-genai not installed. Run: pip install google-genai\n"
        "Docs: https://pypi.org/project/google-genai/"
    )

def create_gemini_client() -> genai.Client:
    load_dotenv()
    api_key = os.getenv("GOOGLE_API_KEY")
    if not api_key:
        raise RuntimeError("GOOGLE_API_KEY environment variable is not set.")
    return genai.Client(api_key=api_key)

# ---------- Tool: create_or_edit_file ----------

def get_file_creation_tool_definition() -> Dict[str, Any]:
    return {
        "name": "create_or_edit_file",
        "description": "Create a new file or edit an existing file by replacing its entire content.",
        "parameters": {
            "type": "object",
            "properties": {
                "file_path": {"type": "string", "description": "POSIX path to create/edit (under temp work dir)"},
                "content": {"type": "string", "description": "Complete file content"},
                "description": {"type": "string", "description": "Short purpose/notes"}
            },
            "required": ["file_path", "content", "description"]
        }
    }

def execute_file_creation(file_path: str, content: str, description: str, work_dir: Path, original_codebase_files: Optional[Set[str]] = None) -> Dict[str, Any]:
    try:
        normalized_path = file_path.replace("\\", "/").lstrip("/")
        target_path = work_dir / normalized_path
        
        # Log if editing a file from original codebase (advisory only, not blocked)
        if original_codebase_files:
            # Normalize the target path for comparison
            path_parts = normalized_path.lower()
            
            # Check if file exists in original codebase
            for orig_file in original_codebase_files:
                if orig_file.lower() == path_parts or path_parts.endswith('/' + orig_file.lower()):
                    logger.warning(
                        f"Modifying original codebase file: {file_path}\n"
                        f"Description: {description}"
                    )
                    break
        
        target_path.parent.mkdir(parents=True, exist_ok=True)
        target_path.write_text(content, encoding='utf-8')
        logger.info(f"File created/edited: {target_path}")
        return {
            "status": "success",
            "file_path": str(target_path),
            "message": f"File '{file_path}' created/edited ({len(content)} bytes)",
            "description": description
        }
    except Exception as e:
        logger.error(f"Error creating file {file_path}: {e}", exc_info=True)
        return {"status": "error", "file_path": file_path, "message": str(e)}

def get_llm_sandbox_tool_definition() -> Dict[str, Any]:
    return {
        "name": "execute_code",
        "description": (
            "Execute code inside the LLM Sandbox with optional library installation and timeout control. "
            "Supports python, javascript, java, cpp, go, ruby, and r."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "code": {
                    "type": "string",
                    "description": "Code to execute inside the sandbox runtime"
                },
                "language": {
                    "type": "string",
                    "description": "Programming language to use",
                    "default": "python"
                },
                "libraries": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "Optional libraries/packages to install before execution"
                },
                "timeout": {
                    "type": "integer",
                    "minimum": 1,
                    "description": "Execution timeout in seconds (defaults to sandbox configuration)"
                }
            },
            "required": ["code", "language"]
        }
    }

def _normalize_sandbox_libraries(libraries: Optional[Union[List[str], str]]) -> Optional[List[str]]:
    if libraries is None:
        return None
    if isinstance(libraries, list):
        normalized = [str(lib).strip() for lib in libraries if str(lib).strip()]
        return normalized or None
    if isinstance(libraries, str):
        stripped = libraries.strip()
        if not stripped:
            return None
        try:
            parsed = json.loads(stripped)
        except json.JSONDecodeError:
            parsed = None
        if isinstance(parsed, list):
            normalized = [str(lib).strip() for lib in parsed if str(lib).strip()]
            return normalized or None
        return [stripped]
    return None

def execute_llm_sandbox_tool(
    code: str,
    language: str = "python",
    libraries: Optional[Union[List[str], str]] = None,
    timeout: Optional[Union[int, str]] = None
) -> Dict[str, Any]:
    if not LLM_SANDBOX_AVAILABLE:
        return {
            "status": "error",
            "message": "llm-sandbox is not installed. Add it to requirements.txt and reinstall dependencies."
        }

    sanitized_code = (code or "").strip()
    if not sanitized_code:
        return {"status": "error", "message": "Code input is empty; nothing to execute."}

    lang = (language or "python").strip() or "python"
    normalized_libs = _normalize_sandbox_libraries(libraries)

    exec_timeout: Optional[int] = None
    if timeout is not None:
        try:
            exec_timeout = int(timeout)
            if exec_timeout <= 0:
                raise ValueError("Timeout must be positive")
        except (TypeError, ValueError) as exc:
            return {
                "status": "error",
                "message": f"Invalid timeout value: {timeout}",
                "exception_type": type(exc).__name__
            }

    run_kwargs: Dict[str, Any] = {}
    if normalized_libs:
        run_kwargs["libraries"] = normalized_libs
    if exec_timeout is not None:
        run_kwargs["timeout"] = exec_timeout

    start_ts = time.perf_counter()
    try:
        with SandboxSession(lang=lang) as session:
            result = session.run(sanitized_code, **run_kwargs)
    except Exception as exc:  # pylint: disable=broad-except
        logger.error("Sandbox execution failed: %s", exc, exc_info=True)
        return {
            "status": "error",
            "message": str(exc),
            "exception_type": type(exc).__name__,
            "language": lang
        }

    duration = time.perf_counter() - start_ts
    response: Dict[str, Any] = {
        "status": "success",
        "language": lang,
        "duration_seconds": round(duration, 3),
        "exit_code": getattr(result, "exit_code", None),
        "stdout": getattr(result, "stdout", ""),
        "stderr": getattr(result, "stderr", "")
    }
    if normalized_libs:
        response["libraries"] = normalized_libs
    if exec_timeout is not None:
        response["timeout"] = exec_timeout

    plots = []
    for plot in getattr(result, "plots", []) or []:
        plot_info = {
            "filename": getattr(plot, "filename", None),
            "mime_type": getattr(plot, "mime_type", None),
            "content_base64": getattr(plot, "content_base64", None)
        }
        filtered = {k: v for k, v in plot_info.items() if v}
        if filtered:
            plots.append(filtered)
    if plots:
        response["plots"] = plots

    return response

# ---------- Sandbox Test Execution ----------

def _has_python_tests(created_files: List[Dict[str, Any]], work_dir: Path) -> bool:
    """Heuristic check to see if the LLM generated any pytest-style files."""
    for entry in created_files or []:
        if entry.get("status") != "success":
            continue
        file_path = Path(entry.get("file_path", ""))
        if not file_path.exists() or file_path.suffix != ".py":
            continue
        try:
            rel_path = file_path.relative_to(work_dir)
        except ValueError:
            rel_path = file_path
        parts = {part.lower() for part in rel_path.parts}
        if "tests" in parts or rel_path.name.startswith("test_"):
            return True
    return False

def _assemble_repo_for_testing(original_codebase_dir: Optional[Path], work_dir: Path, created_files: List[Dict[str, Any]]) -> Optional[Path]:
    """Create a temporary directory that merges the original codebase with generated files."""
    staging_dir = Path(tempfile.mkdtemp(prefix="sandbox_repo_"))
    try:
        if original_codebase_dir and Path(original_codebase_dir).exists():
            shutil.copytree(original_codebase_dir, staging_dir, dirs_exist_ok=True)
        files_copied = False
        for entry in created_files or []:
            if entry.get("status") != "success":
                continue
            file_path = Path(entry.get("file_path", ""))
            if not file_path.exists():
                continue
            try:
                rel_path = file_path.relative_to(work_dir)
            except ValueError:
                rel_path = Path(file_path.name)
            target_path = staging_dir / rel_path
            target_path.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(file_path, target_path)
            files_copied = True
        if not files_copied and not (original_codebase_dir and Path(original_codebase_dir).exists()):
            shutil.rmtree(staging_dir, ignore_errors=True)
            return None
        return staging_dir
    except Exception as exc:  # pylint: disable=broad-except
        shutil.rmtree(staging_dir, ignore_errors=True)
        logger.error("Failed to assemble sandbox repo: %s", exc, exc_info=True)
        return None

def _build_install_steps(repo_root: Path) -> List[Dict[str, Any]]:
    """Infer a lightweight installation sequence for the sandbox environment."""
    steps: List[Dict[str, Any]] = []
    seen: Set[Tuple[str, ...]] = set()

    def add_step(cmd: List[str], label: str) -> None:
        key = tuple(cmd)
        if key in seen:
            return
        seen.add(key)
        steps.append({"cmd": cmd, "label": label})

    add_step(["python", "-m", "pip", "install", "--upgrade", "pip"], "Upgrade pip")
    add_step(["python", "-m", "pip", "install", "pytest"], "Install pytest")

    requirement_candidates = [
        "requirements-test.txt",
        "requirements_dev.txt",
        "requirements-dev.txt",
        "dev-requirements.txt",
        "requirements.txt"
    ]
    for filename in requirement_candidates:
        candidate = repo_root / filename
        if candidate.exists():
            rel = candidate.relative_to(repo_root).as_posix()
            add_step(["python", "-m", "pip", "install", "-r", rel], f"Install {filename}")

    # Search for additional requirement files nested in subdirectories
    for req_path in sorted(repo_root.rglob("requirements*.txt")):
        if not req_path.is_file():
            continue
        try:
            rel = req_path.relative_to(repo_root).as_posix()
        except ValueError:
            continue
        add_step(
            ["python", "-m", "pip", "install", "-r", rel],
            f"Install {rel}"
        )

    setup_roots: Set[Path] = set()
    for marker in ("pyproject.toml", "setup.cfg", "setup.py"):
        for path in repo_root.rglob(marker):
            if not path.is_file():
                continue
            setup_roots.add(path.parent)

    for package_dir in sorted(setup_roots):
        try:
            rel = package_dir.relative_to(repo_root).as_posix()
        except ValueError:
            continue
        target = "." if rel == "." else rel
        label = "Install project package" if target == "." else f"Install package {rel}"
        add_step(["python", "-m", "pip", "install", target], label)

    return steps

def _bundle_repository(repo_root: Path) -> Tuple[Path, Path]:
    """Zip the staged repository for transfer to the sandbox."""
    bundle_dir = Path(tempfile.mkdtemp(prefix="sandbox_bundle_"))
    archive_base = bundle_dir / "repo_bundle"
    archive_path = Path(shutil.make_archive(str(archive_base), "zip", root_dir=repo_root))
    return archive_path, bundle_dir

def _run_repo_tests_in_sandbox(repo_root: Path, install_steps: List[Dict[str, Any]], log_destination: Path, timeout: int = 600) -> Dict[str, Any]:
    """Copy the repo into the sandbox, install dependencies, run pytest, and collect logs."""
    if not LLM_SANDBOX_AVAILABLE:
        return {
            "status": "error",
            "message": "llm-sandbox is not installed.",
            "exit_code": None,
            "log_path": None
        }

    bundle_path, bundle_tmp_dir = _bundle_repository(repo_root)
    remote_bundle = "/sandbox/project_bundle.zip"
    remote_workspace = "/sandbox/project_workspace"
    remote_log = "/sandbox/test_run.log"
    test_command = ["python", "-m", "pytest", "-q"]

    install_steps_literal = json.dumps(install_steps)
    test_command_literal = json.dumps(test_command)
    sandbox_code = textwrap.dedent(f"""
        import json
        import pathlib
        import shutil
        import subprocess
        import sys
        import zipfile

        bundle_path = pathlib.Path("{remote_bundle}")
        workspace = pathlib.Path("{remote_workspace}")
        log_path = pathlib.Path("{remote_log}")

        if workspace.exists():
            shutil.rmtree(workspace)
        workspace.mkdir(parents=True, exist_ok=True)

        if not bundle_path.exists():
            raise SystemExit("Bundle not found")
        with zipfile.ZipFile(bundle_path, "r") as zf:
            zf.extractall(workspace)

        install_steps = json.loads({install_steps_literal!r})
        test_command = json.loads({test_command_literal!r})

        def append_lines(lines, target):
            if not lines:
                return
            if not isinstance(lines, str):
                lines = str(lines)
            for raw in lines.splitlines():
                line = raw.rstrip()
                if not line:
                    continue
                print(line)
                target.append(line)

        log_lines = []
        exit_code = 0

        for step in install_steps:
            cmd = []
            label = ""
            if isinstance(step, dict):
                raw_cmd = step.get("cmd") or []
                cmd = [str(part) for part in raw_cmd]
                label = str(step.get("label") or "")
            elif isinstance(step, (list, tuple)):
                cmd = [str(part) for part in step]
            elif isinstance(step, str):
                cmd = step.split()
                label = step
            if not cmd:
                continue
            if not label:
                label = " ".join(cmd) or "<sandbox step>"
            append_lines(f"\\[sandbox\\] {{label}}", log_lines)
            proc = subprocess.run(cmd, cwd=str(workspace), capture_output=True, text=True)
            append_lines(proc.stdout, log_lines)
            append_lines(proc.stderr, log_lines)
            append_lines(f"\\[sandbox\\] exit={{proc.returncode}}", log_lines)
            if proc.returncode != 0:
                exit_code = proc.returncode
                break

        if exit_code == 0:
            append_lines("[sandbox] running tests", log_lines)
            proc = subprocess.run(test_command, cwd=str(workspace), capture_output=True, text=True)
            append_lines(proc.stdout, log_lines)
            append_lines(proc.stderr, log_lines)
            exit_code = proc.returncode
            append_lines(f"[sandbox] pytest exit={{exit_code}}", log_lines)

        log_path.write_text("\\n".join(log_lines), encoding="utf-8")
        print(f"SANDBOX_TEST_EXIT_CODE={{exit_code}}")
        print(f"SANDBOX_TEST_LOG={{log_path}}")
        sys.exit(exit_code)
    """)

    log_destination.parent.mkdir(parents=True, exist_ok=True)
    run_result: Optional[Any] = None
    try:
        with SandboxSession(lang="python") as session:
            session.copy_to_runtime(str(bundle_path), remote_bundle)
            run_result = session.run(sandbox_code, timeout=timeout)
            exit_code = getattr(run_result, "exit_code", None)
            stdout = getattr(run_result, "stdout", "")
            stderr = getattr(run_result, "stderr", "")
            log_path = None
            try:
                session.copy_from_runtime(remote_log, str(log_destination))
                log_path = str(log_destination)
            except Exception as copy_exc:  # pylint: disable=broad-except
                logger.error("Failed to copy sandbox log: %s", copy_exc, exc_info=True)
                if stdout or stderr:
                    fallback = "\n".join(filter(None, [stdout.strip(), stderr.strip()]))
                    log_destination.write_text(fallback, encoding="utf-8")
                    log_path = str(log_destination)

            status = "success" if exit_code == 0 else "failed"
            return {
                "status": status,
                "message": f"Pytest exit code {exit_code}",
                "exit_code": exit_code,
                "stdout": stdout,
                "stderr": stderr,
                "log_path": log_path
            }
    except Exception as exc:  # pylint: disable=broad-except
        logger.error("Sandbox test execution failed: %s", exc, exc_info=True)
        if not log_destination.exists():
            log_destination.write_text(f"Sandbox execution error: {exc}", encoding="utf-8")
        return {
            "status": "error",
            "message": str(exc),
            "exit_code": None,
            "stdout": getattr(run_result, "stdout", "") if run_result else "",
            "stderr": getattr(run_result, "stderr", "") if run_result else "",
            "log_path": str(log_destination)
        }
    finally:
        try:
            bundle_path.unlink(missing_ok=True)  # type: ignore[attr-defined]
        except TypeError:
            if bundle_path.exists():
                bundle_path.unlink()
        shutil.rmtree(bundle_tmp_dir, ignore_errors=True)

def maybe_run_tests_with_sandbox(
    work_dir: Path,
    original_codebase_dir: Optional[Path],
    created_files: List[Dict[str, Any]],
    log_progress: Optional[Callable[[str], None]] = None
) -> Optional[Dict[str, Any]]:
    """Entrypoint to execute pytest via llm-sandbox and return a created_files entry for the log."""
    progress = log_progress or (lambda _: None)

    if not LLM_SANDBOX_AVAILABLE:
        progress("🧪 Skipping sandbox tests (llm-sandbox unavailable)")
        return None
    if not created_files:
        progress("🧪 Skipping sandbox tests (no generated files)")
        return None
    if not _has_python_tests(created_files, work_dir):
        progress("🧪 Skipping sandbox tests (no Python test files detected)")
        return None

    repo_root = _assemble_repo_for_testing(original_codebase_dir, work_dir, created_files)
    if repo_root is None:
        progress("🧪 Skipping sandbox tests (unable to stage repo)")
        return None

    install_steps = _build_install_steps(repo_root)
    if not install_steps:
        install_steps = [{"cmd": ["python", "-m", "pip", "install", "pytest"], "label": "Install pytest"}]

    log_filename = f"sandbox_test_log_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.txt"
    log_path = work_dir / log_filename

    progress("🧪 Running pytest inside sandbox…")
    try:
        test_result = _run_repo_tests_in_sandbox(repo_root, install_steps, log_path)
    finally:
        shutil.rmtree(repo_root, ignore_errors=True)

    if not log_path.exists():
        progress("🧪 Sandbox tests failed (log missing)")
        return None

    exit_code = test_result.get("exit_code")
    message = test_result.get("message") or "Sandbox pytest execution"
    if exit_code not in (None, 0):
        progress("🧪 Sandbox tests completed with failures ⚠️")
        message = f"{message} (exit code {exit_code})"
    else:
        progress("🧪 Sandbox tests completed ✅")

    return {
        "status": "success",
        "file_path": str(log_path),
        "message": message,
        "description": "LLM Sandbox pytest execution log",
        "exit_code": exit_code
    }

# ---------- Repo-to-Text (single, de-duplicated implementation) ----------

def check_tree_command() -> bool:
    return shutil.which('tree') is not None

def is_ignored_path(file_path: str) -> bool:
    return ('.git' in file_path) or os.path.basename(file_path).startswith('repo-to-text_')

def run_tree_command(path: str) -> str:
    if platform.system() == "Windows":
        cmd = ["cmd", "/c", "tree", "/a", "/f", path]
    else:
        cmd = ["tree", "-a", "-f", "--noreport", path]
    result = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, encoding='utf-8', check=True)
    return result.stdout

def extract_full_path(line: str, path: str) -> Optional[str]:
    idx = line.find('./')
    if idx == -1:
        idx = line.find(path)
    return line[idx:].strip() if idx != -1 else None

def _mark_non_empty_dirs(relative_path: str, non_empty_dirs: Set[str]) -> None:
    dir_path = os.path.dirname(relative_path)
    while dir_path:
        non_empty_dirs.add(dir_path)
        dir_path = os.path.dirname(dir_path)

def _should_ignore_file(
    file_path: str,
    relative_path: str,
    gitignore_spec: Optional[PathSpec],
    content_ignore_spec: Optional[PathSpec],
    tree_and_content_ignore_spec: Optional[PathSpec]
) -> bool:
    relative_path = relative_path.replace(os.sep, '/')
    if relative_path.startswith('./'):
        relative_path = relative_path[2:]
    if os.path.isdir(file_path):
        relative_path += '/'

    return (
        is_ignored_path(file_path) or
        bool(gitignore_spec and gitignore_spec.match_file(relative_path)) or
        bool(content_ignore_spec and content_ignore_spec.match_file(relative_path)) or
        bool(tree_and_content_ignore_spec and tree_and_content_ignore_spec.match_file(relative_path))
    )

def _process_line(
    line: str,
    path: str,
    gitignore_spec: Optional[PathSpec],
    tree_and_content_ignore_spec: Optional[PathSpec],
    non_empty_dirs: Set[str]
) -> Optional[str]:
    full_path = extract_full_path(line, path)
    if not full_path or full_path == '.':
        return None

    try:
        relative_path = os.path.relpath(full_path, path).replace(os.sep, '/')
    except (ValueError, OSError):
        relative_path = os.path.basename(full_path)

    if _should_ignore_file(full_path, relative_path, gitignore_spec, None, tree_and_content_ignore_spec):
        return None

    if not os.path.isdir(full_path):
        _mark_non_empty_dirs(relative_path, non_empty_dirs)

    if not os.path.isdir(full_path) or os.path.dirname(relative_path) in non_empty_dirs:
        return line.replace('./', '', 1)
    return None

def _filter_tree_output(
    tree_output: str,
    path: str,
    gitignore_spec: Optional[PathSpec],
    tree_and_content_ignore_spec: Optional[PathSpec]
) -> str:
    lines = tree_output.splitlines()
    non_empty_dirs: Set[str] = set()
    filtered = [_process_line(line, path, gitignore_spec, tree_and_content_ignore_spec, non_empty_dirs) for line in lines]
    return '\n'.join(filter(None, filtered))

def get_tree_structure(
    path: str = '.',
    gitignore_spec: Optional[PathSpec] = None,
    tree_and_content_ignore_spec: Optional[PathSpec] = None
) -> str:
    if not check_tree_command():
        return ""
    try:
        tree_output = run_tree_command(path)
    except Exception:
        return ""
    if not gitignore_spec and not tree_and_content_ignore_spec:
        return tree_output
    return _filter_tree_output(tree_output, path, gitignore_spec, tree_and_content_ignore_spec)

def load_ignore_specs(path: str = '.', cli_ignore_patterns: Optional[List[str]] = None) -> Tuple[Optional[PathSpec], Optional[PathSpec], PathSpec]:
    gitignore_spec = None
    content_ignore_spec = None
    tree_and_content_ignore_list: List[str] = []
    use_gitignore = True

    repo_settings_path = os.path.join(path, '.repo-to-text-settings.yaml')
    if os.path.exists(repo_settings_path):
        try:
            with open(repo_settings_path, 'r', encoding='utf-8') as f:
                settings: Dict[str, Any] = yaml.safe_load(f) or {}
                use_gitignore = settings.get('gitignore-import-and-ignore', True)
                if 'ignore-content' in settings:
                    content_ignore_spec = pathspec.PathSpec.from_lines('gitwildmatch', settings['ignore-content'])
                if 'ignore-tree-and-content' in settings:
                    tree_and_content_ignore_list.extend(settings.get('ignore-tree-and-content', []))
        except Exception:
            pass

    if cli_ignore_patterns:
        tree_and_content_ignore_list.extend(cli_ignore_patterns)

    if use_gitignore:
        gitignore_path = os.path.join(path, '.gitignore')
        if os.path.exists(gitignore_path):
            try:
                with open(gitignore_path, 'r', encoding='utf-8') as f:
                    gitignore_spec = pathspec.PathSpec.from_lines('gitwildmatch', f)
            except Exception:
                pass

    tree_and_content_ignore_spec = pathspec.PathSpec.from_lines('gitwildmatch', tree_and_content_ignore_list)
    return gitignore_spec, content_ignore_spec, tree_and_content_ignore_spec

def _read_file_content(file_path: str) -> str:
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return f.read()
    except UnicodeDecodeError:
        with open(file_path, 'rb') as f_bin:
            return f_bin.read().decode('latin1')
    except FileNotFoundError:
        if os.path.islink(file_path) and not os.path.exists(file_path):
            try:
                target = os.readlink(file_path)
            except OSError:
                target = ''
            return f"[symlink] -> {target}"
        raise

def _load_additional_specs(path: str = '.') -> Dict[str, Any]:
    kv: Dict[str, Any] = {'maximum_word_count_per_file': None}
    repo_settings_path = os.path.join(path, '.repo-to-text-settings.yaml')
    if os.path.exists(repo_settings_path):
        try:
            with open(repo_settings_path, 'r', encoding='utf-8') as f:
                settings = yaml.safe_load(f) or {}
                max_words = settings.get('maximum_word_count_per_file')
                if isinstance(max_words, int) and max_words > 0:
                    kv['maximum_word_count_per_file'] = max_words
        except Exception:
            pass
    return kv

def _generate_output_content(
    path: str,
    tree_structure: str,
    gitignore_spec: Optional[PathSpec],
    content_ignore_spec: Optional[PathSpec],
    tree_and_content_ignore_spec: Optional[PathSpec],
    maximum_word_count_per_file: Optional[int] = None
) -> List[str]:
    output_segments: List[str] = []
    current: List[str] = []
    current_wc = 0
    project_name = os.path.basename(os.path.abspath(path))

    def count_words(text: str) -> int:
        return len(text.split())

    def finalize():
        nonlocal current_wc
        if current:
            output_segments.append("".join(current))
            current.clear()
            current_wc = 0

    def add(chunk: str):
        nonlocal current_wc
        wc = count_words(chunk)
        if maximum_word_count_per_file and current and current_wc + wc > maximum_word_count_per_file:
            finalize()
        current.append(chunk)
        current_wc += wc

    add('<repo-to-text>\n')
    add(f'Directory: {project_name}\n\n')
    add('Directory Structure:\n')
    add('<directory_structure>\n.\n')
    if os.path.exists(os.path.join(path, '.gitignore')):
        add('├── .gitignore\n')
    add(tree_structure + '\n' + '</directory_structure>\n')

    for root, _, files in os.walk(path):
        for filename in files:
            file_path = os.path.join(root, filename)
            relative_path = os.path.relpath(file_path, path)
            if _should_ignore_file(file_path, relative_path, gitignore_spec, content_ignore_spec, tree_and_content_ignore_spec):
                continue
            cleaned_relative_path = relative_path.replace('./', '', 1)
            add(f'\n<content full_path="{cleaned_relative_path}">\n')
            add(_read_file_content(file_path))
            add('\n</content>\n')

    add('\n</repo-to-text>\n')
    finalize()
    return output_segments or ["<repo-to-text>\n</repo-to-text>\n"]

def save_repo_to_text(
        path: str = '.',
        output_dir: Optional[str] = None,
        to_stdout: bool = False,
        cli_ignore_patterns: Optional[List[str]] = None
) -> str:
    gitignore_spec, content_ignore_spec, tree_and_content_ignore_spec = load_ignore_specs(path, cli_ignore_patterns)
    additional_specs = _load_additional_specs(path)
    tree_structure = get_tree_structure(path, gitignore_spec, tree_and_content_ignore_spec)
    segments = _generate_output_content(
        path,
        tree_structure,
        gitignore_spec,
        content_ignore_spec,
        tree_and_content_ignore_spec,
        additional_specs.get('maximum_word_count_per_file')
    )

    if to_stdout:
        for s in segments:
            print(s, end='')
        return "".join(segments)

    timestamp = datetime.now(timezone.utc).strftime('%Y-%m-%d-%H-%M-%S-UTC')
    base_stem = f'repo-to-text_{timestamp}'
    output_paths: List[str] = []

    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)

    if len(segments) == 1:
        filename = f"{base_stem}.txt"
        full = os.path.join(output_dir, filename) if output_dir else filename
        Path(full).write_text(segments[0], encoding='utf-8')
        print(f'[SUCCESS] Repository structure and contents saved to "{os.path.basename(full)}"')
        output_paths.append(full)
    else:
        for i, segment in enumerate(segments, 1):
            filename = f"{base_stem}_part_{i}.txt"
            full = os.path.join(output_dir, filename) if output_dir else filename
            Path(full).write_text(segment, encoding='utf-8')
            output_paths.append(full)
        print(f"[SUCCESS] Saved to {len(output_paths)} files:")
        for fp in output_paths:
            print(f'  - "{os.path.basename(fp)}"')

    return output_paths[0] if output_paths else ""

# ---------- Helpers: file ingest & Google Files ----------

def ensure_dir(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path

def copy_file_to_work_dir(src: Path, work_dir: Path) -> Path:
    dst = work_dir / (src.stem + ".txt" if src.suffix.lower() == ".md" else src.name)
    shutil.copy2(src, dst)
    return dst

def categorize_upload(file_path: Path) -> Optional[str]:
    ext = file_path.suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        logger.warning(f"Unsupported file extension: {file_path.name} ({ext})")
        return None
    return "zip" if ext == ".zip" else "context"

def get_mime_type(file_path: Path) -> str:
    mime_type, _ = mimetypes.guess_type(str(file_path))
    if mime_type:
        return mime_type
    return {
        '.md': 'text/markdown',
        '.txt': 'text/plain',
        '.pdf': 'application/pdf',
        '.jpg': 'image/jpeg',
        '.jpeg': 'image/jpeg',
        '.png': 'image/png',
        '.gif': 'image/gif',
    }.get(file_path.suffix.lower(), 'application/octet-stream')

def upload_file_to_gemini(client: genai.Client, file_path: Path) -> Dict[str, str]:
    return upload_path_to_gemini(client, file_path, None)

def upload_path_to_gemini(client: genai.Client, path: Path, mime_hint: Optional[str] = None) -> Optional[Dict[str, str]]:
    mime_type = mime_hint or get_mime_type(path)
    try:
        config = types.UploadFileConfig(mime_type=mime_type)
    except (ImportError, AttributeError):
        config = None

    uploaded_file = None
    last_error: Optional[Exception] = None

    attempt_id = 0

    def _log_failure(exc: Exception, label: str) -> None:
        logger.warning(f"Upload attempt {label} failed for {path}: {exc}")

    attempts: List[Tuple[str, Any]] = []

    if config:
        attempts.append(("file+config", lambda: client.files.upload(file=str(path), config=config)))

    attempts.append(("file+mime", lambda: client.files.upload(file=str(path), mime_type=mime_type)))

    def attempt_file_handle():
        with open(path, 'rb') as fh:
            return client.files.upload(file=fh, mime_type=mime_type)

    attempts.append(("fh+mime", attempt_file_handle))

    if hasattr(client.files, "upload_from_path"):
        attempts.append(("upload_from_path", lambda: client.files.upload_from_path(path=str(path), mime_type=mime_type)))

    for label, func in attempts:
        attempt_id += 1
        try:
            uploaded_file = func()
            if uploaded_file:
                break
        except KeyError as exc:
            last_error = exc
            _log_failure(exc, label)
        except Exception as exc:  # pylint: disable=broad-except
            last_error = exc
            _log_failure(exc, label)

    if not uploaded_file:
        logger.error(f"Giving up uploading {path}: {last_error}")
        return None

    logger.info(f"Uploaded to Google Files: {path.name}")
    file_uri = getattr(uploaded_file, "uri", None)
    if not file_uri:
        logger.warning(f"Uploaded file object missing URI for {path}; response: {uploaded_file}")
        return None
    return {"file_uri": file_uri, "mime_type": mime_type, "file_path": str(path)}


def prepare_file_for_upload(source: Path) -> Tuple[Path, Optional[Path]]:
    """Return path to upload and optional temp artifact to clean up."""
    try:
        if source.suffix.lower() == ".md":
            with tempfile.NamedTemporaryFile(
                mode='w',
                encoding='utf-8',
                suffix='.txt',
                prefix=source.stem + '_',
                dir=str(source.parent),
                delete=False
            ) as tmp_fh:
                tmp_fh.write(source.read_text(encoding='utf-8'))
                temp_path = Path(tmp_fh.name)
            logger.info(f"Converted markdown to text for upload: {source.name} -> {temp_path.name}")
            return temp_path, temp_path
    except Exception as exc:  # pylint: disable=broad-except
        logger.warning(f"Failed converting markdown {source} to text: {exc}; uploading original file.")
    return source, None

def extract_and_flatten_repo(zip_path: Path) -> Tuple[str, str]:
    temp_dir = Path(tempfile.mkdtemp(prefix="codebase_"))
    with zipfile.ZipFile(zip_path, 'r') as zf:
        zf.extractall(temp_dir)
    context_dir = ensure_dir(Path("context"))
    output_file_path = save_repo_to_text(path=str(temp_dir), output_dir=str(context_dir), to_stdout=False)
    output_file = Path(output_file_path)
    content = output_file.read_text(encoding="utf-8", errors="replace")
    if len(content) > MAX_TOTAL_TEXT:
        content = content[:MAX_TOTAL_TEXT]
        logger.warning(f"Truncated {zip_path.stem} from {len(content)} to {MAX_TOTAL_TEXT} chars")
    return output_file.as_posix(), content

def save_json_file(data: Dict[str, Any], directory: str, prefix: str) -> Path:
    dir_path = ensure_dir(Path(directory))
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    file_path = dir_path / f"{prefix}_{timestamp}.json"
    file_path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding='utf-8')
    logger.info(f"Saved {prefix} to: {file_path.name} ({file_path.stat().st_size} bytes)")
    return file_path

def save_llm_context(user_story: str, repo_files: List[Tuple[str, str]], uploaded_files: List[Path]) -> Path:
    context_data = {
        "timestamp": datetime.now().isoformat(),
        "user_story": user_story,
        "code_files": [{"path": path, "size_chars": len(content)} for path, content in repo_files],
        "uploaded_files": [{"name": f.name, "size_bytes": f.stat().st_size} for f in uploaded_files],
        "file_content_sample": {path: content[:500] for path, content in repo_files[:5]}
    }
    return save_json_file(context_data, "context_logs", "llm_context")


def persist_session_artifacts(
    session_id: str,
    user_story_text: str,
    attachments: List[Path],
    zip_path: Optional[Path],
    conversation_history: List[Dict[str, Any]],
    tool_events: Optional[List[Dict[str, Any]]] = None,
    session_metadata: Optional[Dict[str, Any]] = None,
    conversion_updates: Optional[List[Dict[str, Any]]] = None,
    summary: Optional[str] = None,
    created_files: Optional[List[Dict[str, Any]]] = None,
) -> Optional[Path]:
    """
    Persist per-session artifacts (inputs, conversation, tool calls, output bundle) locally.
    """
    try:
        base_dir = ensure_dir(Path("session_logs"))
        session_dir = ensure_dir(base_dir / session_id)
        run_dir = ensure_dir(session_dir / datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S_%f'))

        uploads_dir: Optional[Path] = None
        saved_uploads: List[Dict[str, Any]] = []
        seen_sources: Set[str] = set()

        for attachment in attachments or []:
            if not attachment:
                continue
            candidate = Path(attachment)
            source_key = str(candidate.resolve()) if candidate.exists() else str(candidate)
            if source_key in seen_sources or not candidate.exists():
                continue
            seen_sources.add(source_key)
            uploads_dir = uploads_dir or ensure_dir(run_dir / "uploads")
            target_path = uploads_dir / candidate.name
            shutil.copy2(candidate, target_path)
            saved_uploads.append({
                "original_path": str(candidate),
                "saved_path": str(target_path),
                "size_bytes": target_path.stat().st_size
            })

        saved_zip_info: Optional[Dict[str, Any]] = None
        if zip_path:
            candidate_zip = Path(zip_path)
            if candidate_zip.exists():
                output_dir = ensure_dir(run_dir / "output")
                target_zip = output_dir / candidate_zip.name
                shutil.copy2(candidate_zip, target_zip)
                saved_zip_info = {
                    "original_path": str(candidate_zip),
                    "saved_path": str(target_zip),
                    "size_bytes": target_zip.stat().st_size
                }

        metadata: Dict[str, Any] = {
            "session_id": session_id,
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "user_story_text": user_story_text,
            "summary": summary,
            "conversation_history": conversation_history,
            "tool_events": tool_events or [],
            "conversion_updates": conversion_updates or [],
            "created_files": created_files or [],
            "session_metadata": session_metadata or {},
            "uploaded_files": saved_uploads,
            "output_zip": saved_zip_info,
            "run_directory": str(run_dir)
        }

        metadata_path = run_dir / "session.json"
        metadata_path.write_text(json.dumps(metadata, indent=2, ensure_ascii=False), encoding='utf-8')
        logger.info("Saved session artifacts to %s", metadata_path)
        return run_dir

    except Exception as exc:  # pylint: disable=broad-except
        logger.error("Failed to persist session artifacts: %s", exc, exc_info=True)
        return None

def convert_markdown_to_pdf(md_file_path: Path, margin_top: str = "1in", margin_bottom: str = "1in", 
                            margin_left: str = "1in", margin_right: str = "1in") -> Optional[Path]:
    """
    Convert a Markdown file to PDF using markdown-pdf library.
    
    Args:
        md_file_path: Path to the markdown file
        margin_top: Top margin (e.g., "1in", "2cm")
        margin_bottom: Bottom margin (e.g., "1in", "2cm")
        margin_left: Left margin (e.g., "1in", "2cm")
        margin_right: Right margin (e.g., "1in", "2cm")
    
    Returns:
        The PDF file path if successful, None otherwise (markdown will be used as fallback).
    """
    pdf_file_path = md_file_path.with_suffix('.pdf')
    
    try:
        from markdown_pdf import MarkdownPdf, Section
        md_content = md_file_path.read_text(encoding='utf-8')
        pdf = MarkdownPdf(toc_level=0, optimize=True)
        pdf.add_section(Section(md_content))
        pdf.meta["title"] = md_file_path.stem
        pdf.save(str(pdf_file_path))
        logger.info(f"Converted {md_file_path} to {pdf_file_path} using markdown_pdf library")
        return pdf_file_path
    except ImportError:
        logger.debug("markdown_pdf library not installed, skipping PDF conversion")
    except Exception as e:
        logger.warning(f"Error converting {md_file_path} to PDF with markdown_pdf: {e}")
    
    # Fall back to markdown only
    logger.info(f"Could not convert {md_file_path} to PDF. Keeping markdown file only.")
    return None


def _write_debug_log(debug_dir: Path, md_filename: str, stage: str, data: Dict[str, Any]) -> None:
    """
    Write debug information to md2excel folder for troubleshooting.
    """
    try:
        debug_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')
        log_name = md_filename.replace('.md', '')
        log_file = debug_dir / f"{log_name}_{stage}_{timestamp}.json"
        
        with open(log_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, default=str)
    except Exception as e:
        logger.error(f"Failed to write debug log: {e}")

def process_markdown_files_to_excel(
    client: genai.Client,
    work_dir: Path,
    status_updates: Optional[List[Dict[str, Any]]] = None
) -> Dict[str, Dict[str, Any]]:
    """
    Scan work_dir for markdown files and attempt to convert them to Excel.
    Returns a mapping of original md path -> conversion result.
    This runs an independent LLM analyzer for each markdown file.
    Files are converted in-place (same directory as original markdown).
    """
    results = {}
    md_files = list(work_dir.glob("**/*.md"))
    
    if not md_files:
        return results
    
    logger.info(f"Found {len(md_files)} markdown files to analyze for Excel conversion")
    logger.info(f"Work dir for markdown processing: {work_dir}")
    
    debug_dir = ensure_dir(work_dir / "md2excel")
    
    for md_file in md_files:
        try:
            logger.info(f"Processing markdown: {md_file.name}")
            event_id: Optional[str] = None
            if status_updates is not None:
                event_id = f"md2excel_{hashlib.sha1(str(md_file).encode('utf-8')).hexdigest()[:8]}"
                status_updates.append({
                    "event": "start",
                    "event_id": event_id,
                    "file": str(md_file),
                    "display_name": md_file.name,
                    "message": f"Creating Excel file from {md_file.name}"
                })
            _write_debug_log(debug_dir, md_file.name, "input", {
                "file": md_file.name,
                "file_size_bytes": md_file.stat().st_size,
                "file_path": str(md_file),
                "timestamp": datetime.now(timezone.utc).isoformat()
            })
            
            # Output directory is same as markdown directory (in-place conversion)
            output_dir = md_file.parent
            
            conversion_result = call_md_to_excel_analyzer(client, md_file, output_dir, debug_dir)
            
            if conversion_result["status"] == "success":
                results[str(md_file)] = conversion_result
                logger.info(f"Successfully converted {md_file.name} to {conversion_result['format']}")
                _write_debug_log(debug_dir, md_file.name, "success", conversion_result)
                if status_updates is not None:
                    fmt = (conversion_result.get("format") or "").lower()
                    file_path_value = conversion_result.get("file_path")
                    if fmt == "excel" and file_path_value:
                        message = f"Created Excel file {Path(file_path_value).name}"
                    elif fmt == "pdf" and file_path_value:
                        message = f"Created PDF {Path(file_path_value).name}"
                    elif fmt == "markdown" and file_path_value:
                        message = f"Kept markdown file {Path(file_path_value).name}"
                    else:
                        message = "Completed conversion"
                    status_updates.append({
                        "event": "success",
                        "event_id": event_id,
                        "file": str(md_file),
                        "display_name": md_file.name,
                        "message": message,
                        "details": conversion_result
                    })
            else:
                logger.warning(f"Failed to convert {md_file.name}: {conversion_result.get('message', 'Unknown error')}")
                results[str(md_file)] = conversion_result
                _write_debug_log(debug_dir, md_file.name, "failure", conversion_result)
                if status_updates is not None:
                    status_updates.append({
                        "event": "error",
                        "event_id": event_id,
                        "file": str(md_file),
                        "display_name": md_file.name,
                        "message": conversion_result.get('message', 'Unknown error'),
                        "details": conversion_result
                    })
                
        except Exception as e:
            logger.error(f"Error processing {md_file.name}: {e}")
            error_info = {
                "status": "error",
                "message": str(e),
                "exception_type": type(e).__name__,
                "traceback": traceback.format_exc()
            }
            results[str(md_file)] = error_info
            _write_debug_log(debug_dir, md_file.name, "exception", error_info)
            if status_updates is not None:
                status_updates.append({
                    "event": "error",
                    "event_id": event_id,
                    "file": str(md_file),
                    "display_name": md_file.name,
                    "message": str(e),
                    "details": error_info
                })
    
    return results


def write_outputs_to_zip_from_workdir(
    result: Dict[str, Any],
    work_dir: Path,
    original_codebase_dir: Optional[Path] = None,
    client: Optional[genai.Client] = None,
    status_updates: Optional[List[Dict[str, Any]]] = None
) -> Path:
    output_dir = ensure_dir(Path("results"))
    timestamp = datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')
    zip_filename = output_dir / f"generated_files_{timestamp}.zip"

    original_dir = Path(original_codebase_dir) if original_codebase_dir and Path(original_codebase_dir).exists() else None
    original_root_dir = None
    root_folder_name = "generated_output"

    if original_dir:
        try:
            candidates = [p for p in original_dir.iterdir() if not p.name.startswith('__MACOSX')]
        except FileNotFoundError:
            candidates = []

        if len(candidates) == 1 and candidates[0].is_dir():
            original_root_dir = candidates[0]
            root_folder_name = candidates[0].name
        else:
            original_root_dir = original_dir
            root_folder_name = original_dir.name

    created_files: Dict[Path, Path] = {}
    for entry in result.get("created_files", []):
        file_path = Path(entry.get("file_path", ""))
        if not file_path.exists():
            continue
        try:
            rel_path = file_path.relative_to(work_dir)
        except ValueError:
            logger.debug(f"Skipping generated file outside work dir: {file_path}")
            continue
        created_files[rel_path] = file_path

    # Process markdown files: attempt Excel conversion, fallback to PDF or keep as-is
    # Keep original markdown AND add converted files alongside
    converted_files_to_add: List[Tuple[Path, Path]] = []  # (rel_path, abs_path) pairs to add to ZIP
    
    if client:
        try:
            logger.info("Starting markdown-to-Excel analysis...")
            conversion_results = process_markdown_files_to_excel(client, work_dir, status_updates=status_updates)
            
            for md_path_str, conv_result in conversion_results.items():
                original_md = Path(md_path_str)
                
                # Add converted file (Excel or PDF) to the ZIP
                output_file = conv_result.get("file_path")
                if output_file and Path(output_file).exists():
                    try:
                        converted_path = Path(output_file)
                        rel_converted = converted_path.relative_to(work_dir)
                        converted_files_to_add.append((rel_converted, converted_path))
                        
                        fmt = conv_result.get('format', 'unknown')
                        logger.info(f"Will add {converted_path.name} ({fmt}) alongside {original_md.name}")
                    except ValueError:
                        logger.debug(f"Converted file outside work dir: {output_file}")
                else:
                    logger.debug(f"No converted file for {original_md.name}")
                    
        except Exception as e:
            logger.error(f"Markdown-to-Excel processing failed: {e}")
    
    # Add all converted files to created_files (alongside originals, not replacing)
    for rel_path, abs_path in converted_files_to_add:
        created_files[rel_path] = abs_path
        logger.info(f"Added converted file to ZIP: {rel_path}")


    try:
        with zipfile.ZipFile(zip_filename, 'w', zipfile.ZIP_DEFLATED) as zf:
            if original_root_dir:
                for path in original_root_dir.rglob('*'):
                    if not path.is_file():
                        continue
                    rel = path.relative_to(original_root_dir)
                    # Skip original file if a generated file overrides the same relative path
                    if rel in created_files:
                        continue
                    arcname = Path(root_folder_name) / rel
                    zf.write(path, arcname)

            for rel_path, abs_path in created_files.items():
                arcname = Path(root_folder_name) / rel_path
                zf.write(abs_path, arcname)

        logger.info(f"ZIP file created: {zip_filename}")
        return zip_filename
    except Exception as e:
        logger.error(f"Failed to create ZIP: {e}")
        raise

def collect_and_upload_created_files(client: genai.Client, work_dir: Path, since_ts: float) -> List[Dict[str, str]]:
    """Upload files created/edited since `since_ts` under work_dir to Google Files."""
    uploaded: List[Dict[str, str]] = []
    for root, _, files in os.walk(work_dir):
        for fn in files:
            fp = Path(root) / fn
            try:
                if fp.stat().st_mtime >= since_ts:
                    upload_target, cleanup_path = prepare_file_for_upload(fp)
                    upload_info = upload_path_to_gemini(client, upload_target)
                    if upload_info:
                        uploaded.append(upload_info)
                    if cleanup_path and cleanup_path != fp:
                        try:
                            cleanup_path.unlink(missing_ok=True)
                        except Exception as exc:  # pylint: disable=broad-except
                            logger.warning(f"Failed to clean temporary upload artifact {cleanup_path}: {exc}")
            except FileNotFoundError:
                continue
            except Exception as exc:  # pylint: disable=broad-except
                logger.error(f"Failed to upload generated file {fp}: {exc}")
    return uploaded

# ---------- Excel Test Case Helpers (Internal, not exposed to LLM yet) ----------

# Core styles (reusable)
_EXCEL_GREEN = PatternFill("solid", fgColor="FF00B050") if OPENPYXL_AVAILABLE else None
_EXCEL_LIGHT_GRAY = PatternFill("solid", fgColor="FFD3D3D3") if OPENPYXL_AVAILABLE else None
_EXCEL_WHITE_BOLD = Font(color="FFFFFFFF", bold=True) if OPENPYXL_AVAILABLE else None
_EXCEL_WRAP = Alignment(wrap_text=True, vertical="center") if OPENPYXL_AVAILABLE else None
_EXCEL_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True) if OPENPYXL_AVAILABLE else None
_EXCEL_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True) if OPENPYXL_AVAILABLE else None
_EXCEL_THIN = Side(style="thin", color="FF000000") if OPENPYXL_AVAILABLE else None
_EXCEL_BORDER_ALL = Border(left=_EXCEL_THIN, right=_EXCEL_THIN, top=_EXCEL_THIN, bottom=_EXCEL_THIN) if OPENPYXL_AVAILABLE else None
_EXCEL_HEADERS = ["Test Case ID", "Test Steps", "Test Input", "Expected Results", "Actual Results", "Status", "Comments"]

def _apply_table_header(ws, start_row: int) -> None:
    """Apply header styling to a row."""
    if not OPENPYXL_AVAILABLE:
        return
    for i, h in enumerate(_EXCEL_HEADERS, start=1):
        cell = ws.cell(row=start_row, column=i, value=h)
        cell.fill = _EXCEL_GREEN
        cell.font = _EXCEL_WHITE_BOLD
        cell.alignment = _EXCEL_CENTER

def _apply_grid(ws, max_row: int, max_col: int = 9) -> None:
    """Apply borders and alignment to grid cells."""
    if not OPENPYXL_AVAILABLE:
        return
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = _EXCEL_BORDER_ALL
            cell.alignment = _EXCEL_WRAP if c != 1 else _EXCEL_LEFT

def _status_dropdown(ws, start_row: int, end_row: int = 500) -> None:
    """Add data validation dropdown for Status column."""
    if not OPENPYXL_AVAILABLE:
        return
    dv = DataValidation(type="list", formula1='"Pass,Fail,Blocked,N/A"', allow_blank=True)
    dv.add(f"F{start_row}:F{end_row}")
    ws.add_data_validation(dv)

def init_testcase_excel(
    path: Union[str, Path],
    module_name: str,
    spec_id: str,
    description: str,
    prerequisites: str,
    env_info: str,
    scenario: str,
) -> Optional[Path]:
    """
    Create a test case workbook with metadata and the steps table.
    Returns the saved file path, or None if openpyxl is not available.
    Session-independent: file created in results/ directory with timestamp.
    """
    if not OPENPYXL_AVAILABLE:
        logger.warning("openpyxl not installed, cannot create Excel test case file")
        return None
    
    path = Path(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Case Template"

    # Column widths
    for col, w in {"A": 18, "B": 26, "C": 30, "D": 30, "E": 30, "F": 12, "G": 18, "H": 10, "I": 10}.items():
        ws.column_dimensions[col].width = w

    def label(row, text, height=18):
        ws.row_dimensions[row].height = height
        cell = ws.cell(row=row, column=1, value=text)
        cell.fill = _EXCEL_GREEN
        cell.font = _EXCEL_WHITE_BOLD
        cell.alignment = _EXCEL_LEFT
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)

    # Top metadata
    label(1, "Module Name:-")
    ws["B1"].value = module_name
    label(2, "Test Case Spec ID")
    ws["B2"].value = spec_id

    ws.merge_cells(start_row=3, start_column=1, end_row=5, end_column=1)
    c = ws.cell(row=3, column=1, value="Test Case Description")
    c.fill = _EXCEL_GREEN
    c.font = _EXCEL_WHITE_BOLD
    c.alignment = _EXCEL_CENTER
    ws.merge_cells(start_row=3, start_column=2, end_row=5, end_column=9)
    ws["B3"].value = description
    ws["B3"].alignment = _EXCEL_LEFT

    ws.merge_cells(start_row=6, start_column=1, end_row=8, end_column=1)
    c = ws.cell(row=6, column=1, value="Prerequisites:")
    c.fill = _EXCEL_GREEN
    c.font = _EXCEL_WHITE_BOLD
    c.alignment = _EXCEL_CENTER
    ws.merge_cells(start_row=6, start_column=2, end_row=8, end_column=9)
    ws["B6"].value = prerequisites
    ws["B6"].alignment = _EXCEL_LEFT

    ws.merge_cells(start_row=9, start_column=1, end_row=11, end_column=1)
    c = ws.cell(row=9, column=1, value="Environmental Information:-")
    c.fill = _EXCEL_GREEN
    c.font = _EXCEL_WHITE_BOLD
    c.alignment = _EXCEL_CENTER
    ws.merge_cells(start_row=9, start_column=2, end_row=11, end_column=9)
    ws["B9"].value = env_info
    ws["B9"].alignment = _EXCEL_LEFT

    ws.merge_cells(start_row=12, start_column=1, end_row=12, end_column=9)
    c = ws.cell(row=12, column=1, value="Test Scenario")
    c.fill = _EXCEL_GREEN
    c.font = _EXCEL_WHITE_BOLD
    c.alignment = _EXCEL_LEFT
    ws.merge_cells(start_row=13, start_column=1, end_row=13, end_column=9)
    ws["A13"].value = scenario
    ws["A13"].alignment = _EXCEL_LEFT

    # Steps table
    table_header_row = 15
    _apply_table_header(ws, table_header_row)
    _status_dropdown(ws, table_header_row + 1, table_header_row + 500)
    ws.freeze_panes = ws[f"A{table_header_row + 1}"]

    # Borders/grid (leave extra space for future rows)
    _apply_grid(ws, table_header_row + 500)

    # Save
    wb.save(path)
    logger.info(f"Created Excel test case file: {path}")
    return path

def add_testcase_to_excel(
    path: Union[str, Path],
    case_id: str,
    steps: List[str],
    test_input: Union[str, List[str]],
    expected_results: str,
) -> Optional[int]:
    """
    Appends a multi-row test case record to an existing Excel file.
    - steps: list[str] of actions (each becomes a new row)
    - test_input: str or list[str] (broadcast if str)
    - expected_results: placed on the FIRST row of the block
    Returns the first row index written, or None if openpyxl is not available.
    """
    if not OPENPYXL_AVAILABLE:
        logger.warning("openpyxl not installed, cannot add test case to Excel file")
        return None
    
    path = Path(path)
    if not path.exists():
        logger.error(f"Excel file not found: {path}")
        return None
    
    wb = load_workbook(path)
    ws = wb.active

    # Find first empty row under table header
    header_row = 15
    row = header_row + 1
    while ws.cell(row=row, column=1).value or ws.cell(row=row, column=2).value:
        row += 1

    # Normalize test_input to list
    if isinstance(test_input, list):
        inputs = test_input + [""] * max(0, len(steps) - len(test_input))
    else:
        inputs = [test_input] + [""] * (len(steps) - 1)

    first_row = row
    for i, step in enumerate(steps):
        ws.cell(row=row, column=1, value=case_id if i == 0 else "")
        ws.cell(row=row, column=2, value=step)
        ws.cell(row=row, column=3, value=inputs[i] if i < len(inputs) else "")
        ws.cell(row=row, column=4, value=expected_results if i == 0 else "")
        # Cols E,F,G left blank by default (Actual, Status, Comments)

        # Apply light gray fill to first row of test case
        if i == 0:
            for c in range(1, 8):  # Columns A through G
                ws.cell(row=row, column=c).fill = _EXCEL_LIGHT_GRAY

        row += 1

    wb.save(path)
    logger.info(f"Added test case {case_id} to Excel file: {path}")
    return first_row

# ---------- Markdown-to-Excel Analyzer (Internal LLM Session) ----------

MD_TO_EXCEL_ANALYZER_PROMPT = """\
You are a deterministic converter that MUST use the provided tools to transform markdown test specs.

### Workflow
1. Read the markdown content.
2. Determine whether the required fields are present (Module Name, Test Case Spec ID, Description, Prerequisites, Environmental Information, Test Scenario, and at least one test case with ID, steps, expected results).
3. If the required fields exist (even if some values are TBD):
    a. Call `initialize_markdown_excel` once with the metadata fields.
    b. Call `append_test_case_to_excel` separately for each test case row you can extract (one function call per test case).
    c. After all tool calls complete, send a brief acknowledgement message.
4. If the required fields are missing or malformed, call `markdown_to_pdf_fallback` with a short reason.

### Non-negotiable rules
- Always begin with a tool call; do not send free-form text before the first function call.
- Never refuse, apologize, or ask questions. Default to `markdown_to_pdf_fallback` when unsure.
- When adding test cases, convert `<br>` separators into individual step strings, join them with newline characters, and preserve the source wording.
- Provide `test_input` as a single string (use a JSON array string when multiple inputs are present).
- Avoid inventing data. Trim whitespace only.

### Example payloads (for guidance only)
```
initialize_markdown_excel({
    "module_name": "Example Module",
    "spec_id": "EX-001",
    "description": "Example description",
    "prerequisites": "1. Item one\\n2. Item two",
    "env_info": "| Item | Details | ...",
    "scenario": "Scenario text"
})

append_test_case_to_excel({
    "case_id": "CASE-001",
    "steps": ["Step 1", "Step 2"],
    "test_input": "Input details",
    "expected_results": "Expected outcome"
})

markdown_to_pdf_fallback({
    "reason": "Missing test case table"
})
```

Return tool calls (in order) followed by the brief acknowledgement mandated by the system.
"""

def create_md_to_excel_analyzer_tool_defs() -> List[Dict[str, Any]]:
    """Return tool definitions for the MD-to-Excel analyzer LLM session."""
    return [
        {
            "name": "initialize_markdown_excel",
            "description": "Create or overwrite an Excel workbook using metadata parsed from markdown",
            "parameters": {
                "type": "object",
                "properties": {
                    "module_name": {"type": "string", "description": "Module name being tested"},
                    "spec_id": {"type": "string", "description": "Test specification ID"},
                    "description": {"type": "string", "description": "Test case description"},
                    "prerequisites": {"type": "string", "description": "Prerequisites (multiline)"},
                    "env_info": {"type": "string", "description": "Environmental information"},
                    "scenario": {"type": "string", "description": "Test scenario"},
                    "output_file": {"type": "string", "description": "Optional explicit Excel output path"}
                },
                "required": ["module_name", "spec_id", "description", "prerequisites", "env_info", "scenario"]
            }
        },
        {
            "name": "append_test_case_to_excel",
            "description": "Append a single test case row block to the Excel workbook",
            "parameters": {
                "type": "object",
                "properties": {
                    "case_id": {"type": "string", "description": "Test case ID"},
                    "steps": {
                        "type": "string",
                        "description": "Ordered test steps (use line breaks or `<br>` to separate steps)"
                    },
                    "test_input": {
                        "type": "string",
                        "description": "Input data (stringified value, JSON array string if multiple)"
                    },
                    "expected_results": {"type": "string", "description": "Expected results"},
                    "excel_file": {"type": "string", "description": "Optional explicit Excel file path"}
                },
                "required": ["case_id", "steps", "expected_results"]
            }
        },
        {
            "name": "markdown_to_pdf_fallback",
            "description": "Convert markdown file to PDF as fallback when not suitable for Excel",
            "parameters": {
                "type": "object",
                "properties": {
                    "reason": {"type": "string", "description": "Why conversion to Excel is not possible"}
                },
                "required": ["reason"]
            }
        }
    ]

def execute_initialize_markdown_excel(
    module_name: str,
    spec_id: str,
    description: str,
    prerequisites: str,
    env_info: str,
    scenario: str,
    output_file: Optional[str] = None
) -> Dict[str, Any]:
    """Initialize an Excel workbook using markdown metadata."""
    try:
        if not output_file:
            raise ValueError("output_file must be provided for Excel generation")
        output_path = Path(output_file)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        excel_path = init_testcase_excel(
            output_path,
            module_name=module_name,
            spec_id=spec_id,
            description=description,
            prerequisites=prerequisites,
            env_info=env_info,
            scenario=scenario
        )

        if not excel_path:
            return {
                "status": "error",
                "message": "Failed to initialize Excel file (openpyxl not available)"
            }

        return {
            "status": "success",
            "file_path": str(excel_path),
            "message": "Initialized Excel workbook"
        }

    except Exception as exc:  # pylint: disable=broad-except
        logger.error(f"Excel initialization failed: {exc}")
        return {
            "status": "error",
            "message": str(exc)
        }


def execute_append_test_case_to_excel(
    case_id: str,
    steps: Union[str, List[str]],
    expected_results: str,
    test_input: Union[str, List[str], None] = "",
    excel_file: Optional[str] = None
) -> Dict[str, Any]:
    """Append a single test case to an existing Excel workbook."""
    try:
        if not excel_file:
            raise ValueError("excel_file must be provided when appending test cases")

        excel_path = Path(excel_file)
        if not excel_path.exists():
            return {
                "status": "error",
                "message": f"Excel file not found: {excel_file}"
            }

        # Normalize steps into a clean list of strings
        if isinstance(steps, str):
            raw_steps = steps.replace("<br>", "\n").splitlines()
            normalized_steps = [s.strip() for s in raw_steps if s.strip()]
        else:
            normalized_steps = [str(s).strip() for s in steps if str(s).strip()]

        if not normalized_steps:
            return {
                "status": "error",
                "message": f"No valid steps provided for case {case_id}"
            }

        # Normalize test input
        inputs: Union[str, List[str]]
        if isinstance(test_input, list):
            inputs = [str(item).strip() for item in test_input]
        elif isinstance(test_input, str):
            stripped = test_input.strip()
            if stripped.startswith("[") and stripped.endswith("]"):
                try:
                    import json
                    parsed = json.loads(stripped)
                    inputs = parsed if isinstance(parsed, list) else stripped
                except (json.JSONDecodeError, ValueError):
                    inputs = stripped
            else:
                inputs = stripped
        else:
            inputs = ""

        row = add_testcase_to_excel(
            excel_path,
            case_id=case_id,
            steps=normalized_steps,
            test_input=inputs,
            expected_results=expected_results
        )

        if row is None:
            return {
                "status": "error",
                "message": f"Failed to append test case {case_id}"
            }

        return {
            "status": "success",
            "file_path": str(excel_path),
            "row": row,
            "message": f"Appended test case {case_id}"
        }

    except Exception as exc:  # pylint: disable=broad-except
        logger.error(f"Failed to append test case {case_id}: {exc}")
        return {
            "status": "error",
            "message": str(exc)
        }

def execute_markdown_to_pdf_fallback(
    reason: str,
    markdown_file: Optional[str] = None,
    output_file: Optional[str] = None
) -> Dict[str, Any]:
    """
    Fallback conversion: markdown to PDF. If PDF fails, keeps markdown.
    """
    try:
        if not markdown_file:
            raise ValueError("markdown_file must be provided for fallback conversion")
        md_path = Path(markdown_file)
        if not md_path.exists():
            return {"status": "error", "message": f"Markdown file not found: {markdown_file}"}
        
        # Try PDF conversion
        pdf_path = convert_markdown_to_pdf(md_path)
        
        if pdf_path:
            return {
                "status": "success",
                "format": "pdf",
                "file_path": str(pdf_path),
                "message": "Converted to PDF",
                "reason": reason
            }
        else:
            # Fallback: keep markdown as-is and reference the original path
            import shutil
            if output_file:
                output_path = Path(output_file)
            else:
                output_path = md_path.with_suffix('.md')
            if output_path.suffix.lower() != ".md":
                output_path = output_path.with_suffix(".md")
            if output_path.resolve() != md_path.resolve():
                shutil.copy2(md_path, output_path)
                fallback_path = output_path
            else:
                fallback_path = md_path
            return {
                "status": "success",
                "format": "markdown",
                "file_path": str(fallback_path),
                "message": "PDF conversion unavailable; kept markdown file",
                "reason": reason
            }
            
    except Exception as e:
        logger.error(f"Markdown fallback conversion failed: {e}")
        return {
            "status": "error",
            "message": str(e)
        }

def call_md_to_excel_analyzer(client: genai.Client, md_file_path: Path, output_dir: Path, debug_dir: Optional[Path] = None) -> Dict[str, Any]:
    """
    Spawn an independent LLM session to analyze markdown and convert to Excel.
    Returns status of conversion and file paths.
    """
    md_content = md_file_path.read_text(encoding='utf-8', errors='replace')
    
    # Log the markdown content being analyzed
    if debug_dir:
        _write_debug_log(debug_dir, md_file_path.name, "markdown_content", {
            "content_length": len(md_content),
            "first_300_chars": md_content[:300],
            "has_test_spec_markers": "##" in md_content or "Test Case" in md_content
        })
    
    # Build a more directive user message that explicitly tells the LLM to call a tool
    user_message = f"""\
You MUST analyze this markdown and use the available tools.

MARKDOWN FILE: {md_file_path.name}

---
{md_content}
---

INSTRUCTIONS:
1. If this markdown contains Module Name, Spec ID, Description, Prerequisites, Environment, Scenario, and well-formed Test Cases:
    a. Call initialize_markdown_excel with the metadata fields.
    b. Call append_test_case_to_excel once per test case row you can extract (provide steps as a newline-separated string and test_input as a string).
    c. After completing all tool calls, send a short acknowledgement message.
2. If any required section is missing or malformed:
    Call markdown_to_pdf_fallback with a short reason (e.g., "missing prerequisites").
Do not send free-form text before the first required tool call.
"""

    if debug_dir:
        _write_debug_log(debug_dir, md_file_path.name, "user_message", {
            "message_length": len(user_message),
            "first_400_chars": user_message[:400]
        })

    analyzer_tools = create_md_to_excel_analyzer_tool_defs()
    
    config = types.GenerateContentConfig(
        temperature=0.1,  # Minimal temperature for deterministic tool calling
        thinking_config = types.ThinkingConfig(
            thinking_budget=0,
        ),
        tools=[types.Tool(function_declarations=analyzer_tools)],
        system_instruction=MD_TO_EXCEL_ANALYZER_PROMPT
    )
    
    if debug_dir:
        _write_debug_log(debug_dir, md_file_path.name, "config", {
            "model": get_model_name(),
            "temperature": 0.1,
            "num_tools": len(analyzer_tools),
            "tool_names": [t.get("name") for t in analyzer_tools]
        })
    
    # Use non-streaming approach first for reliability
    try:
        response = client.models.generate_content(
            model=get_model_name(),
            contents=[types.Content(role='user', parts=[types.Part.from_text(text=user_message)])],
            config=config
        )
    except Exception as e:
        error_detail = {
            "error_type": type(e).__name__,
            "error_message": str(e),
            "traceback": traceback.format_exc()
        }
        if debug_dir:
            _write_debug_log(debug_dir, md_file_path.name, "stream_error", error_detail)
        raise
    
    result = {
        "file_path": None,
        "format": None,
        "status": "pending",
        "message": "",
        "tool_calls": [],  # Track all tool calls
        "test_cases_added": 0,
        "errors": [],
        "initialized": False
    }
    
    response_text = ""
    
    # Process the response (non-streaming)
    if response.candidates:
        candidate = response.candidates[0]
        
        # Process function calls AND text responses
        if candidate.content and candidate.content.parts:
            for part in candidate.content.parts:
                # Handle function calls
                if part.function_call:
                    func_name = part.function_call.name
                    args = {k: v for k, v in part.function_call.args.items()}
                    
                    # Log the tool call
                    tool_call_info = {
                        "function": func_name,
                        "args_keys": list(args.keys()),
                        "args_summary": {k: (v[:100] if isinstance(v, str) else v) for k, v in args.items()}
                    }
                    result["tool_calls"].append(tool_call_info)
                    
                    if debug_dir:
                        _write_debug_log(debug_dir, md_file_path.name, f"tool_call_{func_name}", {
                            "function": func_name,
                            "full_args": args
                        })
                    
                    try:
                        if func_name == "initialize_markdown_excel":
                            args.setdefault("output_file", str(output_dir / f"{md_file_path.stem}.xlsx"))
                            exec_result = execute_initialize_markdown_excel(**args)

                            if exec_result.get("status") == "success":
                                result.update({
                                    "format": "excel",
                                    "status": "initialized",
                                    "file_path": exec_result.get("file_path"),
                                    "message": exec_result.get("message", "Initialized Excel workbook")
                                })
                                result["initialized"] = True
                                logger.info(f"MD-to-Excel init: {result['message']}")
                            else:
                                error_msg = exec_result.get("message", "Unknown initialization error")
                                result["errors"].append(error_msg)
                                result["status"] = "error"
                                result["message"] = error_msg
                                logger.error(f"MD-to-Excel init failed: {error_msg}")

                            if debug_dir:
                                _write_debug_log(debug_dir, md_file_path.name, "exec_result_excel_init", exec_result)

                        elif func_name == "append_test_case_to_excel":
                            default_excel_path = result["file_path"] or str(output_dir / f"{md_file_path.stem}.xlsx")
                            args.setdefault("excel_file", default_excel_path)
                            exec_result = execute_append_test_case_to_excel(**args)

                            if exec_result.get("status") == "success":
                                result["file_path"] = exec_result.get("file_path", result["file_path"])
                                result["format"] = "excel"
                                result["test_cases_added"] += 1
                                result["status"] = "success"
                                result["message"] = f"Added {result['test_cases_added']} test case(s)"
                                logger.info(f"MD-to-Excel append: {exec_result.get('message', '')}")
                            else:
                                error_msg = exec_result.get("message", "Unknown append error")
                                result["errors"].append(error_msg)
                                # Preserve success state if some cases added; mark partial success later
                                if result["test_cases_added"] == 0:
                                    result["status"] = "error"
                                logger.error(f"MD-to-Excel append failed: {error_msg}")

                            if debug_dir:
                                _write_debug_log(debug_dir, md_file_path.name, "exec_result_excel_append", exec_result)

                        elif func_name == "markdown_to_pdf_fallback":
                            args.setdefault("markdown_file", str(md_file_path))
                            args.setdefault("output_file", str(output_dir / f"{md_file_path.stem}.pdf"))
                            exec_result = execute_markdown_to_pdf_fallback(**args)
                            result.update({
                                "format": exec_result.get("format", "unknown"),
                                "status": exec_result.get("status"),
                                "file_path": exec_result.get("file_path"),
                                "message": exec_result.get("message", "Fallback conversion completed")
                            })
                            logger.info(f"MD-to-PDF/Fallback: {result['message']}")
                            
                            if debug_dir:
                                _write_debug_log(debug_dir, md_file_path.name, "exec_result_pdf", exec_result)
                    
                    except Exception as exec_error:
                        error_detail = {
                            "tool_function": func_name,
                            "error_type": type(exec_error).__name__,
                            "error_message": str(exec_error),
                            "traceback": traceback.format_exc(),
                            "args_passed": args
                        }
                        if debug_dir:
                            _write_debug_log(debug_dir, md_file_path.name, f"exec_error_{func_name}", error_detail)
                        logger.error(f"Error executing {func_name}: {exec_error}")
                        raise
                
                # Handle text responses (for debugging)
                if part.text:
                    response_text += part.text
    
    # If no tool calls were made, report a clear error
    if not result["tool_calls"]:
        if debug_dir:
            _write_debug_log(debug_dir, md_file_path.name, "llm_response_no_tools", {
                "response_text": response_text,
                "note": "LLM did not call any tools"
            })
        logger.warning("LLM did not call any tools for %s", md_file_path.name)
        result["message"] = (
            "Markdown-to-Excel analyzer did not invoke required tools; conversion skipped."
            if not response_text else
            f"LLM response without tool call: {response_text[:200]}"
        )
        result["status"] = "error"

        return result

    if result["format"] == "excel" and result["initialized"]:
        base_message = (
            f"Initialized Excel and added {result['test_cases_added']} test case(s)"
            if result["test_cases_added"] > 0
            else "Initialized Excel workbook; no test cases appended"
        )
        if result["errors"]:
            result["status"] = "partial_success" if result["test_cases_added"] > 0 else "error"
            joined_errors = "; ".join(err for err in result["errors"] if err)
            result["message"] = f"{base_message}. Issues: {joined_errors}" if joined_errors else base_message
        else:
            result["status"] = "success"
            result["message"] = base_message
    elif result["format"] == "excel" and not result["initialized"] and result["errors"]:
        # Excel operations attempted but initialization failed
        result["status"] = "error"
        joined_errors = "; ".join(err for err in result["errors"] if err)
        if joined_errors:
            result["message"] = joined_errors

    if not result["errors"]:
        result["errors"] = None

    # Final summary log
    if debug_dir:
        _write_debug_log(debug_dir, md_file_path.name, "final_result", result)
    
    return result

def _python_type_to_schema(type_obj: Dict[str, Any]) -> str:
    """Helper to convert python type dict to schema type string."""
    if isinstance(type_obj, dict):
        if type_obj.get("type") == "array":
            return "ARRAY"
        if type_obj.get("type") == "object":
            return "OBJECT"
        if type_obj.get("type") == "string":
            return "STRING"
        if "oneOf" in type_obj:
            return "STRING"  # Default for oneOf
    return "STRING"

# ---------- Pipeline ----------

def process_uploads(uploads, work_dir: Path) -> Tuple[List[Path], List[Path]]:
    context_files, zip_files = [], []
    if not uploads:
        return context_files, zip_files
    for upload in uploads:
        candidate, _ = normalize_uploaded_entry(upload)
        if candidate is None:
            logger.warning("Skipping upload with no resolvable path: %r", upload)
            continue
        if not candidate.exists():
            logger.warning("Skipping missing uploaded file: %s", candidate)
            continue
        src_path = candidate
        category = categorize_upload(src_path)
        if not category:
            continue
        dst_path = copy_file_to_work_dir(src_path, work_dir)
        (zip_files if category == "zip" else context_files).append(dst_path)
    return context_files, zip_files

def process_repositories(zip_files: List[Path]) -> List[Tuple[str, str]]:
    repo_contents = []
    for zip_file in zip_files:
        try:
            logger.info(f"Processing ZIP: {zip_file.name}")
            file_path, content = extract_and_flatten_repo(zip_file)
            repo_contents.append((file_path, content))
        except Exception as e:
            logger.error(f"Failed to process {zip_file.name}: {e}")
    return repo_contents

def build_llm_request(
    user_story: str,
    repo_contents: List[Tuple[str, str]],
    uploaded_refs: List[Any],
    conversation_history: List[Dict[str, str]] = None
) -> List[Any]:
    
    user_parts: List[types.Part] = []

    if conversation_history:
        buf = io.StringIO()
        buf.write("### CONVERSATION HISTORY\n")
        for i, turn in enumerate(conversation_history, 1):
            buf.write(f"\n--- Turn {i} ---\nUser: {turn['user']}\n")
            if 'assistant_summary' in turn:
                buf.write(f"Assistant: {turn['assistant_summary']}\n")
        user_parts.append(types.Part.from_text(text=buf.getvalue()))

    user_parts.append(
        types.Part.from_text(text="### CURRENT USER REQUEST\n" + user_story.strip())
    )

    for ref in uploaded_refs or []:
        if not isinstance(ref, dict):
            continue
        file_uri = ref.get('file_uri')
        if not file_uri:
            continue
        mime_type = ref.get('mime_type') or 'application/octet-stream'
        user_parts.append(types.Part.from_uri(file_uri=file_uri, mime_type=mime_type))

    for file_path, content in repo_contents:
        chunk = content[:MAX_TOTAL_TEXT]
        user_parts.append(
            types.Part.from_text(
                text=f"### FLATTENED CODEBASE: {file_path}\n{chunk}"
            )
        )

    return [types.Content(role='user', parts=user_parts)]

def call_gemini_model(client: genai.Client, contents: List[Any], work_dir: Path, original_codebase_files: Optional[Set[str]] = None):
    """
    Stream Gemini model responses with thoughts and tool calls.
    Yields dictionaries with:
    - type: 'thought' | 'tool_call' | 'tool_response' | 'text' | 'final'
    - content: the actual content
    - metadata: additional info for display
    """
    file_tool = get_file_creation_tool_definition()
    sandbox_tool = get_llm_sandbox_tool_definition()

    def _extract_response_parts(resp):
        if not resp or not getattr(resp, "candidates", None):
            return None, None, []
        candidate = resp.candidates[0]
        content = getattr(candidate, "content", None)
        parts = getattr(content, "parts", None) or []
        return candidate, content, parts

    config = types.GenerateContentConfig(
        temperature=get_model_temperature(),
        thinking_config=types.ThinkingConfig(
            thinking_budget=-1,
        ),
        tools=[types.Tool(function_declarations=[file_tool, sandbox_tool])],
        system_instruction=SYSTEM_PROMPT
    )

    created_files: List[Dict[str, Any]] = []
    tool_responses: List[Dict[str, Any]] = []
    tool_event_counter = 0
    thinking_buffer = ""
    thinking_message_id = "thinking"

    # Stream the initial response
    response_stream = client.models.generate_content_stream(
        model=get_model_name(),
        contents=contents,
        config=config
    )

    last_candidate = None
    last_content = None
    accumulated_text = ""

    for chunk in response_stream:
        candidate, assistant_content, assistant_parts = _extract_response_parts(chunk)
        last_candidate = candidate
        last_content = assistant_content

        if candidate is None:
            continue

        # Check for thinking content (if model has extended thinking)
        for part in assistant_parts:
            # Stream thinking/reasoning
            if hasattr(part, 'thought') and part.thought:
                thinking_buffer += part.thought
                yield {
                    "type": "thought",
                    "content": thinking_buffer,
                    "metadata": {
                        "title": "🤔 Analyzing requirements...",
                        "id": thinking_message_id,
                        "status": "pending"
                    }
                }

            # Stream regular text
            if getattr(part, "text", None):
                accumulated_text += part.text
                yield {
                    "type": "text",
                    "content": accumulated_text,
                    "metadata": {}
                }

            # Handle function calls
            if getattr(part, "function_call", None):
                fc = part.function_call
                logger.info(f"Function call: {fc.name}")

                if fc.name == "create_or_edit_file":
                    file_path = fc.args.get("file_path", "")
                    description = fc.args.get("description", "")
                    tool_id = tool_event_counter
                    tool_event_counter += 1

                    # Show tool call
                    yield {
                        "type": "tool_call",
                        "content": f"**File:** `{file_path}`\n**Purpose:** {description}",
                        "metadata": {
                            "title": f"🛠️ Creating: {file_path}",
                            "id": f"tool_call_{tool_id}",
                            "status": "pending"
                        }
                    }

                    # Execute the tool with codebase file restrictions
                    result = execute_file_creation(
                        file_path=file_path,
                        content=fc.args.get("content", ""),
                        description=description,
                        work_dir=work_dir,
                        original_codebase_files=original_codebase_files
                    )
                    created_files.append(result)
                    tool_responses.append({
                        "name": "create_or_edit_file",
                        "response": result
                    })

                    # Show tool result
                    if result.get("status") == "success":
                        yield {
                            "type": "tool_response",
                            "content": f"✅ {result.get('message', 'File created successfully')}",
                            "metadata": {
                                "title": f"✅ Completed: {file_path}",
                                "id": f"tool_result_{tool_id}",
                                "status": "done"
                            }
                        }
                    else:
                        yield {
                            "type": "tool_response",
                            "content": f"❌ Error: {result.get('message', 'Unknown error')}",
                            "metadata": {
                                "title": f"❌ Failed: {file_path}",
                                "id": f"tool_result_{tool_id}",
                                "status": "done"
                            }
                        }
                elif fc.name == "execute_code":
                    code_snippet = fc.args.get("code", "")
                    language = fc.args.get("language", "python")
                    libraries = fc.args.get("libraries")
                    timeout = fc.args.get("timeout")
                    tool_id = tool_event_counter
                    tool_event_counter += 1

                    summary_lines = [f"**Language:** `{language}`"]
                    if libraries:
                        if isinstance(libraries, list):
                            lib_str = ", ".join(str(lib) for lib in libraries)
                        else:
                            lib_str = str(libraries)
                        summary_lines.append(f"**Libraries:** {lib_str}")
                    if timeout is not None:
                        summary_lines.append(f"**Timeout:** {timeout}s")
                    preview = (code_snippet or "").strip()
                    if preview:
                        trimmed = preview if len(preview) <= 400 else preview[:400] + "…"
                        summary_lines.append(f"```{language}\n{trimmed}\n```")

                    yield {
                        "type": "tool_call",
                        "content": "\n".join(summary_lines),
                        "metadata": {
                            "title": "🛡️ LLM Sandbox execution",
                            "id": f"tool_call_{tool_id}",
                            "status": "pending"
                        }
                    }

                    sandbox_result = execute_llm_sandbox_tool(
                        code=code_snippet,
                        language=language,
                        libraries=libraries,
                        timeout=timeout
                    )
                    tool_responses.append({
                        "name": "execute_code",
                        "response": sandbox_result
                    })

                    if sandbox_result.get("status") == "success":
                        exit_code = sandbox_result.get("exit_code")
                        stdout_preview = (sandbox_result.get("stdout") or "").strip()
                        if stdout_preview and len(stdout_preview) > 500:
                            stdout_preview = stdout_preview[:500] + "…"
                        response_lines = [f"✅ Exit code: {exit_code}"]
                        if stdout_preview:
                            response_lines.append(f"```\n{stdout_preview}\n```")
                        yield {
                            "type": "tool_response",
                            "content": "\n".join(response_lines),
                            "metadata": {
                                "title": "✅ Sandbox execution finished",
                                "id": f"tool_result_{tool_id}",
                                "status": "done"
                            }
                        }
                    else:
                        error_msg = sandbox_result.get("message", "Sandbox execution failed")
                        stderr_text = sandbox_result.get("stderr") or ""
                        if stderr_text:
                            stderr_preview = stderr_text if len(stderr_text) <= 500 else stderr_text[:500] + "…"
                            error_msg = f"{error_msg}\n```\n{stderr_preview}\n```"
                        yield {
                            "type": "tool_response",
                            "content": f"❌ {error_msg}",
                            "metadata": {
                                "title": "❌ Sandbox execution failed",
                                "id": f"tool_result_{tool_id}",
                                "status": "done"
                            }
                        }
                else:
                    logger.warning(f"Received unsupported function call: {fc.name}")

    # Mark thinking as done if it was shown
    if thinking_buffer:
        yield {
            "type": "thought",
            "content": thinking_buffer,
            "metadata": {
                "title": "🤔 Analysis complete",
                "id": thinking_message_id,
                "status": "done"
            }
        }

    base_text = accumulated_text
    final_text = ""

    # If there were function calls, continue the conversation
    if tool_responses:
        # Build tool response content
        tool_parts = [
            types.Part.from_function_response(
                name=entry["name"],
                response=entry["response"]
            )
            for entry in tool_responses
        ]

        # Add assistant's last content with explicit role
        if last_content is not None:
            # Ensure the content has role="model" (Gemini uses "model" not "assistant")
            assistant_content = types.Content(role="model", parts=last_content.parts if hasattr(last_content, 'parts') else [])
            contents.append(assistant_content)
        else:
            # Add empty model response if none exists
            contents.append(types.Content(role="model", parts=[]))
        
        # Add tool responses from user
        contents.append(types.Content(role="user", parts=tool_parts))

        # Get final summary from model
        final_stream = client.models.generate_content_stream(
            model=get_model_name(),
            contents=contents,
            config=config
        )

        final_text = ""
        for chunk in final_stream:
            _, _, parts = _extract_response_parts(chunk)
            for part in parts:
                if getattr(part, "text", None):
                    final_text += part.text
                    yield {
                        "type": "text",
                        "content": base_text + final_text,
                        "metadata": {}
                    }

    # Yield final result
    combined_text = accumulated_text if not tool_responses else (base_text + final_text)
    if not combined_text and last_candidate is not None:
        finish_info = getattr(last_candidate, "finish_reason", "")
        if finish_info:
            combined_text = f"[finish_reason: {finish_info}]"

    yield {
        "type": "final",
        "status": "success",
        "created_files": created_files,
        "summary": combined_text,
        "total_files_created": sum(1 for f in created_files if f.get("status") == "success"),
        "tool_responses": tool_responses
    }

def infer(
    user_story: str,
    attached_files: List[Path],
    create_zip: bool,
    conversation_history: List[Dict[str, str]],
    session_state: Dict[str, Any],
    session_metadata: Optional[Dict[str, Any]] = None,
):
    session_state = dict(session_state or {})
    session_id = session_state.get('session_id') or datetime.now().strftime('%Y%m%d_%H%M%S')
    logger.info(f"=== Processing request in session: {session_id} ===")

    progress_lines: List[str] = []

    def log_progress(line: str):
        """Log progress to backend only, not shown in chat"""
        progress_lines.append(line)
        logger.info(line)

    sanitized_story = (user_story or "").strip()
    attachment_paths = [Path(p) for p in (attached_files or [])]

    if not sanitized_story and not attachment_paths:
        yield {
            "type": "final",
            "message": "⚠️ Please provide a request or attach files.",
            "zip_path": None,
            "conversation_history": conversation_history,
            "session_state": session_state
        }
        return

    if not sanitized_story and attachment_paths:
        sanitized_story = "User uploaded new context files. Continue processing with these artifacts."

    try:
        log_progress("🔌 Initializing Gemini client…")
        client = create_gemini_client()
        log_progress("🔌 Gemini client ready ✅")

        repo_contents = session_state.get('repo_contents', [])
        uploaded_refs = session_state.get('uploaded_refs', [])
        work_dir_path = session_state.get('work_dir')
        work_dir = Path(work_dir_path) if work_dir_path else None
        original_codebase_dir = session_state.get('original_codebase_dir')
        original_codebase_path = Path(original_codebase_dir) if original_codebase_dir else None
        original_codebase_files = session_state.get('original_codebase_files', set())
        context_files: List[Path] = []
        zip_files: List[Path] = []

        if not session_state or work_dir is None:
            work_dir = Path(tempfile.mkdtemp(prefix="poc_ctx_"))
            log_progress("📥 Ingesting uploads…")
            context_files, zip_files = process_uploads(attachment_paths, work_dir)
            log_progress("📥 Ingesting uploads ✅")

            log_progress("🧩 Flattening codebase (zip → text)…")
            repo_contents = process_repositories(zip_files)
            log_progress("🧩 Flattening codebase (zip → text) ✅")

            original_codebase_path = None
            original_codebase_files = set()
            if zip_files:
                original_codebase_path = Path(tempfile.mkdtemp(prefix="codebase_"))
                with zipfile.ZipFile(zip_files[0], 'r') as zf:
                    zf.extractall(original_codebase_path)
                    # Track original file names for restriction checks
                    original_codebase_files = set(zf.namelist())

            log_progress("☁️ Uploading context to Google Files…")
            uploaded_refs = [upload_file_to_gemini(client, f) for f in context_files] if context_files else []
            for file_path, _ in repo_contents:
                flattened_ref = upload_path_to_gemini(client, Path(file_path), "text/plain")
                uploaded_refs.append(flattened_ref)
            log_progress("☁️ Uploading context to Google Files ✅")

            session_state = {
                'session_id': session_id,
                'work_dir': str(work_dir),
                'repo_contents': repo_contents,
                'uploaded_refs': uploaded_refs,
                'context_files': [str(f) for f in context_files],
                'original_codebase_dir': str(original_codebase_path) if original_codebase_path else None,
                'original_codebase_files': list(original_codebase_files)
            }
            save_llm_context(sanitized_story, repo_contents, context_files)
        else:
            work_dir = Path(work_dir)
            original_codebase_path = Path(original_codebase_path) if original_codebase_path else None
            original_codebase_files = set(session_state.get('original_codebase_files', []))

        log_progress("🤖 Calling LLM (with agent tools)…")
        turn_start = time.time()

        contents = build_llm_request(sanitized_story, repo_contents, uploaded_refs, conversation_history)

        # Stream the model responses - pass original codebase files for validation
        result = None
        for stream_update in call_gemini_model(client, contents, work_dir, original_codebase_files):
            if stream_update.get("type") == "final":
                result = stream_update
            else:
                update_type = stream_update.get("type", "")
                if update_type in ["thought", "tool_call", "tool_response", "text"]:
                    yield stream_update

        log_progress("🤖 LLM response received ✅")

        if result is None:
            result = {
                "status": "error",
                "created_files": [],
                "summary": "No response received from the model",
                "total_files_created": 0
            }

        sandbox_log_entry = maybe_run_tests_with_sandbox(
            work_dir=work_dir,
            original_codebase_dir=original_codebase_path,
            created_files=result.get("created_files", []),
            log_progress=log_progress
        )
        if sandbox_log_entry:
            result.setdefault("created_files", []).append(sandbox_log_entry)
            result["total_files_created"] = sum(
                1 for f in result.get("created_files", []) if f.get("status") == "success"
            )

        log_progress("☁️ Uploading newly created files to Google Files…")
        new_refs = collect_and_upload_created_files(client, work_dir, since_ts=turn_start)
        if new_refs:
            session_state.setdefault("uploaded_refs", []).extend(new_refs)
        log_progress("☁️ Uploading newly created files to Google Files ✅")

        save_json_file(result, "results", "function_call_result")

        zip_path: Optional[Path] = None
        zip_path: Optional[Path] = None
        conversion_updates: List[Dict[str, Any]] = []
        if create_zip:
            log_progress("📦 Bundling ZIP (original + generated files)…")
            original_dir = session_state.get('original_codebase_dir')
            original_dir_path = Path(original_dir) if original_dir else None
            zip_path = write_outputs_to_zip_from_workdir(
                result=result,
                work_dir=work_dir,
                original_codebase_dir=original_dir_path,
                client=client,
                status_updates=conversion_updates
            )
            log_progress("📦 ZIP bundle ready ✅")

        for event in conversion_updates:
            event_id = event.get("event_id") or f"md2excel_{hashlib.sha1(event.get('file', '').encode('utf-8')).hexdigest()[:8]}"
            display_name = event.get("display_name") or Path(event.get("file", "")).name
            if event.get("event") == "start":
                yield {
                    "type": "tool_call",
                    "content": event.get("message", ""),
                    "metadata": {
                        "title": f"🛠️ {display_name}",
                        "id": event_id,
                        "status": "pending"
                    }
                }
            elif event.get("event") == "success":
                yield {
                    "type": "tool_response",
                    "content": event.get("message", ""),
                    "metadata": {
                        "title": f"✅ {display_name}",
                        "id": event_id,
                        "status": "done"
                    }
                }
            elif event.get("event") == "error":
                yield {
                    "type": "tool_response",
                    "content": f"❌ {event.get('message', '')}",
                    "metadata": {
                        "title": f"❌ {display_name}",
                        "id": event_id,
                        "status": "done"
                    }
                }

        summary = result.get('summary', '')
        success_count = result.get('total_files_created', 0)
        created_files = result.get('created_files', [])

        updated_history = conversation_history + [{
            'user': sanitized_story,
            'assistant_summary': (summary[:500] if summary else f"Created {success_count} files")
        }]

        # Persist session artifacts locally after ZIP creation
        if zip_path:
            saved_dir = persist_session_artifacts(
                session_id=session_id,
                user_story_text=sanitized_story,
                attachments=attachment_paths,
                zip_path=zip_path,
                conversation_history=updated_history,
                tool_events=result.get("tool_responses"),
                session_metadata=session_metadata,
                conversion_updates=conversion_updates,
                summary=summary,
                created_files=created_files,
            )
            if saved_dir:
                log_progress(f"🗂️ Session artifacts saved to {saved_dir}")

        # Yield final result with summary, file list, and zip
        yield {
            "type": "final",
            "message": summary,
            "created_files": created_files,
            "total_files_created": success_count,
            "zip_path": str(zip_path) if zip_path else None,
            "conversation_history": updated_history,
            "session_state": session_state
        }

    except Exception as e:
        logger.error(f"Inference failed: {e}", exc_info=True)
        progress_lines.append(f"❌ Error: {str(e)}")
        yield {
            "type": "final",
            "message": "\n".join(progress_lines) if progress_lines else f"❌ Error: {str(e)}",
            "zip_path": None,
            "conversation_history": conversation_history,
            "session_state": session_state
        }

# ---------- Form-based UI Helpers ----------


def _compose_user_story_input(title: str, description: str) -> str:
    sections = []
    if title:
        sections.append(f"User Story Title:\n{title.strip()}")
    if description:
        sections.append(f"User Story Description:\n{description.strip()}")
    return "\n\n".join(sections).strip()


def _format_created_files(created_files: List[Dict[str, Any]]) -> str:
    rows = []
    for info in created_files or []:
        if info.get("status") != "success":
            continue
        file_path = info.get("file_path")
        if not file_path:
            continue
        name = Path(file_path).name
        desc = info.get("description", "")
        rows.append(f"- **{name}**{f': {desc}' if desc else ''}")
    return "\n".join(rows)


def _persist_streamlit_upload(uploaded_file: Any) -> Optional[Path]:
    """Save a Streamlit UploadedFile to disk for downstream processing."""
    if uploaded_file is None:
        return None
    filename = Path(getattr(uploaded_file, "name", "upload.bin")).name or "upload.bin"
    temp_dir = Path(tempfile.mkdtemp(prefix="streamlit_upload_"))
    target_path = temp_dir / filename
    buffer = uploaded_file.getbuffer()
    target_path.write_bytes(buffer.tobytes() if hasattr(buffer, "tobytes") else bytes(buffer))
    return target_path


def _prepare_zip_download(zip_path: Optional[str]) -> Optional[Tuple[str, bytes]]:
    """Return (filename, bytes) for a generated ZIP so Streamlit can serve it."""
    if not zip_path:
        return None
    candidate = Path(zip_path)
    if not candidate.exists():
        return None
    return candidate.name, candidate.read_bytes()


def _get_all_previous_sessions() -> List[Tuple[str, str, str]]:
    """
    Scan session_logs directory and return list of (session_id, run_timestamp, display_name).
    Returns sessions sorted by timestamp (newest first).
    """
    sessions_list: List[Tuple[str, str, str]] = []
    session_logs_dir = Path("session_logs")
    
    if not session_logs_dir.exists():
        return sessions_list
    
    try:
        for session_folder in sorted(session_logs_dir.iterdir(), reverse=True):
            if not session_folder.is_dir():
                continue
            session_id = session_folder.name
            
            # Find run directories (subdirectories with session.json)
            for run_folder in sorted(session_folder.iterdir(), reverse=True):
                if not run_folder.is_dir():
                    continue
                session_json_path = run_folder / "session.json"
                if session_json_path.exists():
                    run_timestamp = run_folder.name
                    
                    # Try to extract title from session.json
                    title = "Untitled"
                    try:
                        session_data = json.loads(session_json_path.read_text(encoding='utf-8'))
                        title = session_data.get("session_metadata", {}).get("user_story_title", "Untitled")
                        if not title or not title.strip():
                            title = "Untitled"
                    except Exception:
                        pass
                    
                    display_name = f"{title} ({session_id})"
                    sessions_list.append((session_id, run_timestamp, display_name))
    except Exception as e:
        logger.warning(f"Error scanning session_logs: {e}")
    
    return sessions_list


def _load_session_data(session_id: str, run_timestamp: str) -> Optional[Dict[str, Any]]:
    """Load session.json data for a given session and run timestamp."""
    session_json_path = Path("session_logs") / session_id / run_timestamp / "session.json"
    
    if not session_json_path.exists():
        logger.warning(f"Session file not found: {session_json_path}")
        return None
    
    try:
        data = json.loads(session_json_path.read_text(encoding='utf-8'))
        return data
    except Exception as e:
        logger.error(f"Error loading session data: {e}")
        return None


def _restore_session_to_state(session_data: Dict[str, Any]) -> None:
    """Restore a loaded session into Streamlit session state for display."""
    st.session_state["session_id"] = session_data.get("session_id", "")
    st.session_state["current_story_title"] = session_data.get("session_metadata", {}).get("user_story_title", "")
    st.session_state["current_story_description"] = session_data.get("session_metadata", {}).get("user_story_description", "")
    st.session_state["form_submitted"] = True
    st.session_state["reviewed_session_data"] = session_data
    st.session_state["viewing_previous_session"] = True
    st.session_state["session_stopped"] = False
    st.session_state["suppress_session_autoload"] = False
    st.session_state["generation_complete"] = False
    st.session_state["last_generation_snapshot"] = None
    logger.info(f"Restored session: {session_data.get('session_id')}")


def _initialize_session():
    """Initialize or reset session state with a new session ID."""
    session_id = datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')
    st.session_state["session_id"] = session_id
    st.session_state.pop("session_selector", None)  # Reset any previous session selection
    st.session_state["reset_session_selector"] = True
    st.session_state["conversation_history"] = []
    st.session_state["agent_session_state"] = {}
    st.session_state["current_story_title"] = ""
    st.session_state["current_story_description"] = ""
    st.session_state["current_uploaded_file"] = None
    st.session_state["form_submitted"] = False
    st.session_state["viewing_previous_session"] = False
    st.session_state["reviewed_session_data"] = None
    st.session_state["last_loaded_session_display"] = None  # Reset the loaded session tracker
    st.session_state["session_stopped"] = False
    st.session_state["suppress_session_autoload"] = True
    st.session_state["generation_complete"] = False
    st.session_state["last_generation_snapshot"] = None
    logger.info(f"New session initialized: {session_id}")
    return session_id


def _stop_session():
    """Mark the current session as stopped and clear transient state."""
    session_id = st.session_state.get("session_id")
    st.session_state["session_stopped"] = True
    st.session_state["form_submitted"] = False
    st.session_state["viewing_previous_session"] = False
    st.session_state["reviewed_session_data"] = None
    st.session_state["last_loaded_session_display"] = None
    st.session_state["conversation_history"] = []
    st.session_state["agent_session_state"] = {}
    st.session_state["current_story_title"] = ""
    st.session_state["current_story_description"] = ""
    st.session_state["current_uploaded_file"] = None
    st.session_state.pop("session_selector", None)
    st.session_state["reset_session_selector"] = True
    st.session_state["suppress_session_autoload"] = True
    st.session_state["generation_complete"] = False
    st.session_state["last_generation_snapshot"] = None
    logger.info(f"Session stopped: {session_id}")


def render_generator_page():
    # Auto-initialize session on first page load
    if "session_id" not in st.session_state:
        _initialize_session()
    
    suppress_session_autoload = st.session_state.pop("suppress_session_autoload", False)
    
    # Check if form has been submitted
    form_submitted = st.session_state.get("form_submitted", False)
    viewing_previous = st.session_state.get("viewing_previous_session", False)
    
    # Session management header with selector
    col1, col2, col3 = st.columns([2, 1, 1])

    with col1:
        if viewing_previous:
            st.subheader("Historical Session")
        else:
            st.subheader("Provide Your User Story")
    
    with col2:
        # Load previous sessions for selector
        previous_sessions = _get_all_previous_sessions()
        if previous_sessions:
            if st.session_state.pop("reset_session_selector", False):
                st.session_state["session_selector"] = SESSION_SELECTOR_DEFAULT
            # Determine default index: 0 for new sessions, or find current session index if viewing previous
            default_index = 0
            # Only show the selected session if we're actively viewing a previous session
            if st.session_state.get("viewing_previous_session") and st.session_state.get("last_loaded_session_display"):
                current_display = st.session_state["last_loaded_session_display"]
                for idx, (_, _, display) in enumerate(previous_sessions):
                    if display == current_display:
                        default_index = idx + 1  # +1 because of "Select..." option
                        break
            
            session_options = [SESSION_SELECTOR_DEFAULT] + [display for _, _, display in previous_sessions]
            selected_session = st.selectbox(
                "Load Previous Session",
                session_options,
                index=default_index,
                label_visibility="collapsed",
                key="session_selector"
            )
            
            # Only process selection if:
            # 1. It's not the default "Select..." option
            # 2. It's different from the last loaded session
            # 3. Autoload suppression flag is not set (prevents unintended loads after reset)
            last_loaded_session = st.session_state.get("last_loaded_session_display")
            
            # Only load if user actively selected a new session (not already viewing it)
            if (selected_session != SESSION_SELECTOR_DEFAULT and 
                selected_session != last_loaded_session and
                not suppress_session_autoload):
                # Find the corresponding session_id and run_timestamp
                for sess_id, run_ts, display in previous_sessions:
                    if display == selected_session:
                        session_data = _load_session_data(sess_id, run_ts)
                        if session_data:
                            st.session_state["last_loaded_session_display"] = selected_session
                            _restore_session_to_state(session_data)
                            st.rerun()
                        break
            
            if suppress_session_autoload:
                # Allow future manual loads after the initial rerun
                st.session_state["suppress_session_autoload"] = False
    
    with col3:
        if st.button("🔄 New Session", use_container_width=True, help="Clear current session and start fresh"):
            _initialize_session()
            st.rerun()
        if st.button("⏹ Stop Session", use_container_width=True, help="Stop the current session and discard progress"):
            _stop_session()
            st.rerun()

    if st.session_state.get("session_stopped"):
        st.info("This session has been stopped. Start a new session to continue.")
        return
    
    # If viewing a previous session, display it the same way as current session
    if viewing_previous and st.session_state.get("reviewed_session_data"):
        session_data = st.session_state["reviewed_session_data"]
        session_meta = session_data.get("session_metadata", {})
        
        # Don't show the form, skip directly to results display
        st.divider()
        st.subheader("📋 User Story Summary")
        
        summary_col1, summary_col2 = st.columns([2, 1])
        with summary_col1:
            st.markdown("**Title:**")
            st.write(session_meta.get("user_story_title", "N/A"))
            st.markdown("**Description:**")
            st.write(session_meta.get("user_story_description", "N/A"))
        
        with summary_col2:
            st.markdown("**Uploaded Files:**")
            uploaded_files = session_meta.get("uploaded_files", [])
            if uploaded_files:
                for uf in uploaded_files:
                    st.success(f"✓ {uf.get('original_name', 'File')}")
            else:
                st.info("No files uploaded")
        
        st.divider()
        
        status_placeholder = st.empty()
        summary_placeholder = st.empty()
        download_placeholder = st.empty()
        
        # Build status lines from tool events and conversion updates
        status_lines = ["✅ Generation complete (Previous Session)"]
        
        tool_events = session_data.get("tool_events", [])
        if tool_events:
            for event in tool_events:
                event_type = event.get("type", "")
                content = event.get("content", "")
                if content:
                    if event_type == "tool_call":
                        status_lines.append(f"🔧 {content}")
                    elif event_type == "tool_response":
                        status_lines.append(f"✓ {content}")
                    elif event_type == "thought":
                        status_lines.append(f"💭 {content}")
        
        conversion_updates = session_data.get("conversion_updates", [])
        if conversion_updates:
            for update in conversion_updates:
                msg = update.get("message", "")
                if msg:
                    status_lines.append(msg)
        
        status_placeholder.success("\n\n".join(status_lines))
        
        # Display summary and created files
        summary = session_data.get("summary", "")
        created_files = session_data.get("created_files", [])
        
        # Format summary with created files list
        final_summary = summary or "Generation complete. See files below."
        if created_files:
            files_section = _format_created_files(created_files)
            if files_section:
                final_summary = (final_summary + "\n\n### Generated Files\n" + files_section).strip()
        
        summary_placeholder.markdown(f"#### Summary of Test Case Result\n\n{final_summary}")
        
        # Download generated ZIP
        output_zip = session_data.get("output_zip", {})
        if output_zip and output_zip.get("saved_path"):
            zip_path = output_zip.get("saved_path")
            zip_payload = _prepare_zip_download(zip_path)
            
            if zip_payload:
                file_name, file_bytes = zip_payload
                download_placeholder.download_button(
                    "Download Generated ZIP",
                    data=file_bytes,
                    file_name=file_name,
                    mime="application/zip",
                    use_container_width=True,
                )
            else:
                download_placeholder.info("⚠️ Previous session ZIP is no longer available for download.")
        else:
            download_placeholder.info("ℹ️ No output ZIP was generated for this session.")
        
        return
    
    # Show form only if not submitted
    if not form_submitted:
        with st.form("user-story-form", clear_on_submit=True):
            col_left, col_right = st.columns(2)
            with col_left:
                story_title = st.text_input(
                    "User Story Title",
                    placeholder="e.g., Login Flow Happy Path",
                    value=st.session_state.get("current_story_title", "")
                )
            with col_right:
                uploaded_zip = st.file_uploader(
                    "Upload project ZIP or supporting artifacts",
                    type=[ext.lstrip(".") for ext in sorted(ALLOWED_EXTENSIONS)],
                    help="Uses Streamlit's documented st.file_uploader widget to collect optional context files.",
                )
            story_description = st.text_area(
                "User Story Description",
                placeholder="Explain what the feature does, acceptance criteria, and testing focus areas.",
                height=200,
                value=st.session_state.get("current_story_description", "")
            )
            submitted = st.form_submit_button("Generate Test Cases", use_container_width=True)

        if submitted:
            # Save form inputs to session state for persistence
            st.session_state["current_story_title"] = story_title
            st.session_state["current_story_description"] = story_description
            st.session_state["current_uploaded_file"] = uploaded_zip
            st.session_state["form_submitted"] = True
            st.session_state["generation_complete"] = False
            st.session_state["last_generation_snapshot"] = None
            st.rerun()
        else:
            return
    
    # Display summary when form is submitted
    st.divider()
    st.subheader("📋 User Story Summary")
    
    summary_col1, summary_col2 = st.columns([2, 1])
    with summary_col1:
        st.markdown("**Title:**")
        st.write(st.session_state.get("current_story_title", "N/A"))
        st.markdown("**Description:**")
        st.write(st.session_state.get("current_story_description", "N/A"))
    
    with summary_col2:
        st.markdown("**Uploaded Files:**")
        uploaded_file_name = st.session_state.get("current_uploaded_file")
        if uploaded_file_name:
            st.success(f"✓ {getattr(uploaded_file_name, 'name', 'File uploaded')}")
        else:
            st.info("No files uploaded")
    
    st.divider()

    status_placeholder = st.empty()
    summary_placeholder = st.empty()
    download_placeholder = st.empty()

    snapshot = st.session_state.get("last_generation_snapshot")
    if st.session_state.get("generation_complete") and snapshot:
        status_lines = snapshot.get("status_lines", [])
        if status_lines:
            status_placeholder.success("\n\n".join(status_lines))
        summary_markdown = snapshot.get("summary_markdown", "#### Summary of Test Case Result\n\nGeneration complete.")
        summary_placeholder.markdown(summary_markdown)
        download_placeholder.empty()
        zip_path = snapshot.get("zip_path")
        if zip_path:
            zip_payload = _prepare_zip_download(zip_path)
            if zip_payload:
                file_name, file_bytes = zip_payload
                download_placeholder.download_button(
                    "Download Generated ZIP",
                    data=file_bytes,
                    file_name=file_name,
                    mime="application/zip",
                    use_container_width=True,
                )
            else:
                download_placeholder.info("No downloadable bundle was produced for this run.")
        else:
            download_placeholder.info("No downloadable bundle was produced for this run.")
        return

    # Get saved values from session state
    story_title = st.session_state.get("current_story_title", "")
    story_description = st.session_state.get("current_story_description", "")
    uploaded_zip = st.session_state.get("current_uploaded_file")

    story_text = _compose_user_story_input(story_title, story_description)
    session_metadata: Dict[str, Any] = {
        "user_story_title": story_title,
        "user_story_description": story_description,
        "form_story_text": story_text,
        "submitted_at": datetime.now(timezone.utc).isoformat()
    }
    attachments: List[Path] = []
    temp_paths: List[Path] = []
    upload_records: List[Dict[str, Any]] = []

    if uploaded_zip is not None:
        persisted = _persist_streamlit_upload(uploaded_zip)
        if persisted:
            attachments.append(persisted)
            temp_paths.append(persisted)
            upload_records.append({
                "original_name": getattr(uploaded_zip, "name", Path(persisted).name),
                "temp_path": str(persisted),
                "size_bytes": persisted.stat().st_size
            })

    session_metadata["uploaded_files"] = upload_records

    if not story_text and not attachments:
        summary_placeholder.warning("⚠️ Please enter a user story or upload a project ZIP before generating.")
        return

    status_lines = ["🚀 Generating test cases..."]
    status_placeholder.info("\n\n".join(status_lines))
    summary_placeholder.info("Preparing artifacts…")
    download_placeholder.empty()

    conversation_history: List[Dict[str, str]] = st.session_state.get("conversation_history", [])
    agent_session_state: Dict[str, Any] = st.session_state.get("agent_session_state", {})

    try:
        for update in infer(
            story_text,
            attachments,
            True,
            conversation_history,
            agent_session_state,
            session_metadata=session_metadata,
        ):
            update_type = update.get("type")

            if update_type in {"thought", "tool_call", "tool_response"}:
                content = update.get("content", "")
                if content:
                    status_lines.append(content)
                status_placeholder.info("\n\n".join(status_lines))

            elif update_type == "text":
                content = update.get("content", "")
                summary_placeholder.markdown(content or "Processing…")

            elif update_type == "final":
                summary = update.get("message") or update.get("summary") or ""
                files_section = _format_created_files(update.get("created_files", []))
                if files_section:
                    summary = (summary + "\n\n### Generated Files\n" + files_section).strip()
                zip_payload = _prepare_zip_download(update.get("zip_path"))

                status_lines.append("✅ Generation complete")
                status_placeholder.success("\n\n".join(status_lines))

                final_summary = summary or "Generation complete. Download the bundle for details."
                summary_markdown = f"#### Summary of Test Case Result\n\n{final_summary}"
                summary_placeholder.markdown(summary_markdown)

                download_placeholder.empty()
                zip_path = update.get("zip_path")
                if zip_payload:
                    file_name, file_bytes = zip_payload
                    download_placeholder.download_button(
                        "Download Generated ZIP",
                        data=file_bytes,
                        file_name=file_name,
                        mime="application/zip",
                        use_container_width=True,
                    )
                else:
                    download_placeholder.info("No downloadable bundle was produced for this run.")

                st.session_state["conversation_history"] = update.get("conversation_history", conversation_history)
                st.session_state["agent_session_state"] = update.get("session_state", agent_session_state)
                st.session_state["generation_complete"] = True
                st.session_state["last_generation_snapshot"] = {
                    "status_lines": list(status_lines),
                    "summary_markdown": summary_markdown,
                    "zip_path": zip_path,
                }
                return
    except Exception as exc:  # pylint: disable=broad-except
        logger.error("Streamlit handler encountered an error: %s", exc, exc_info=True)
        summary_placeholder.error(f"❌ Error: {exc}")
    finally:
        for temp_path in temp_paths:
            try:
                temp_path.unlink(missing_ok=True)
            except Exception as cleanup_err:  # pylint: disable=broad-except
                logger.warning("Failed to remove temp upload %s: %s", temp_path, cleanup_err)


def main():
    st.set_page_config(page_title="User Story Test Generator", layout="wide")
    apply_peppermint_theme()
    st.title("User Story Test Generator")
    st.caption("LLM-assisted workflow for turning user stories into runnable tests.")

    # Initialize session state
    st.session_state.setdefault("session_id", datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S'))
    st.session_state.setdefault("conversation_history", [])
    st.session_state.setdefault("agent_session_state", {})
    st.session_state.setdefault("current_story_title", "")
    st.session_state.setdefault("current_story_description", "")
    st.session_state.setdefault("current_uploaded_file", None)
    st.session_state.setdefault("form_submitted", False)
    st.session_state.setdefault("viewing_previous_session", False)
    st.session_state.setdefault("reviewed_session_data", None)
    st.session_state.setdefault("generation_complete", False)
    st.session_state.setdefault("last_generation_snapshot", None)
    st.session_state.setdefault("session_stopped", False)

    render_generator_page()


if __name__ == "__main__":
    try:
        main()
    except RuntimeError as runtime_error:
        print(
            f"Runtime error: {runtime_error}. "
            "If you're running this with `python app.py`, try `streamlit run app.py` instead."
        )
