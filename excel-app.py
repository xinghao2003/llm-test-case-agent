# Modular openpyxl helpers: (1) initialize template, (2) append test case records
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path
from typing import List, Union

# ------------------------------
# Core styles (reusable)
# ------------------------------
GREEN = PatternFill("solid", fgColor="FF00B050")
LIGHT_GRAY = PatternFill("solid", fgColor="FFD3D3D3")
WHITE_BOLD = Font(color="FFFFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="FF000000")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADERS = ["Test Case ID", "Test Steps", "Test Input", "Expected Results", "Actual Results", "Status", "Comments"]

def _apply_table_header(ws, start_row:int) -> None:
    for i, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=start_row, column=i, value=h)
        cell.fill = GREEN
        cell.font = WHITE_BOLD
        cell.alignment = CENTER

def _apply_grid(ws, max_row:int, max_col:int=9) -> None:
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER_ALL
            cell.alignment = WRAP if c != 1 else LEFT

def _status_dropdown(ws, start_row:int, end_row:int=500) -> None:
    dv = DataValidation(type="list", formula1='"Pass,Fail,Blocked,N/A"', allow_blank=True)
    dv.add(f"F{start_row}:F{end_row}")
    ws.add_data_validation(dv)

# ---------------------------------------------
# (1) Initialize template
# ---------------------------------------------
def init_testcase_template(
    path: Union[str, Path],
    module_name: str,
    spec_id: str,
    description: str,
    prerequisites: str,
    env_info: str,
    scenario: str,
) -> Path:
    """
    Create a test case workbook with metadata and the steps table.
    Returns the saved file path.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Case Template"

    # Column widths
    for col, w in {"A":18, "B":26, "C":30, "D":30, "E":30, "F":12, "G":18, "H":10, "I":10}.items():
        ws.column_dimensions[col].width = w

    def label(row, text, height=18):
        ws.row_dimensions[row].height = height
        cell = ws.cell(row=row, column=1, value=text)
        cell.fill = GREEN; cell.font = WHITE_BOLD; cell.alignment = LEFT
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)

    # Top metadata
    label(1, "Module Name:-");            ws["B1"].value = module_name
    label(2, "Test Case Spec ID");        ws["B2"].value = spec_id

    ws.merge_cells(start_row=3, start_column=1, end_row=5, end_column=1)
    c = ws.cell(row=3, column=1, value="Test Case Description"); c.fill = GREEN; c.font = WHITE_BOLD; c.alignment = CENTER
    ws.merge_cells(start_row=3, start_column=2, end_row=5, end_column=9)
    ws["B3"].value = description; ws["B3"].alignment = LEFT

    ws.merge_cells(start_row=6, start_column=1, end_row=8, end_column=1)
    c = ws.cell(row=6, column=1, value="Prerequisites:"); c.fill = GREEN; c.font = WHITE_BOLD; c.alignment = CENTER
    ws.merge_cells(start_row=6, start_column=2, end_row=8, end_column=9)
    ws["B6"].value = prerequisites; ws["B6"].alignment = LEFT

    ws.merge_cells(start_row=9, start_column=1, end_row=11, end_column=1)
    c = ws.cell(row=9, column=1, value="Environmental Information:-"); c.fill = GREEN; c.font = WHITE_BOLD; c.alignment = CENTER
    ws.merge_cells(start_row=9, start_column=2, end_row=11, end_column=9)
    ws["B9"].value = env_info; ws["B9"].alignment = LEFT

    ws.merge_cells(start_row=12, start_column=1, end_row=12, end_column=9)
    c = ws.cell(row=12, column=1, value="Test Scenario"); c.fill = GREEN; c.font = WHITE_BOLD; c.alignment = LEFT
    ws.merge_cells(start_row=13, start_column=1, end_row=13, end_column=9)
    ws["A13"].value = scenario; ws["A13"].alignment = LEFT

    # Steps table
    table_header_row = 15
    _apply_table_header(ws, table_header_row)
    _status_dropdown(ws, table_header_row + 1, table_header_row + 500)
    ws.freeze_panes = ws[f"A{table_header_row+1}"]

    # Borders/grid (leave extra space for future rows)
    _apply_grid(ws, table_header_row + 500)

    # Save
    path = Path(path)
    wb.save(path)
    return path

# ---------------------------------------------
# (2) Add a test case row block
# ---------------------------------------------
def add_test_case(
    path: Union[str, Path],
    case_id: str,
    steps: List[str],
    test_input: Union[str, List[str]],
    expected_results: str,
) -> int:
    """
    Appends a multi-row record under the steps table.
    - steps: list[str] of actions (each becomes a new row)
    - test_input: str or list[str] (broadcast if str)
    - expected_results: placed on the FIRST row of the block
    Leaves Actual Results, Status, Comments empty.
    Returns the first row index written.
    """
    wb = load_workbook(path)
    ws = wb.active

    # Find first empty row under table header
    header_row = 15
    row = header_row + 1
    while ws.cell(row=row, column=1).value or ws.cell(row=row, column=2).value:
        row += 1

    # Normalize test_input to list
    if isinstance(test_input, list):
        inputs = test_input + [""] * max(0, len(steps) - len(test_input))
    else:
        inputs = [test_input] + [""] * (len(steps) - 1)

    first_row = row
    for i, step in enumerate(steps):
        ws.cell(row=row, column=1, value=case_id if i == 0 else "")
        ws.cell(row=row, column=2, value=step)
        ws.cell(row=row, column=3, value=inputs[i] if i < len(inputs) else "")
        ws.cell(row=row, column=4, value=expected_results if i == 0 else "")
        # Cols E,F,G left blank by default (Actual, Status, Comments)
        
        # Apply light gray fill to first row of test case
        if i == 0:
            for c in range(1, 8):  # Columns A through G
                ws.cell(row=row, column=c).fill = LIGHT_GRAY
        
        row += 1

    wb.save(path)
    return first_row

# ------------------------------
# Demo: build a file and append two cases
# ------------------------------
out = Path("test_cases_modular.xlsx")
init_testcase_template(
    out,
    module_name="Login",
    spec_id="TC_SPEC_001",
    description="Verify login functionality for valid credentials.",
    prerequisites="1) Stable internet\n2) Test account available\n3) App deployed to staging",
    env_info="OS: Windows/Linux/Mac\nBrowser: Chrome/Firefox/Edge\nSystem: Laptop/Desktop",
    scenario="User should reach dashboard after entering valid username & password.",
)

add_test_case(
    out,
    case_id="TC-001",
    steps=["Enter username", "Enter password", "Click Login"],
    test_input=["user: alice@example.com", "pass: CorrectHorse!23", ""],
    expected_results="Welcome banner appears; Dashboard loads.",
)

add_test_case(
    out,
    case_id="TC-002",
    steps=["Open login page", "Type locked user credentials", "Click Login"],
    test_input="locked user",
    expected_results="Account locked message shown.",
)